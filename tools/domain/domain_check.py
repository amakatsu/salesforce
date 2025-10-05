#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ドメイン定義総合チェックツール（LLM補助）

機能1【ドメイン提案】: 対象一覧のドメイン指定をチェックし、適切なドメインを提案
機能2【整合性チェック】: ドメイン定義とテーブル定義の型・桁数の整合性をチェック
"""
from __future__ import annotations

import argparse
import concurrent.futures as cf
import json
import os
import re
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests

# 共通モジュール
sys.path.insert(0, str(Path(__file__).parent.parent))
from common.api_backend import create_api_backend
from common.config import get_common_config, get_tool_config
from common.excel_utils import read_excel_with_auto_header
from common.normalizers import normalize_text as norm_text, normalize_data_type as norm_dtype
from common.cli_utils import app_root, ask_directory

# domainモジュール
from domain.models import DomainDef, TableDef, TargetItem
from domain.data_loader import load_domains, load_tables, load_targets

# ====== GUI（任意） ===========================================================
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

# ====== 定数 ==================================================================
# 共通設定を取得してツール固有設定を追加
DEFAULT_CONFIG: Dict[str, Any] = {
    **get_common_config(),      # 共通設定（API、LLMパラメータ等）
    **get_tool_config("domain")  # ツール固有設定（config.yamlから取得）
}

# ====== 正規化関数（エイリアス） ==============================================
# Use imported functions directly with shorter names for convenience
normalize_text = norm_text
normalize_data_type = norm_dtype

# ====== LLMプロンプト（機能1: ドメイン提案） ==================================
LLM_SUGGESTION_SYSTEM = (
    "あなたはデータベース設計の専門家です。\n"
    "項目名の業務的意味を分析し、判断材料を提供します（推奨はしません）。\n"
    "\n"
    "重要な分析視点:\n"
    "1. 業務的意味: 項目名が表す業務上の概念を特定\n"
    "2. 技術的項目の判定: 業務的意味を持たない技術的カラムかどうか\n"
    "3. 類似性分析: 既存ドメインとの意味的な類似性を評価\n"
    "4. 特性分析: この項目に必要な特性（バリデーション、制約等）を抽出\n"
    "\n"
    "技術的項目の例:\n"
    "- id, created_at, updated_at, version, seq_no\n"
    "- delete_flag, is_active, enabled\n"
    "- lock_version, row_version\n"
    "\n"
    "注意: 同じデータ型・桁数でも、意味が異なれば別ドメイン（例: 顧客コード vs 商品コード）\n"
)

LLM_SUGGESTION_USER = """対象項目: {item_name}
テーブル: {table_name}.{column_name}
テーブル定義: データ型={table_type}, 桁数={table_length}

既存ドメイン定義一覧:
{domain_candidates}

タスク:
以下の分析を実施し、人間が判断するための材料を提供してください。

1. 業務的意味の分析
2. 技術的項目かどうかの判定
3. 既存ドメインとの類似性分析（意味的な類似性）
4. この項目に必要な特性の抽出

出力JSON:
{{
  "business_meaning": string,
  "is_technical": boolean,
  "similar_domains": [string],
  "required_characteristics": string,
  "notes": string
}}

フィールド説明:
- business_meaning: 項目の業務的意味を簡潔に説明（例: "顧客を一意に識別するコード"）
  技術的項目の場合も説明を記載（例: "レコードの作成日時を保持する技術的カラム"）
- is_technical: 技術的項目=true、業務的項目=false
- similar_domains: 既存ドメインの中で意味的に類似するものをリスト（最大3件）
  完全一致がなくても、関連性があるものを含める
  例: ["顧客コード", "取引先コード"]
- required_characteristics: この項目に必要と思われる特性（バリデーション、制約等）
  例: "英数字8桁固定、ユニーク制約"
- notes: その他の補足情報や注意点

JSON以外出力禁止。
"""

# ====== LLMプロンプト（機能2: 整合性チェック） ================================
LLM_VALIDATION_SYSTEM = (
    "あなたはデータベース設計の専門家です。\n"
    "ドメイン定義とテーブル定義の差異を分析し、重要度を評価します。\n"
)

LLM_VALIDATION_USER = """テーブル: {table_name}.{column_name}
指定ドメイン: {domain_name}

ドメイン定義: 型={domain_type}, 桁数={domain_length}, 精度={domain_precision}, スケール={domain_scale}
テーブル定義: 型={table_type}, 桁数={table_length}, 精度={table_precision}, スケール={table_scale}

差異: {differences}

出力JSON:
{{
  "severity": "critical" | "warning" | "info" | "acceptable",
  "is_valid": true | false,
  "reason": string,
  "recommendation": string
}}

重要度:
- critical: データ損失の危険
- warning: 業務上問題の可能性
- info: 統一推奨だが問題なし
- acceptable: 実質的に同じ

JSON以外出力禁止。
"""

def call_llm(prompt_system: str, prompt_user: str, cfg: Dict[str, Any],
             client, api_semaphore: threading.Semaphore) -> Dict[str, Any]:
    payload = {
        "model": cfg["OPENAI_MODEL"],
        "max_tokens": cfg["MAX_TOKENS"],
        "temperature": cfg["TEMPERATURE"],
        "top_p": cfg["TOP_P"],
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": prompt_system},
            {"role": "user", "content": prompt_user}
        ]
    }
    if api_semaphore:
        api_semaphore.acquire()
    try:
        for attempt in range(cfg["RETRY"] + 1):
            try:
                data = client.post_json(payload)
                content = data["choices"][0]["message"]["content"]
                return json.loads(content)
            except Exception as e:
                if attempt < cfg["RETRY"]:
                    time.sleep(1.2 * (attempt + 1))
                    continue
                return {"error": str(e)}
    finally:
        if api_semaphore:
            api_semaphore.release()

# ====== 処理関数 ==============================================================
def compare_spec(domain: DomainDef, table: TableDef) -> Tuple[bool, List[str]]:
    """ドメイン定義とテーブル定義を比較"""
    type_match = normalize_data_type(domain.data_type) == normalize_data_type(table.data_type)
    diffs = []
    if not type_match:
        diffs.append(f"データ型: {domain.data_type} vs {table.data_type}")

    # 桁数比較
    if domain.length and table.length:
        if normalize_text(domain.length) != normalize_text(table.length):
            diffs.append(f"桁数: {domain.length} vs {table.length}")

    # 精度・スケール比較
    if domain.precision and table.precision:
        if normalize_text(domain.precision) != normalize_text(table.precision):
            diffs.append(f"精度: {domain.precision} vs {table.precision}")
    if domain.scale and table.scale:
        if normalize_text(domain.scale) != normalize_text(table.scale):
            diffs.append(f"スケール: {domain.scale} vs {table.scale}")

    return (type_match and len(diffs) == 0), diffs


def analyze_compatibility(table_def: TableDef, domains: Dict[str, DomainDef]) -> Tuple[List[str], str]:
    """テーブル定義と各ドメインの整合性を分析（提案時の参考情報）"""
    domain_list = []
    analysis_notes = []

    table_dtype_norm = normalize_data_type(table_def.data_type)

    # データ型一致のドメインを優先、その後は全ドメイン
    type_matched = []
    type_mismatched = []

    for domain_name, domain in domains.items():
        domain_dtype_norm = normalize_data_type(domain.data_type)
        is_match, diffs = compare_spec(domain, table_def)

        domain_info = f"{domain.name} ({domain.data_type}"
        if domain.length:
            domain_info += f", {domain.length}桁"
        domain_info += ")"

        if domain_dtype_norm == table_dtype_norm:
            # データ型一致
            if is_match:
                # 完全一致
                domain_info += " - 完全一致 ✓"
                analysis_notes.append(f"✓ {domain.name}: データ型・桁数完全一致")
                type_matched.insert(0, domain_info)  # 先頭に追加
            elif len(diffs) == 1 and "桁数" in diffs[0]:
                # 型一致、桁数のみ差異
                domain_info += " - 型一致、桁数差異"
                analysis_notes.append(f"△ {domain.name}: 型一致、{diffs[0]}")
                type_matched.append(domain_info)
            else:
                # 型一致、その他差異
                domain_info += " - 型一致"
                type_matched.append(domain_info)
        else:
            # データ型不一致（参考として含める）
            domain_info += " - 型不一致"
            type_mismatched.append(domain_info)

    # データ型一致を優先、その後全ドメイン
    domain_list = type_matched + type_mismatched

    # 上位候補のみ返却（LLMへの入力を減らす）
    top_domains = domain_list[:30]  # 上位30件（意味重視なので多めに）
    analysis_text = "\n".join(analysis_notes[:15]) if analysis_notes else "データ型が一致するドメインがありません（新規定義を推奨）"

    return (top_domains, analysis_text)


def process_domain_suggestion(targets: List[TargetItem], domains: Dict[str, DomainDef],
                              tables: Dict[str, TableDef],
                              cfg: Dict[str, Any], client: ApiClient,
                              api_semaphore: threading.Semaphore) -> pd.DataFrame:
    """
    ドメイン提案機能（新仕様）

    処理フロー:
    1. 対象一覧から項目名を読み込み
    2. ドメイン定義で完全一致チェック → 一致あれば採用して終了
    3. テーブル定義から該当項目を検索
    4. LLMで判断材料を提供（提案はしない）
    5. 結果を出力（人間が最終判断）
    """
    rows = []
    for item in targets:
        # Step 1: 項目名で正規化キーを作成
        item_key = normalize_text(item.item_name)

        # Step 2: ドメイン定義で完全一致チェック
        if item_key in domains:
            matched_domain = domains[item_key]
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "判定結果": "完全一致",
                "一致ドメイン": matched_domain.name,
                "ドメイン行番号": matched_domain.row_number,
                "ドメインファイル": matched_domain.source_file,
                "データ型": matched_domain.data_type,
                "桁数": matched_domain.length or "-",
                "バリデーション": matched_domain.validation,
                "備考": "ドメイン定義に完全一致"
            })
            continue

        # Step 3: テーブル定義から該当項目を検索
        table_def = tables.get(item_key)
        if not table_def:
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "判定結果": "エラー",
                "備考": "テーブル定義が見つかりません"
            })
            continue

        # Step 4: LLMで判断材料を提供
        # すべてのドメインを候補として渡す（型でフィルタしない）
        domain_list = []
        for domain in domains.values():
            domain_list.append(
                f"- {domain.name}: {domain.data_type}({domain.length or '-'}), バリデーション: {domain.validation}"
            )

        llm_result = call_llm(
            LLM_SUGGESTION_SYSTEM,
            LLM_SUGGESTION_USER.format(
                item_name=item.item_name,
                table_name=table_def.table_name,
                column_name=table_def.column_name,
                table_type=table_def.data_type,
                table_length=table_def.length or "未指定",
                domain_candidates="\n".join(domain_list) if domain_list else "(ドメイン候補なし)"
            ),
            cfg, client, api_semaphore
        )

        if "error" in llm_result:
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "テーブル名": table_def.table_name,
                "カラム名": table_def.column_name,
                "テーブル行番号": table_def.row_number,
                "データ型": table_def.data_type,
                "桁数": table_def.length or "-",
                "判定結果": "LLMエラー",
                "備考": llm_result.get("error", "不明なエラー")
            })
        else:
            # Step 5: LLM分析結果を出力（人間が最終判断）
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "テーブル名": table_def.table_name,
                "カラム名": table_def.column_name,
                "テーブル行番号": table_def.row_number,
                "テーブルファイル": table_def.source_file,
                "データ型": table_def.data_type,
                "桁数": table_def.length or "-",
                "判定結果": "要判断",
                "業務的意味": llm_result.get("business_meaning", ""),
                "類似ドメイン": ", ".join(llm_result.get("similar_domains", [])),
                "必要な特性": llm_result.get("required_characteristics", ""),
                "技術的項目判定": "はい" if llm_result.get("is_technical", False) else "いいえ",
                "備考": llm_result.get("notes", "")
            })

    return pd.DataFrame(rows)

def process_domain_validation(tables: Dict[Tuple[str, str], TableDef],
                              domains: Dict[str, DomainDef],
                              cfg: Dict[str, Any], client: ApiClient,
                              api_semaphore: threading.Semaphore) -> pd.DataFrame:
    """機能2: 整合性チェック"""
    rows = []
    for (table_key), table_def in tables.items():
        if not table_def.domain:
            continue

        domain_key = normalize_text(table_def.domain)
        domain = domains.get(domain_key)

        if not domain:
            rows.append({
                "table_name": table_def.table_name,
                "column_name": table_def.column_name,
                "domain_name": table_def.domain,
                "check_result": "エラー",
                "severity": "critical",
                "reason": "ドメイン定義が存在しません",
                "recommendation": "ドメイン定義を追加"
            })
            continue

        is_match, diffs = compare_spec(domain, table_def)

        if is_match:
            rows.append({
                "table_name": table_def.table_name,
                "column_name": table_def.column_name,
                "domain_name": domain.name,
                "check_result": "OK",
                "severity": "acceptable",
                "domain_type": domain.data_type,
                "table_type": table_def.data_type,
                "reason": "完全一致",
                "recommendation": "問題なし"
            })
        else:
            # LLM判定
            llm_result = call_llm(
                LLM_VALIDATION_SYSTEM,
                LLM_VALIDATION_USER.format(
                    table_name=table_def.table_name,
                    column_name=table_def.column_name,
                    domain_name=domain.name,
                    domain_type=domain.data_type,
                    domain_length=domain.length or "未指定",
                    domain_precision=domain.precision or "未指定",
                    domain_scale=domain.scale or "未指定",
                    table_type=table_def.data_type,
                    table_length=table_def.length or "未指定",
                    table_precision=table_def.precision or "未指定",
                    table_scale=table_def.scale or "未指定",
                    differences="\n".join(diffs)
                ),
                cfg, client, api_semaphore
            )
            if "error" not in llm_result:
                rows.append({
                    "table_name": table_def.table_name,
                    "column_name": table_def.column_name,
                    "domain_name": domain.name,
                    "check_result": "OK" if llm_result.get("is_valid") else "不一致",
                    "severity": llm_result.get("severity", "warning"),
                    "domain_type": domain.data_type,
                    "table_type": table_def.data_type,
                    "reason": llm_result.get("reason", ""),
                    "recommendation": llm_result.get("recommendation", "")
                })
    return pd.DataFrame(rows)

# ====== 出力 ==================================================================
def save_outputs(df_suggestion: Optional[pd.DataFrame], df_validation: Optional[pd.DataFrame],
                cfg: Dict[str, Any]) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule

    out_dir = Path(cfg["OUT_DIR"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "domain_check_result.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        if df_suggestion is not None and len(df_suggestion) > 0:
            df_suggestion.to_excel(w, sheet_name="ドメイン提案", index=False)
            ws = w.sheets["ドメイン提案"]
            # 色付け（判定結果列）
            if "判定結果" in df_suggestion.columns:
                result_col_idx = list(df_suggestion.columns).index("判定結果") + 1
                result_col = get_column_letter(result_col_idx)
                fill_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                rng = f"{result_col}2:{result_col}{len(df_suggestion)+1}"
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{result_col}2="完全一致"'], fill=fill_green))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{result_col}2="要判断"'], fill=fill_yellow))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{result_col}2="エラー"'], fill=fill_red))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{result_col}2="LLMエラー"'], fill=fill_red))

        if df_validation is not None and len(df_validation) > 0:
            df_validation.to_excel(w, sheet_name="整合性チェック", index=False)
            ws = w.sheets["整合性チェック"]
            # 色付け
            if "severity" in df_validation.columns:
                sev_col_idx = list(df_validation.columns).index("severity") + 1
                sev_col = get_column_letter(sev_col_idx)
                fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                fill_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                rng = f"{sev_col}2:{sev_col}{len(df_validation)+1}"
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{sev_col}2="critical"'], fill=fill_red))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{sev_col}2="warning"'], fill=fill_yellow))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{sev_col}2="acceptable"'], fill=fill_green))

    print(f"保存: {xlsx_path}")

# ====== CLI ===================================================================
def main() -> None:
    """
    メイン処理（CLI実行用）
    ※ 設定はconfig.yamlから読み込まれます（環境変数で上書き可）
    """
    parser = argparse.ArgumentParser(description="ドメイン定義総合チェックツール")
    parser.add_argument("--dir", help="入力ディレクトリ")
    parser.add_argument("--out-dir", help="出力ディレクトリ")
    parser.add_argument("--mode", choices=["both", "suggestion", "validation"], default="both",
                       help="実行モード: both=両方, suggestion=ドメイン提案のみ, validation=整合性チェックのみ")
    parser.add_argument("--no-gui", action="store_true")
    args = parser.parse_args()

    cfg = DEFAULT_CONFIG.copy()
    cfg["CHECK_MODE"] = args.mode

    in_dir = args.dir
    if not in_dir and not args.no_gui:
        in_dir = ask_directory("入力ディレクトリを選択")
    if not in_dir:
        print("入力ディレクトリが指定されていません")
        return

    if args.out_dir:
        cfg["OUT_DIR"] = args.out_dir

    root_dir = Path(in_dir)
    if not root_dir.exists():
        print(f"ディレクトリが存在しません: {root_dir}")
        return

    # データ読み込み
    domains = load_domains(root_dir, cfg)
    tables = load_tables(root_dir, cfg)

    api_client = create_api_backend(cfg)
    api_semaphore = threading.Semaphore(cfg["MAX_CONCURRENT_API"])

    df_suggestion = None
    df_validation = None

    # 機能1: ドメイン提案
    if cfg["CHECK_MODE"] in ["both", "suggestion"]:
        targets = load_targets(root_dir, cfg)
        if targets:
            print("\n[機能1] ドメイン提案処理開始...")
            df_suggestion = process_domain_suggestion(targets, domains, tables, cfg, api_client, api_semaphore)

    # 機能2: 整合性チェック
    if cfg["CHECK_MODE"] in ["both", "validation"]:
        print("\n[機能2] 整合性チェック処理開始...")
        df_validation = process_domain_validation(tables, domains, cfg, api_client, api_semaphore)

    save_outputs(df_suggestion, df_validation, cfg)

    if TK_AVAILABLE and not args.no_gui:
        try:
            messagebox.showinfo("完了", f"処理完了\n保存先: {Path(cfg['OUT_DIR']).resolve()}")
        except:
            pass

if __name__ == "__main__":
    main()
