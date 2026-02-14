#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ドメイン定義チェックツール

画面項目定義・テーブル定義とドメイン定義一覧を照合し、
一致状況を判定する。
"""
from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd

try:
    from ..config import get_domain_config
except ImportError:
    from config import get_domain_config

try:
    from .domain_matcher import collect_evidence, resolve_without_llm, digit_match_rate_from_details
    from .llm_matcher import run_llm_matching, build_llm_call, group_new_domain_proposals
except ImportError:
    collect_evidence = None
    resolve_without_llm = None
    digit_match_rate_from_details = None
    run_llm_matching = None
    build_llm_call = None
    group_new_domain_proposals = None


# ====== 正規化 ================================================================

def normalize_text(text: str) -> str:
    """NFKC正規化 + 小文字化 + 同義語統一"""
    if text is None or text == "":
        return ""
    s = unicodedata.normalize("NFKC", str(text)).lower().strip()
    s = re.sub(r"[\u3000\s]+", " ", s)

    # 日付・時刻
    s = re.sub(r"取引日$", "日付", s)
    s = re.sub(r"年月日$", "日付", s)
    s = re.sub(r"(\w+)日$", r"\1日付", s)
    s = re.sub(r"日時$", "日時", s)
    s = re.sub(r"時刻$", "時刻", s)
    s = re.sub(r"タイムスタンプ$", "日時", s)

    # 金額・数量
    s = re.sub(r"金額$", "金額", s)
    s = re.sub(r"価格$", "金額", s)
    s = re.sub(r"料金$", "金額", s)
    s = re.sub(r"単価$", "単価", s)
    s = re.sub(r"数量$", "数量", s)
    s = re.sub(r"件数$", "数量", s)
    s = re.sub(r"個数$", "数量", s)

    # コード・ID
    s = re.sub(r"コード$", "コード", s)
    s = re.sub(r"cd$", "コード", s)
    s = re.sub(r"識別子$", "id", s)
    s = re.sub(r"番号$", "番号", s)
    s = re.sub(r"no$", "番号", s)

    # 名称
    s = re.sub(r"名称$", "名称", s)
    s = re.sub(r"名前$", "名称", s)
    s = re.sub(r"氏名$", "名称", s)
    s = re.sub(r"name$", "名称", s)

    # 区分・種別
    s = re.sub(r"区分$", "区分", s)
    s = re.sub(r"種別$", "区分", s)
    s = re.sub(r"タイプ$", "区分", s)
    s = re.sub(r"type$", "区分", s)

    # フラグ・状態
    s = re.sub(r"フラグ$", "フラグ", s)
    s = re.sub(r"flag$", "フラグ", s)
    s = re.sub(r"状態$", "状態", s)
    s = re.sub(r"ステータス$", "状態", s)
    s = re.sub(r"status$", "状態", s)

    # 備考・メモ
    s = re.sub(r"備考$", "備考", s)
    s = re.sub(r"メモ$", "備考", s)
    s = re.sub(r"コメント$", "備考", s)
    s = re.sub(r"摘要$", "備考", s)

    return s


def normalize_data_type(dtype: str) -> str:
    """データ型の正規化"""
    if not dtype:
        return ""
    dt = normalize_text(dtype)
    type_map = {
        "varchar": "varchar", "varchar2": "varchar", "char": "char",
        "int": "integer", "integer": "integer", "bigint": "bigint",
        "decimal": "decimal", "numeric": "decimal", "number": "decimal",
        "date": "date", "datetime": "datetime", "timestamp": "timestamp",
        "boolean": "boolean", "bool": "boolean",
    }
    for key, value in type_map.items():
        if key in dt:
            return value
    return dt


# ====== データクラス ==========================================================

@dataclass
class ScreenItem:
    """画面項目定義（確定仕様: 10列）"""
    item_name: str             # 項目名
    data_type: str             # データ型
    text_type: str             # テキストタイプ
    min_digits: Optional[str]  # 最小桁
    max_digits: Optional[str]  # 最大桁
    max_bytes: Optional[str]   # 最大バイト数
    min_value: Optional[str]   # 最小値
    max_value: Optional[str]   # 最大値
    code_id: str               # 外部コード
    row_number: int
    source_file: str
    source_sheet: str


@dataclass
class TableItem:
    """テーブル定義（確定仕様: 6列）"""
    item_name: str             # 項目名
    data_type: str             # データ型
    length: Optional[str]      # length
    integer_part: Optional[str]  # 数値の整数
    decimal_part: Optional[str]  # 小数点
    row_number: int
    source_file: str
    source_sheet: str


@dataclass
class DomainDef:
    """ドメイン定義（殿の確定仕様: 12列）"""
    name: str                      # ドメイン名
    data_type: str                 # データ型
    min_char: Optional[str]        # 最小文字数
    max_char: Optional[str]        # 最大文字数
    min_byte: Optional[str]        # 最小バイト長
    max_byte: Optional[str]        # 最大バイト長
    integer_digits: Optional[str]  # 整数部桁数
    decimal_digits: Optional[str]  # 小数部桁数
    min_value: Optional[str]       # 最小値
    max_value: Optional[str]       # 最大値
    regex: Optional[str]           # 書式（正規表現）
    code_id: str                   # 参照外部コード
    row_number: int
    source_file: str
    source_sheet: str


# ====== Excel I/O =============================================================

HEADER_DETECT = os.getenv("HEADER_DETECT", "true").lower() != "false"
HEADER_SCAN_ROWS = int(os.getenv("HEADER_SCAN_ROWS", "15"))


def _pick_matching_sheets(xls: pd.ExcelFile, preferred: Optional[str]) -> List[str]:
    if not preferred:
        return [xls.sheet_names[0]]
    norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
    patterns = [p.strip() for p in preferred.split(",") if p.strip()]
    if not patterns:
        return [xls.sheet_names[0]]
    all_matches = []
    for pattern in patterns:
        want = norm(pattern)
        regex_pattern = want.replace("*", ".*")
        for name in xls.sheet_names:
            if re.match(regex_pattern, norm(name)):
                all_matches.append(name)
    seen: set[str] = set()
    result = []
    for sheet in all_matches:
        if sheet not in seen:
            result.append(sheet)
            seen.add(sheet)
    return result if result else [xls.sheet_names[0]]


def _domain_column_names(cfg: Dict[str, Any]) -> Dict[str, str]:
    """ドメイン定義シートの列名を返す。"""
    return {
        "name":           cfg.get("DOMAIN_NAME_COL", "ドメイン名"),
        "data_type":      cfg.get("DOMAIN_TYPE_COL", "データ型"),
        "min_char":       cfg.get("DOMAIN_STR_MIN_CHARS_COL", "最小文字数"),
        "max_char":       cfg.get("DOMAIN_STR_MAX_CHARS_COL", "最大文字数"),
        "min_byte":       cfg.get("DOMAIN_BYTES_MIN_COL", "最小バイト長"),
        "max_byte":       cfg.get("DOMAIN_BYTES_MAX_COL", "最大バイト長"),
        "integer_digits": cfg.get("DOMAIN_INT_DIGITS_COL", "整数部桁数"),
        "decimal_digits": cfg.get("DOMAIN_DEC_DIGITS_COL", "小数部桁数"),
        "min_value":      cfg.get("DOMAIN_NUM_MIN_COL", "最小値"),
        "max_value":      cfg.get("DOMAIN_NUM_MAX_COL", "最大値"),
        "regex":          cfg.get("DOMAIN_REGEX_COL", "書式（正規表現）"),
        "code_id":        cfg.get("DOMAIN_EXT_CODE_COL", "参照外部コード"),
    }


def _detect_header_row(
    path: Path, sheet_name: str, required_cols: List[str], scan_rows: int,
) -> int:
    head_df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows)
    req = [normalize_text(c) for c in required_cols if c]
    for i in range(len(head_df)):
        row_vals = [
            normalize_text(x) for x in head_df.iloc[i].values if str(x) not in {"", "nan"}
        ]
        if all(any(r in v for v in row_vals) for r in req):
            return i
    raise KeyError(
        f"必須列{sorted(required_cols)}を含むヘッダ行が見つかりません: {path.name}/{sheet_name} (scan_rows={scan_rows})"
    )


def read_excel_with_header_detection(
    path: Path,
    sheet_name: Optional[str],
    required_cols: List[str],
    explicit_header_row_1based: Optional[int] = None,
    scan_rows: int = 30,
) -> Tuple[pd.DataFrame, int]:
    """ヘッダ行が1行目とは限らないExcelに対応。"""
    if explicit_header_row_1based is not None:
        hdr0 = max(0, explicit_header_row_1based - 1)
        return pd.read_excel(path, sheet_name=sheet_name, header=hdr0), hdr0
    if not HEADER_DETECT:
        return pd.read_excel(path, sheet_name=sheet_name), 0
    header_row = _detect_header_row(path, sheet_name, required_cols, scan_rows)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row), header_row


# ====== セル値取得ユーティリティ ==============================================

_EMPTY_VALUES = {"", "nan", "None", " ", "\u3000"}


def _get_cell_value(
    row: pd.Series, df: pd.DataFrame, col_name: str, col_idx: int = -1,
) -> str:
    """列名で値を取得し、見つからなければ列位置でフォールバック。

    空値・NaN・"nan"等は空文字列として返す。
    """
    val = None
    if col_name in df.columns:
        val = row.get(col_name, "")
    elif 0 <= col_idx < len(row):
        try:
            val = row.iloc[col_idx]
        except Exception:
            val = ""

    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    val_str = str(val).strip()
    if val_str in _EMPTY_VALUES or val_str.isspace():
        return ""
    return val_str


def _find_columns(df: pd.DataFrame, column_names: Dict[str, str]) -> Dict[str, int]:
    """ヘッダー行から列名→列インデックスの対応を動的に取得する（部分一致）。

    列名が「項目名称（和名）」のように装飾されていても、
    検索キー「項目名称」でヒットする。完全一致を優先し、
    なければ部分一致にフォールバックする。

    Args:
        df: ヘッダー検出済みのDataFrame（df.columnsがヘッダー行）
        column_names: {内部キー: Excel列名} の辞書

    Returns:
        {内部キー: 列インデックス} の辞書（見つかった列のみ）
    """
    headers = [str(h).strip() for h in df.columns]
    col_map: Dict[str, int] = {}
    for key, col_name in column_names.items():
        target = col_name.strip()
        # 完全一致を優先
        exact = [i for i, h in enumerate(headers) if h == target]
        if exact:
            col_map[key] = exact[0]
            continue
        # 部分一致にフォールバック
        partial = [i for i, h in enumerate(headers) if target in h]
        if partial:
            col_map[key] = partial[0]
    return col_map


def _is_domain_unnecessary(row: pd.Series, col_map: Dict[str, int]) -> bool:
    """項目名称だけあって他が全て空 → ドメイン不要項目として除外対象。

    テキストタイプ〜外部コードの7列が全て空の場合にTrueを返す。
    """
    check_keys = [
        "text_type", "min_digits", "max_digits",
        "max_bytes", "min_value", "max_value", "external_code",
    ]
    for key in check_keys:
        if key not in col_map:
            continue
        val = row.iloc[col_map[key]]
        if not (pd.isna(val) or str(val).strip() in _EMPTY_VALUES):
            return False
    return True


# ====== データ読み込み ========================================================

def load_screen_items(dir_path: Path, cfg: Dict[str, Any]) -> List[ScreenItem]:
    """画面項目定義を読み込み（ヘッダー文字列検索方式・行番号付き）。

    殿の確定仕様:
    - 列はヘッダー行から文字列で動的検索（列番号固定禁止）
    - 正式列名: 項目名称, 型, テキストタイプ, 最小桁, 最大桁,
                最大バイト数, 最小値, 最大値, 外部コード
    - 項目名称だけあって他7列が全空の行は「ドメイン不要」（対象外）
      → 読み込みは行うが、process関数でマッチングスキップ＋「対象外」出力
    """
    # 列名マッピング: 内部キー → Excelヘッダーの正式列名
    screen_col_names: Dict[str, str] = {
        "item_name":     cfg.get("SCREEN_ITEM_COL", "項目名称"),
        "data_type":     cfg.get("SCREEN_TYPE_COL", "型"),
        "text_type":     cfg.get("SCREEN_TEXT_TYPE_COL", "テキストタイプ"),
        "min_digits":    cfg.get("SCREEN_MIN_DIGITS_COL", "最小桁"),
        "max_digits":    cfg.get("SCREEN_MAX_DIGITS_COL", "最大桁"),
        "max_bytes":     cfg.get("SCREEN_MAX_BYTES_COL", "最大バイト数"),
        "min_value":     cfg.get("SCREEN_MIN_VALUE_COL", "最小値"),
        "max_value":     cfg.get("SCREEN_MAX_VALUE_COL", "最大値"),
        "external_code": cfg.get("SCREEN_EXT_CODE_COL", "外部コード"),
    }

    files = sorted(dir_path.glob(cfg["SCREEN_GLOB"]))
    if not files:
        raise FileNotFoundError("画面項目定義ファイルが見つかりません")
    items: List[ScreenItem] = []
    for path in files:
        xls = pd.ExcelFile(path)
        for sheet in _pick_matching_sheets(xls, cfg["SCREEN_SHEET"]):
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet, [screen_col_names["item_name"]], None, 30,
                )
                col_map = _find_columns(df, screen_col_names)
                if "item_name" not in col_map:
                    print(f"[警告] {path.name}({sheet}): 「{screen_col_names['item_name']}」列が見つかりません")
                    continue

                def _val(key: str) -> str:
                    if key not in col_map:
                        return ""
                    v = row.iloc[col_map[key]]
                    if pd.isna(v):
                        return ""
                    s = str(v).strip()
                    return "" if s in _EMPTY_VALUES else s

                for idx, row in df.iterrows():
                    if row.isna().all():
                        continue
                    item_name = _val("item_name")
                    if not item_name:
                        continue

                    items.append(ScreenItem(
                        item_name=item_name,
                        data_type=_val("data_type"),
                        text_type=_val("text_type"),
                        min_digits=_val("min_digits") or None,
                        max_digits=_val("max_digits") or None,
                        max_bytes=_val("max_bytes") or None,
                        min_value=_val("min_value") or None,
                        max_value=_val("max_value") or None,
                        code_id=_val("external_code"),
                        row_number=header_row + idx + 2,
                        source_file=path.name,
                        source_sheet=sheet,
                    ))
                print(f"[INFO] 画面項目定義読み込み: {path.name}/{sheet}")
            except KeyError as e:
                print(f"[警告] {path.name}({sheet}): {e} -> スキップ")
                continue
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計画面項目: {len(items)}件")
    return items


def load_domains(dir_path: Path, cfg: Dict[str, Any]) -> Dict[str, DomainDef]:
    """ドメイン定義を読み込み（ヘッダー文字列検索方式・正規化名でキー・行番号付き）。

    殿の確定仕様（12列）:
    - ドメイン名, データ型, 最小文字数, 最大文字数, 最小バイト長, 最大バイト長,
      整数部桁数, 小数部桁数, 最小値, 最大値, 書式（正規表現）, 参照外部コード
    """
    domain_col_names = _domain_column_names(cfg)

    files = sorted(dir_path.glob(cfg["DOMAIN_GLOB"]))
    if not files:
        raise FileNotFoundError("ドメイン定義ファイルが見つかりません")
    domains: Dict[str, DomainDef] = {}
    for path in files:
        xls = pd.ExcelFile(path)
        for sheet in _pick_matching_sheets(xls, cfg["DOMAIN_SHEET"]):
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet, [domain_col_names["name"]], None, 30,
                )
                col_map = _find_columns(df, domain_col_names)
                if "name" not in col_map:
                    print(f"[警告] {path.name}({sheet}): 「{domain_col_names['name']}」列が見つかりません")
                    continue

                def _val(key: str) -> str:
                    if key not in col_map:
                        return ""
                    v = row.iloc[col_map[key]]
                    if pd.isna(v):
                        return ""
                    s = str(v).strip()
                    return "" if s in _EMPTY_VALUES else s

                for idx, row in df.iterrows():
                    if row.isna().all():
                        continue
                    name = _val("name")
                    if not name:
                        continue

                    domains[normalize_text(name)] = DomainDef(
                        name=name,
                        data_type=_val("data_type"),
                        min_char=_val("min_char") or None,
                        max_char=_val("max_char") or None,
                        min_byte=_val("min_byte") or None,
                        max_byte=_val("max_byte") or None,
                        integer_digits=_val("integer_digits") or None,
                        decimal_digits=_val("decimal_digits") or None,
                        min_value=_val("min_value") or None,
                        max_value=_val("max_value") or None,
                        regex=_val("regex") or None,
                        code_id=_val("code_id"),
                        row_number=header_row + idx + 2,
                        source_file=path.name,
                        source_sheet=sheet,
                    )
                print(f"[INFO] ドメイン定義読み込み: {path.name}/{sheet} ({len(domains)}件)")
            except KeyError as e:
                print(f"[警告] {path.name}({sheet}): {e} -> スキップ")
                continue
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計ドメイン定義: {len(domains)}件")
    return domains


def load_table_definitions(dir_path: Path, cfg: Dict[str, Any]) -> List[TableItem]:
    """テーブル定義を読み込み（ヘッダー文字列検索方式・行番号付き）。

    殿の確定仕様:
    - 列はヘッダー行から文字列で動的検索（列番号固定禁止）
    - 正式列名: 論理項目名, データ型, Length, 全体数値, 少数桁
    - 論理項目名だけあって他4列が全空の行は「対象外」
      → 読み込みは行うが、process関数でマッチングスキップ＋「対象外」出力
    """
    # 列名マッピング: 内部キー → Excelヘッダーの正式列名
    table_col_names: Dict[str, str] = {
        "item_name":     cfg.get("TABLE_ITEM_COL", "論理項目名"),
        "data_type":     cfg.get("TABLE_TYPE_COL", "データ型"),
        "length":        cfg.get("TABLE_LENGTH_COL", "Length"),
        "integer_part":  cfg.get("TABLE_INT_COL", "全体数値"),
        "decimal_part":  cfg.get("TABLE_DEC_COL", "少数桁"),
    }

    files = sorted(dir_path.glob(cfg["TABLE_GLOB"]))
    if not files:
        raise FileNotFoundError("テーブル定義ファイルが見つかりません")
    items: List[TableItem] = []
    for path in files:
        xls = pd.ExcelFile(path)
        for sheet in _pick_matching_sheets(xls, cfg["TABLE_SHEET"]):
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet, [table_col_names["item_name"]], None, 30,
                )
                col_map = _find_columns(df, table_col_names)
                if "item_name" not in col_map:
                    print(f"[警告] {path.name}({sheet}): 「{table_col_names['item_name']}」列が見つかりません")
                    continue

                def _val(key: str) -> str:
                    if key not in col_map:
                        return ""
                    v = row.iloc[col_map[key]]
                    if pd.isna(v):
                        return ""
                    s = str(v).strip()
                    return "" if s in _EMPTY_VALUES else s

                for idx, row in df.iterrows():
                    if row.isna().all():
                        continue
                    item_name = _val("item_name")
                    if not item_name:
                        continue

                    items.append(TableItem(
                        item_name=item_name,
                        data_type=_val("data_type"),
                        length=_val("length") or None,
                        integer_part=_val("integer_part") or None,
                        decimal_part=_val("decimal_part") or None,
                        row_number=header_row + idx + 2,
                        source_file=path.name,
                        source_sheet=sheet,
                    ))
                print(f"[INFO] テーブル定義読み込み: {path.name}/{sheet}")
            except KeyError as e:
                print(f"[警告] {path.name}({sheet}): {e} -> スキップ")
                continue
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計テーブル定義: {len(items)}件")
    return items


# ====== マッチング結果型 ======================================================

@dataclass
class MatchResult:
    """ドメインマッチング結果。

    domain_matcher.py (C方式) / llm_matcher.py (B方式) と共通の型。

    match_type の値:
        レガシー: "exact", "similar", "unmatched"
        C方式:    "exact", "special_flag", "special_date",
                  "special_date_individual", "special_comment",
                  "similar", "needs_llm", "unmatched"

    特殊パターン（special_flag, special_comment）は自動割当ではなく
    人間選択方式。reason に選択肢テキストが格納される。
      フラグ: 「以下から選択: 0/1, 0/9, null/1, スペース/1」
      コメント/補記: 「改行あり/なしを確認して選択してください」
    """
    match_type: str
    domain: Optional[DomainDef]
    confidence: float
    reason: str
    generalized_candidates: list[str] | None = None


# match_type → 表示用「判定結果」の変換マップ（確定仕様）
_MATCH_TYPE_DISPLAY = {
    "exact": "完全一致",
    "special_flag": "選択必須",
    "special_comment": "選択必須",
    "human_choice": "選択必須",
    "new_domain": "提案",
    "suggested": "提案",
    "no_match": "提案",
    "unmatched": "提案",
    "no_type_digits_match": "提案",
    "unnecessary": "対象外",
    "out_of_scope": "対象外",
}


def _match_result_to_display(result: MatchResult) -> str:
    """MatchResult.match_type を表示用「判定結果」文字列に変換する。"""
    return _MATCH_TYPE_DISPLAY.get(result.match_type, "提案")


def _resolve_config_path(path_str: str) -> Path:
    """設定ファイルのパスを解決する。"""
    path = Path(path_str)
    if path.is_absolute():
        return path
    cwd_candidate = Path(path_str)
    if cwd_candidate.exists():
        return cwd_candidate.resolve()
    root_candidate = Path(__file__).resolve().parents[2] / path_str
    if root_candidate.exists():
        return root_candidate
    return Path(__file__).resolve().parent / path_str


def _load_synonym_dict(cfg: Dict[str, Any]) -> Dict[str, List[str]]:
    """同義語辞書（YAML）を読み込み、domain_matcher 用の形式に変換する。"""
    path_str = cfg.get("SYNONYM_FILE_PATH") or "word/domain/synonyms.yaml"
    path = _resolve_config_path(path_str)
    try:
        import yaml
    except Exception:
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except Exception:
        return {}

    if isinstance(data, dict) and isinstance(data.get("synonyms"), dict):
        return data["synonyms"]

    groups = data.get("groups") if isinstance(data, dict) else None
    if isinstance(groups, list):
        result: Dict[str, List[str]] = {}
        for group in groups:
            if not group:
                continue
            if not isinstance(group, list):
                continue
            canonical = str(group[0]).strip()
            if not canonical:
                continue
            variants = [str(v).strip() for v in group[1:] if str(v).strip()]
            result[canonical] = variants
        return result

    return data if isinstance(data, dict) else {}


def _make_placeholder_domain(name: str) -> DomainDef:
    """ドメイン名だけが判明している場合のダミー DomainDef を生成する。"""
    return DomainDef(
        name=name,
        data_type="",
        min_char=None,
        max_char=None,
        min_byte=None,
        max_byte=None,
        integer_digits=None,
        decimal_digits=None,
        min_value=None,
        max_value=None,
        regex=None,
        code_id="",
        row_number=0,
        source_file="",
        source_sheet="",
    )


def _lookup_domain_by_name(name: str, domains: Dict[str, DomainDef]) -> Optional[DomainDef]:
    """ドメイン名から DomainDef を取得（なければ None）。"""
    if not name:
        return None
    return domains.get(normalize_text(name))


def _match_from_evidence(
    evidence,
    domains: Dict[str, DomainDef],
    cfg: Dict[str, Any],
) -> MatchResult:
    """C方式の evidence から MatchResult を生成する。"""
    if resolve_without_llm is None:
        return MatchResult("needs_llm", None, 0.0, "C方式未使用")

    resolved = resolve_without_llm(evidence)
    if resolved is None:
        return MatchResult("needs_llm", None, 0.0, evidence.reason)

    domain = None
    reason = resolved.reason
    if resolved.match_type in {"special_flag", "special_comment"}:
        base = evidence.resolved_domain or resolved.reason
        if resolved.match_type == "special_flag":
            reason = f"フラグ: {base}"
        else:
            reason = f"コメント: {base}"
    elif resolved.domain_name:
        domain = _lookup_domain_by_name(resolved.domain_name, domains)
        if domain is None:
            domain = _make_placeholder_domain(resolved.domain_name)

    return MatchResult(resolved.match_type, domain, resolved.confidence, reason)


def _match_from_llm(
    evidence,
    domains: Dict[str, DomainDef],
    cfg: Dict[str, Any],
    llm_call=None,
) -> MatchResult:
    """B方式（LLM/フォールバック）から MatchResult を生成する。"""
    if run_llm_matching is None:
        return MatchResult("needs_llm", None, 0.0, evidence.reason)

    llm_results = run_llm_matching([evidence], cfg, llm_call=llm_call)
    llm_result = llm_results[0]

    domain = None
    reason = llm_result.reason
    if llm_result.match_type == "human_choice":
        base = llm_result.check_point or llm_result.reason
        if base and "改行" in base:
            reason = f"コメント: {base}"
        elif base and "0/1" in base:
            reason = f"フラグ: {base}"
        else:
            reason = base
    elif llm_result.match_type == "suggested":
        if llm_result.suggested_domain:
            domain = _lookup_domain_by_name(llm_result.suggested_domain, domains)
            if domain is None:
                domain = _make_placeholder_domain(llm_result.suggested_domain)

    return MatchResult(
        llm_result.match_type,
        domain,
        llm_result.confidence,
        reason,
        generalized_candidates=getattr(llm_result, "generalized_candidates", None),
    )


def _llm_result_to_match(
    evidence,
    llm_result,
    domains: Dict[str, DomainDef],
) -> MatchResult:
    """llm_matcher の結果を MatchResult に変換する。"""
    domain = None
    reason = llm_result.reason
    if llm_result.match_type == "human_choice":
        reason = llm_result.check_point or llm_result.reason
    elif llm_result.match_type in {"suggested", "new_domain"}:
        if llm_result.suggested_domain:
            domain = _lookup_domain_by_name(llm_result.suggested_domain, domains)
            if domain is None:
                domain = _make_placeholder_domain(llm_result.suggested_domain)

    return MatchResult(
        llm_result.match_type,
        domain,
        llm_result.confidence,
        reason,
        generalized_candidates=getattr(llm_result, "generalized_candidates", None),
    )


def _domain_detail_dict(domain: Optional[DomainDef]) -> Dict[str, str]:
    """DomainDef からドメイン詳細列の辞書を生成する（殿の確定仕様: 12列）。"""
    if domain is None:
        return {
            "D_データ型": "",
            "D_最小文字数": "", "D_最大文字数": "",
            "D_最小バイト長": "", "D_最大バイト長": "",
            "D_最小値": "", "D_最大値": "",
            "D_整数部桁数": "", "D_小数部桁数": "",
            "D_書式（正規表現）": "", "D_参照外部コード": "",
        }

    def _s(val: Optional[str]) -> str:
        if val is None:
            return ""
        v = str(val).strip()
        return "" if v in _EMPTY_VALUES else v

    return {
        "D_データ型": _s(domain.data_type),
        "D_最小文字数": _s(domain.min_char),
        "D_最大文字数": _s(domain.max_char),
        "D_最小バイト長": _s(domain.min_byte),
        "D_最大バイト長": _s(domain.max_byte),
        "D_最小値": _s(domain.min_value),
        "D_最大値": _s(domain.max_value),
        "D_整数部桁数": _s(domain.integer_digits),
        "D_小数部桁数": _s(domain.decimal_digits),
        "D_書式（正規表現）": _s(domain.regex),
        "D_参照外部コード": _s(domain.code_id),
    }


def _make_lookup_key(name: str, digits: Optional[str]) -> str:
    """VLOOKUP用の検索キーを生成する。"""
    d = str(digits).strip() if digits else ""
    return f"{name}_{d}" if d else name


def _make_screen_digits_key(max_digits: Optional[str], max_bytes: Optional[str]) -> str:
    """画面項目定義の桁数キーを生成する。

    優先順位:
      1) 最大桁
      2) 最大バイト数
    """
    d = str(max_digits).strip() if max_digits else ""
    if d:
        return d
    b = str(max_bytes).strip() if max_bytes else ""
    return b


def _make_table_digits_key(length: Optional[str], integer_part: Optional[str], decimal_part: Optional[str]) -> str:
    """テーブル定義の桁数キーを生成する。

    優先順位:
      1) 整数部/小数部（数値系）
      2) Length（文字列系）
    """
    int_s = str(integer_part).strip() if integer_part else ""
    dec_s = str(decimal_part).strip() if decimal_part else ""
    if int_s or dec_s:
        if not int_s:
            int_s = "0"
        if not dec_s:
            dec_s = "0"
        return f"{int_s}.{dec_s}"
    return str(length).strip() if length else ""


def _find_domain_by_stripped_name(domains: Dict[str, DomainDef], item_name: str) -> Optional[DomainDef]:
    """項目名（数字除去後）に一致するドメインを1件返す。"""
    item_key = _strip_digits(normalize_text(item_name))
    for key, domain in domains.items():
        if _strip_digits(key) == item_key:
            return domain
    return None


def _domain_digits_key(domain: DomainDef) -> str:
    """ドメイン定義の桁数キーを生成する（画面/テーブル共通の比較用）。"""
    int_dec = _make_table_digits_key(None, domain.integer_digits, domain.decimal_digits)
    if int_dec:
        return int_dec
    # 文字列系は最大文字数 → 最大バイト数の順で使用
    return _make_screen_digits_key(domain.max_char, domain.max_byte)


def suggest_domain_from_screen_item(item: ScreenItem) -> Dict[str, str]:
    """画面項目定義からドメイン提案を生成（確定仕様対応）"""
    suggestions: Dict[str, str] = {}

    if item.data_type:
        dt = str(item.data_type).lower().strip()
        if any(k in dt for k in ("text", "string", "varchar")):
            suggestions["データ型"] = "VARCHAR"
        elif any(k in dt for k in ("number", "integer", "int")):
            suggestions["データ型"] = "INTEGER"
        elif "date" in dt:
            suggestions["データ型"] = "DATE"
        elif any(k in dt for k in ("decimal", "float", "numeric")):
            suggestions["データ型"] = "DECIMAL"
        elif any(k in dt for k in ("boolean", "bool")):
            suggestions["データ型"] = "BOOLEAN"
        else:
            suggestions["データ型"] = item.data_type

    if item.max_digits is not None:
        mv = str(item.max_digits).strip()
        if mv and mv not in _EMPTY_VALUES:
            suggestions["最大文字数"] = mv

    if item.max_bytes is not None:
        bv = str(item.max_bytes).strip()
        if bv and bv not in _EMPTY_VALUES:
            suggestions["最大バイト数"] = bv

    return suggestions


def suggest_domain_from_table_item(item: TableItem) -> Dict[str, str]:
    """テーブル定義からドメイン提案を生成（仕様書: 型+桁数ベース）"""
    suggestions: Dict[str, str] = {}

    if item.data_type:
        dt = str(item.data_type).lower().strip()
        if any(k in dt for k in ("char", "varchar", "string", "text")):
            suggestions["データ型"] = "VARCHAR"
        elif any(k in dt for k in ("number", "integer", "int")):
            suggestions["データ型"] = "INTEGER"
        elif "date" in dt or "timestamp" in dt:
            suggestions["データ型"] = "DATE"
        elif any(k in dt for k in ("decimal", "float", "numeric")):
            suggestions["データ型"] = "DECIMAL"
        elif any(k in dt for k in ("boolean", "bool")):
            suggestions["データ型"] = "BOOLEAN"
        else:
            suggestions["データ型"] = item.data_type

    if item.length is not None:
        lv = str(item.length).strip()
        if lv and lv not in _EMPTY_VALUES:
            suggestions["最大文字数"] = lv

    if item.integer_part is not None:
        iv = str(item.integer_part).strip()
        if iv and iv not in _EMPTY_VALUES:
            suggestions["整数部桁数"] = iv

    if item.decimal_part is not None:
        dv = str(item.decimal_part).strip()
        if dv and dv not in _EMPTY_VALUES:
            suggestions["小数部桁数"] = dv

    return suggestions


def _should_propose(match_type: str) -> bool:
    return match_type not in {
        "exact",
        "special_flag",
        "special_comment",
        "human_choice",
        "unnecessary",
        "out_of_scope",
    }


def _is_no_match_type(match_type: str) -> bool:
    return match_type in {"unmatched", "no_match", "no_type_digits_match"}


def _is_new_domain_proposal(match_type: str, domain: Optional[DomainDef]) -> bool:
    """ドメイン一覧にない新規提案かどうかを判定する。"""
    return match_type == "new_domain"


def _should_auto_new_domain(match_type: str) -> bool:
    """LLMが明示しない場合のみ新規提案を補助生成する。"""
    return match_type in {"unmatched", "no_type_digits_match", "no_match"}


def _propose_new_domain_name(raw_name: str) -> str:
    """ドメイン一覧にない場合の仮提案名を生成する。"""
    base = _strip_digits(str(raw_name or "")).strip()
    if not base:
        return "新規ドメイン"
    return f"{base}_新規ドメイン"


# ====== 重複排除 ==============================================================

_RE_DIGITS = re.compile(r"\d+")


def _strip_digits(name: str) -> str:
    """項目名から数字を全て除去する。仕様書 Section 3-5 準拠。

    例: 「ああ1」→「ああ」、「金額01」→「金額」
    """
    return _RE_DIGITS.sub("", name).strip()


def dedup_by_name_and_digits(
    df: pd.DataFrame, name_col: str, digits_col: str,
) -> pd.DataFrame:
    """「項目名（数字除去後）＋桁数」が完全一致する行を1つに集約する。

    仕様書 Section 2「重複排除の基準」準拠。
    項目名列は数字除去後の値に置換される（重複排除シート用）。
    """
    df = df.copy()
    df[name_col] = df[name_col].apply(_strip_digits)
    digits = df[digits_col].fillna("").astype(str).str.strip() if digits_col in df.columns else pd.Series([""] * len(df))
    df["検索キー"] = df[name_col] + "_" + digits
    return df.drop_duplicates(subset=["検索キー"], keep="first").reset_index(drop=True)


# ====== 画面項目 突合処理 ====================================================

def process_screen_domain_matching(
    screen_items: List[ScreenItem],
    domains: Dict[str, DomainDef],
    cfg: Dict[str, Any],
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> pd.DataFrame:
    """画面項目とドメインの突合処理。


    Returns:
        全項目の照合結果DataFrame（検索キー付き）
    """
    if collect_evidence is None:
        raise RuntimeError("C方式マッチャーが読み込めません。domain_matcher.py を確認してください。")

    # --- 新ロジック: C方式→B方式パイプライン ---
    synonyms = _load_synonym_dict(cfg)
    results: List[Dict[str, Any]] = [None] * len(screen_items)
    evidences = []
    evidence_indices = []
    evidence_by_index: Dict[int, Any] = {}
    evidence_by_index: Dict[int, Any] = {}
    evidence_by_index: Dict[int, Any] = {}
    # 同名+別桁の検知用（画面）
    screen_name_digits: Dict[str, set] = {}
    for item in screen_items:
        key_name = _strip_digits(item.item_name)
        key_digits = _make_screen_digits_key(item.max_digits, item.max_bytes)
        if key_name:
            screen_name_digits.setdefault(key_name, set()).add(key_digits)

    for i, item in enumerate(screen_items):
        is_unnecessary = all(
            v is None or str(v).strip() in _EMPTY_VALUES
            for v in (item.text_type, item.min_digits, item.max_digits,
                      item.max_bytes, item.min_value, item.max_value, item.code_id)
        )
        if is_unnecessary:
            mr = MatchResult("unnecessary", None, 0.0, "ドメイン不要項目")
            results[i] = (item, mr)
            continue

        item_dict: Dict[str, Any] = {
            "item_name": item.item_name,
            "data_type": item.data_type,
            "text_type": item.text_type,
            "min_chars": item.min_digits,
            "max_chars": item.max_digits,
            "max_bytes": item.max_bytes,
            "min_value": item.min_value,
            "max_value": item.max_value,
            "external_code": item.code_id,
            "source": "screen",
        }
        ev = collect_evidence(item_dict, domains, cfg, synonyms=synonyms)
        evidences.append(ev)
        evidence_indices.append(i)
        evidence_by_index[i] = ev
        evidence_by_index[i] = ev
        evidence_by_index[i] = ev

    # C方式の確定分を先に埋める
    unresolved = []
    hard_resolved = {
        "exact",
        "special_flag",
        "special_comment",
        "special_date",
        "special_date_individual",
        "out_of_scope",
    }
    for ev, idx in zip(evidences, evidence_indices):
        if ev.is_resolved and ev.match_type in hard_resolved:
            mr = _match_from_evidence(ev, domains, cfg)
            results[idx] = (screen_items[idx], mr)
        else:
            unresolved.append((ev, idx))

    # B方式（LLM/フォールバック）
    if unresolved:
        use_llm = bool(cfg.get("LLM_MATCHING_ENABLED")) and run_llm_matching is not None and build_llm_call is not None
        llm_call = build_llm_call(cfg) if use_llm else None
        if use_llm:
            batch_size = int(cfg.get("LLM_MATCHING_BATCH_SIZE", 20))
            for start in range(0, len(unresolved), batch_size):
                chunk = unresolved[start:start + batch_size]
                chunk_evs = [ev for ev, _ in chunk]
                llm_results = run_llm_matching(chunk_evs, cfg, llm_call=llm_call)
                for (ev, idx), llm_result in zip(chunk, llm_results):
                    mr = _llm_result_to_match(ev, llm_result, domains)
                    results[idx] = (screen_items[idx], mr)
                    evidence_by_index[idx] = ev
        else:
            for ev, idx in unresolved:
                mr = _match_from_llm(ev, domains, cfg, llm_call=llm_call)
                results[idx] = (screen_items[idx], mr)
                evidence_by_index[idx] = ev

    # 結果整形
    output_rows: List[Dict[str, Any]] = []
    total = len(screen_items)
    for i, (item, mr) in enumerate(results):
        display_type = _match_result_to_display(mr)
        if display_type == "提案":
            display_type = "提案（候補から）" if not _is_new_domain_proposal(mr.match_type, mr.domain) else "提案（新規）"
        remark = mr.reason
        digit_rate: float | None = None
        if digit_match_rate_from_details and i in evidence_by_index:
            ev = evidence_by_index[i]
            details = getattr(ev, "detail_matches", None)
            rate = digit_match_rate_from_details(details) if details else None
            if isinstance(rate, float):
                digit_rate = rate
            if isinstance(rate, float) and rate < 0.8 and mr.domain:
                remark += " / 注意: 意味は近いが桁数が一致していない可能性"
        if mr.generalized_candidates:
            remark += f" / 汎化候補: {', '.join(mr.generalized_candidates)}"
            if mr.match_type == "exact" and _has_digit_mismatch(details):
                mr = MatchResult("new_domain", None, mr.confidence, mr.reason)
                display_type = "提案（新規）"
                remark += " / 注意: 完全一致だが桁数が一致していない → 新規提案"
        if mr.generalized_candidates:
            remark += f" / 汎化候補: {', '.join(mr.generalized_candidates)}"

        if _should_propose(mr.match_type):
            has_digits = bool(item.max_digits and str(item.max_digits).strip() not in _EMPTY_VALUES)
            has_bytes = bool(item.max_bytes and str(item.max_bytes).strip() not in _EMPTY_VALUES)
            if has_digits or has_bytes:
                remark = "桁数/バイト数の定義あり → 新規ドメイン提案が必要"
            else:
                remark = "桁数/バイト数なし → ドメイン設定不要の可能性"
            same_name_domain = _find_domain_by_stripped_name(domains, item.item_name)
            if same_name_domain:
                item_key = _make_screen_digits_key(item.max_digits, item.max_bytes)
                domain_key = _domain_digits_key(same_name_domain)
                if item_key and domain_key and item_key != domain_key:
                    remark += f" / 同名ドメインあり（桁数違い: 項目={item_key}, ドメイン={domain_key}）"
        # 同名で別桁定義が存在する場合は注意喚起
        name_key = _strip_digits(item.item_name)
        digits_key = _make_screen_digits_key(item.max_digits, item.max_bytes)
        if name_key in screen_name_digits and len(screen_name_digits[name_key]) > 1:
            others = sorted({d for d in screen_name_digits[name_key] if d != digits_key})
            if others:
                remark += f" / 同名で別桁定義あり（他: {', '.join(others)}）"

        proposed_domain = ""
        if mr.domain:
            proposed_domain = mr.domain.name
        elif mr.match_type == "new_domain":
            proposed_domain = _propose_new_domain_name(item.item_name)
        elif _should_auto_new_domain(mr.match_type):
            proposed_domain = _propose_new_domain_name(item.item_name)

        row: Dict[str, Any] = {
            "ファイル名": item.source_file,
            "項目名称": item.item_name,
            "型": item.data_type if item.data_type else "",
            "テキストタイプ": item.text_type if item.text_type else "",
            "最小桁": item.min_digits if item.min_digits else "",
            "最大桁": item.max_digits if item.max_digits else "",
            "最大バイト数": item.max_bytes if item.max_bytes else "",
            "最小値": item.min_value if item.min_value else "",
            "最大値": item.max_value if item.max_value else "",
            "外部コード": item.code_id if item.code_id else "",
            "一致ドメイン名": proposed_domain,
            "判定結果": display_type,
            "備考": remark,
            "桁一致率": "" if digit_rate is None else round(float(digit_rate), 3),
            "汎化候補": "" if not mr.generalized_candidates else ", ".join(mr.generalized_candidates),
        }
        row.update(_domain_detail_dict(mr.domain))

        if _should_propose(mr.match_type):
            suggestions = suggest_domain_from_screen_item(item)
            if "データ型" in suggestions:
                row["D_データ型"] = suggestions["データ型"]
            if "最大文字数" in suggestions:
                row["D_最大文字数"] = suggestions["最大文字数"]
            if "最大バイト数" in suggestions:
                row["D_最大バイト長"] = suggestions["最大バイト数"]

        output_rows.append(row)

        processed = i + 1
        if processed % 10 == 0 or processed == total:
            print(f"[INFO] 画面項目照合: {processed}/{total} 件処理済み")
        if progress_callback:
            progress_callback(processed, total)

    return _clean_dataframe(output_rows)


# ====== テーブル定義 突合処理 ================================================

def process_table_domain_matching(
    table_items: List[TableItem],
    domains: Dict[str, DomainDef],
    cfg: Dict[str, Any],
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> pd.DataFrame:
    """テーブル定義とドメインの突合処理。


    Returns:
        全項目の照合結果DataFrame（検索キー付き）
    """
    if collect_evidence is None:
        raise RuntimeError("C方式マッチャーが読み込めません。domain_matcher.py を確認してください。")

    # --- 新ロジック: C方式→B方式パイプライン ---
    synonyms = _load_synonym_dict(cfg)
    results: List[Dict[str, Any]] = [None] * len(table_items)
    evidences = []
    evidence_indices = []
    evidence_by_index: Dict[int, Any] = {}
    # 同名+別桁の検知用（テーブル）
    table_name_digits: Dict[str, set] = {}
    for item in table_items:
        key_name = _strip_digits(item.item_name)
        key_digits = _make_table_digits_key(item.length, item.integer_part, item.decimal_part)
        if key_name:
            table_name_digits.setdefault(key_name, set()).add(key_digits)

    for i, item in enumerate(table_items):
        is_unnecessary = all(
            v is None or str(v).strip() in _EMPTY_VALUES
            for v in (item.data_type, item.length, item.integer_part, item.decimal_part)
        )
        if is_unnecessary:
            mr = MatchResult("unnecessary", None, 0.0, "ドメイン不要項目")
            results[i] = (item, mr)
            continue

        item_dict: Dict[str, Any] = {
            "item_name": item.item_name,
            "data_type": item.data_type,
            "length": item.length,
            "integer_digits": item.integer_part,
            "decimal_digits": item.decimal_part,
            "source": "table",
        }
        ev = collect_evidence(item_dict, domains, cfg, synonyms=synonyms)
        evidences.append(ev)
        evidence_indices.append(i)

    unresolved = []
    hard_resolved = {
        "exact",
        "special_flag",
        "special_comment",
        "special_date",
        "special_date_individual",
        "out_of_scope",
    }
    for ev, idx in zip(evidences, evidence_indices):
        if ev.is_resolved and ev.match_type in hard_resolved:
            mr = _match_from_evidence(ev, domains, cfg)
            results[idx] = (table_items[idx], mr)
        else:
            unresolved.append((ev, idx))

    if unresolved:
        use_llm = bool(cfg.get("LLM_MATCHING_ENABLED")) and run_llm_matching is not None and build_llm_call is not None
        llm_call = build_llm_call(cfg) if use_llm else None
        if use_llm:
            batch_size = int(cfg.get("LLM_MATCHING_BATCH_SIZE", 20))
            for start in range(0, len(unresolved), batch_size):
                chunk = unresolved[start:start + batch_size]
                chunk_evs = [ev for ev, _ in chunk]
                llm_results = run_llm_matching(chunk_evs, cfg, llm_call=llm_call)
                for (ev, idx), llm_result in zip(chunk, llm_results):
                    mr = _llm_result_to_match(ev, llm_result, domains)
                    results[idx] = (table_items[idx], mr)
                    evidence_by_index[idx] = ev
        else:
            for ev, idx in unresolved:
                mr = _match_from_llm(ev, domains, cfg, llm_call=llm_call)
                results[idx] = (table_items[idx], mr)
                evidence_by_index[idx] = ev

    output_rows: List[Dict[str, Any]] = []
    total = len(table_items)
    for i, (item, mr) in enumerate(results):
        display_type = _match_result_to_display(mr)
        if display_type == "提案":
            display_type = "提案（候補から）" if not _is_new_domain_proposal(mr.match_type, mr.domain) else "提案（新規）"
        remark = mr.reason
        digit_rate: float | None = None
        if digit_match_rate_from_details and i in evidence_by_index:
            ev = evidence_by_index[i]
            details = getattr(ev, "detail_matches", None)
            rate = digit_match_rate_from_details(details) if details else None
            if isinstance(rate, float):
                digit_rate = rate
            if isinstance(rate, float) and rate < 0.8 and mr.domain:
                remark += " / 注意: 意味は近いが桁数が一致していない可能性"

        if _should_propose(mr.match_type):
            has_type = bool(item.data_type and item.data_type.strip())
            has_length = bool(item.length and str(item.length).strip() not in _EMPTY_VALUES)
            if has_type or has_length:
                remark = "データ型/桁数の定義あり → 新規ドメイン提案が必要"
            else:
                remark = "データ型/桁数なし → ドメイン設定不要の可能性"
            same_name_domain = _find_domain_by_stripped_name(domains, item.item_name)
            if same_name_domain:
                item_key = _make_table_digits_key(item.length, item.integer_part, item.decimal_part)
                domain_key = _domain_digits_key(same_name_domain)
                if item_key and domain_key and item_key != domain_key:
                    remark += f" / 同名ドメインあり（桁数違い: 項目={item_key}, ドメイン={domain_key}）"
        # 同名で別桁定義が存在する場合は注意喚起
        name_key = _strip_digits(item.item_name)
        digits_key = _make_table_digits_key(item.length, item.integer_part, item.decimal_part)
        if name_key in table_name_digits and len(table_name_digits[name_key]) > 1:
            others = sorted({d for d in table_name_digits[name_key] if d != digits_key})
            if others:
                remark += f" / 同名で別桁定義あり（他: {', '.join(others)}）"

        if mr.domain and item.data_type:
            item_type_norm = normalize_data_type(item.data_type)
            domain_type_norm = normalize_data_type(mr.domain.data_type)
            if item_type_norm and domain_type_norm and item_type_norm != domain_type_norm:
                remark += f" / 型不一致: テーブル={item.data_type}, ドメイン={mr.domain.data_type}"

        proposed_domain = ""
        if mr.domain:
            proposed_domain = mr.domain.name
        elif mr.match_type == "new_domain":
            proposed_domain = _propose_new_domain_name(item.item_name)
        elif _should_auto_new_domain(mr.match_type):
            proposed_domain = _propose_new_domain_name(item.item_name)

        row: Dict[str, Any] = {
            "ファイル名": item.source_file,
            "論理項目名": item.item_name,
            "データ型": item.data_type if item.data_type else "",
            "Length": item.length if item.length else "",
            "全体数値": item.integer_part if item.integer_part else "",
            "少数桁": item.decimal_part if item.decimal_part else "",
            "一致ドメイン名": proposed_domain,
            "判定結果": display_type,
            "備考": remark,
            "桁一致率": "" if digit_rate is None else round(float(digit_rate), 3),
            "汎化候補": "" if not mr.generalized_candidates else ", ".join(mr.generalized_candidates),
        }
        row.update(_domain_detail_dict(mr.domain))
        if _should_propose(mr.match_type):
            suggestions = suggest_domain_from_table_item(item)
            if "データ型" in suggestions:
                row["D_データ型"] = suggestions["データ型"]
            if "最大文字数" in suggestions:
                row["D_最大文字数"] = suggestions["最大文字数"]
            if "整数部桁数" in suggestions:
                row["D_整数部桁数"] = suggestions["整数部桁数"]
            if "小数部桁数" in suggestions:
                row["D_小数部桁数"] = suggestions["小数部桁数"]
        output_rows.append(row)

        processed = i + 1
        if processed % 10 == 0 or processed == total:
            print(f"[INFO] テーブル定義照合: {processed}/{total} 件処理済み")
        if progress_callback:
            progress_callback(processed, total)

    return _clean_dataframe(output_rows)


def _clean_dataframe(results: List[Dict[str, Any]]) -> pd.DataFrame:
    """結果辞書リストをDataFrameに変換し、NaN/nan文字列を除去する。"""
    df = pd.DataFrame(results)
    df = df.fillna("")
    df = df.replace("nan", "")
    df = df.replace("None", "")
    return df


# ====== 出力 ==================================================================

# ドメイン詳細列の定義（殿の確定仕様: 11列）
_DOMAIN_DETAIL_COLS = [
    "D_データ型",
    "D_最小文字数", "D_最大文字数",
    "D_最小バイト長", "D_最大バイト長",
    "D_最小値", "D_最大値",
    "D_整数部桁数", "D_小数部桁数",
    "D_書式（正規表現）", "D_参照外部コード",
]

# 抽出シートに含める列（画面項目 — 殿の確定仕様シート1: A〜J）
_SCREEN_RAW_COLS = [
    "ファイル名", "項目名称", "型", "テキストタイプ",
    "最小桁", "最大桁", "最大バイト数", "最小値", "最大値", "外部コード",
]

# 抽出シートに含める列（テーブル定義 — 殿の確定仕様シート2: A〜F）
_TABLE_RAW_COLS = [
    "ファイル名", "論理項目名", "データ型", "Length", "全体数値", "少数桁",
]

# 重複排除シートの列（画面項目 — 殿の確定仕様シート3: A〜L）
_SCREEN_DEDUP_COLS = [
    "項目名称", "型", "テキストタイプ",
    "最小桁", "最大桁", "最大バイト数", "最小値", "最大値", "外部コード",
    "一致ドメイン名", "判定結果", "備考", "桁一致率", "汎化候補",
]

# 重複排除シートの列（テーブル定義 — 殿の確定仕様シート4: A〜H）
_TABLE_DEDUP_COLS = [
    "論理項目名", "データ型", "Length", "全体数値", "少数桁",
    "一致ドメイン名", "判定結果", "備考", "桁一致率", "汎化候補",
]


def save_domain_check_results(
    screen_df: pd.DataFrame,
    table_df: pd.DataFrame,
    screen_dedup_df: pd.DataFrame,
    table_dedup_df: pd.DataFrame,
    cfg: Dict[str, Any],
    include_screen: bool = True,
    include_table: bool = True,
) -> Path:
    """4シート構成のExcelファイルを保存する（確定仕様）。

    シート1: 画面項目_抽出（生データ + VLOOKUP列）
    シート2: テーブル定義_抽出（生データ + VLOOKUP列）
    シート3: 画面項目_重複排除（A〜L + ドメイン詳細M〜）
    シート4: テーブル定義_重複排除（A〜H + ドメイン詳細I〜）

    Returns:
        出力ファイルのPath
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    def _find_col(ws, header: str) -> int:
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == header:
                return c
        raise ValueError(f"列が見つかりません: {ws.title}/{header}")

    out_dir = Path(cfg["OUT_DIR"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    output_file = out_dir / "domain_check_result.xlsx"

    # 抽出シート用に列を絞る（A〜J / A〜F）
    screen_raw = _select_columns(screen_df, _SCREEN_RAW_COLS)
    table_raw = _select_columns(table_df, _TABLE_RAW_COLS)

    # 重複排除シート用に列を絞る（基本列 + ドメイン詳細列）
    screen_dedup = _select_columns(screen_dedup_df, _SCREEN_DEDUP_COLS + _DOMAIN_DETAIL_COLS)
    table_dedup = _select_columns(table_dedup_df, _TABLE_DEDUP_COLS + _DOMAIN_DETAIL_COLS)
    proposal_df = _build_domain_proposal_df(
        screen_dedup_df, table_dedup_df, cfg, include_screen, include_table,
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if include_screen:
            screen_raw.to_excel(writer, sheet_name="画面項目_抽出", index=False)
            screen_dedup.to_excel(writer, sheet_name="画面項目_重複排除", index=False)
        if include_table:
            table_raw.to_excel(writer, sheet_name="テーブル定義_抽出", index=False)
            table_dedup.to_excel(writer, sheet_name="テーブル定義_重複排除", index=False)
        if not proposal_df.empty:
            if include_screen:
                screen_prop = proposal_df[proposal_df.get("由来") == "画面"]
                if not screen_prop.empty:
                    screen_prop.to_excel(writer, sheet_name="提案_ドメイン一覧_画面", index=False)
            if include_table:
                table_prop = proposal_df[proposal_df.get("由来") == "テーブル"]
                if not table_prop.empty:
                    table_prop.to_excel(writer, sheet_name="提案_ドメイン一覧_テーブル", index=False)

    wb = load_workbook(output_file)

    if include_screen:
        # シート1: 検索キー列 + VLOOKUP列
        ws1 = wb["画面項目_抽出"]
        vlookup_col_s = len(_SCREEN_RAW_COLS) + 1  # K列
        ws1.cell(row=1, column=vlookup_col_s, value="VLOOKUP")
        # 検索キー（L列）: 項目名称（数字除去後）＋桁数(最大桁 or 最大バイト数)
        key_col_s = vlookup_col_s + 1  # L列
        ws1.cell(row=1, column=key_col_s, value="検索キー")
        for r in range(2, ws1.max_row + 1):
            raw_name = ws1.cell(row=r, column=2).value  # B列
            raw_digits = ws1.cell(row=r, column=6).value  # F列=最大桁
            raw_bytes = ws1.cell(row=r, column=7).value  # G列=最大バイト数
            digit_key = _make_screen_digits_key(raw_digits, raw_bytes)
            key = _make_lookup_key(_strip_digits(str(raw_name or "")), digit_key or None)
            ws1.cell(row=r, column=key_col_s, value=key)
        # 画面項目_重複排除の一致ドメイン名（J列）を検索キーで引く
        ws1_key_col_letter = get_column_letter(key_col_s)
        ws3 = wb["画面項目_重複排除"]
        ws3_domain_col = _find_col(ws3, "一致ドメイン名")
        ws3_result_col = _find_col(ws3, "判定結果")
        ws3_remark_col = _find_col(ws3, "備考")
        ws3_key_col = ws3.max_column + 1
        ws3.cell(row=1, column=ws3_key_col, value="検索キー")
        for r in range(2, ws3.max_row + 1):
            name = ws3.cell(row=r, column=1).value  # A列
            digits = ws3.cell(row=r, column=5).value  # E列=最大桁
            bytes_ = ws3.cell(row=r, column=6).value  # F列=最大バイト数
            digit_key = _make_screen_digits_key(digits, bytes_)
            key = _make_lookup_key(str(name or ""), digit_key or None)
            ws3.cell(row=r, column=ws3_key_col, value=key)
        dedup_key_col_s = get_column_letter(ws3_key_col)
        # 画面項目_重複排除の検索キー列でMATCHし、J列をINDEX
        for r in range(2, ws1.max_row + 1):
            ws1.cell(
                row=r,
                column=vlookup_col_s,
                value=(
                    f'=IFERROR(INDEX(画面項目_重複排除!{get_column_letter(ws3_domain_col)}:{get_column_letter(ws3_domain_col)},'
                    f'MATCH({ws1_key_col_letter}{r},画面項目_重複排除!{dedup_key_col_s}:{dedup_key_col_s},0)),"")'
                ),
            )

        # VLOOKUP列を分割（一致ドメイン名 / 判定結果 / 備考）
        split_start = key_col_s + 1
        ws1.cell(row=1, column=split_start, value="一致ドメイン名")
        ws1.cell(row=1, column=split_start + 1, value="判定結果")
        ws1.cell(row=1, column=split_start + 2, value="備考")
        for r in range(2, ws1.max_row + 1):
            ws1.cell(
                row=r,
                column=split_start,
                value=(
                    f'=IFERROR(INDEX(画面項目_重複排除!{get_column_letter(ws3_domain_col)}:{get_column_letter(ws3_domain_col)},'
                    f'MATCH({ws1_key_col_letter}{r},画面項目_重複排除!{dedup_key_col_s}:{dedup_key_col_s},0)),"")'
                ),
            )
            ws1.cell(
                row=r,
                column=split_start + 1,
                value=(
                    f'=IFERROR(INDEX(画面項目_重複排除!{get_column_letter(ws3_result_col)}:{get_column_letter(ws3_result_col)},'
                    f'MATCH({ws1_key_col_letter}{r},画面項目_重複排除!{dedup_key_col_s}:{dedup_key_col_s},0)),"")'
                ),
            )
            ws1.cell(
                row=r,
                column=split_start + 2,
                value=(
                    f'=IFERROR(INDEX(画面項目_重複排除!{get_column_letter(ws3_remark_col)}:{get_column_letter(ws3_remark_col)},'
                    f'MATCH({ws1_key_col_letter}{r},画面項目_重複排除!{dedup_key_col_s}:{dedup_key_col_s},0)),"")'
                ),
            )

    if include_table:
        # シート2: 検索キー列 + VLOOKUP列
        ws2 = wb["テーブル定義_抽出"]
        vlookup_col_t = len(_TABLE_RAW_COLS) + 1  # G列
        ws2.cell(row=1, column=vlookup_col_t, value="VLOOKUP")
        # 検索キー（H列）: 論理項目名（数字除去後）＋桁数(整数/小数 or Length)
        key_col_t = vlookup_col_t + 1  # H列
        ws2.cell(row=1, column=key_col_t, value="検索キー")
        for r in range(2, ws2.max_row + 1):
            raw_name = ws2.cell(row=r, column=2).value  # B列
            raw_length = ws2.cell(row=r, column=4).value  # D列=Length
            raw_int = ws2.cell(row=r, column=5).value  # E列=全体数値
            raw_dec = ws2.cell(row=r, column=6).value  # F列=少数桁
            digit_key = _make_table_digits_key(raw_length, raw_int, raw_dec)
            key = _make_lookup_key(_strip_digits(str(raw_name or "")), digit_key or None)
            ws2.cell(row=r, column=key_col_t, value=key)
        ws4 = wb["テーブル定義_重複排除"]
        ws4_domain_col = _find_col(ws4, "一致ドメイン名")
        ws4_result_col = _find_col(ws4, "判定結果")
        ws4_remark_col = _find_col(ws4, "備考")
        ws4_key_col = ws4.max_column + 1
        ws4.cell(row=1, column=ws4_key_col, value="検索キー")
        for r in range(2, ws4.max_row + 1):
            name = ws4.cell(row=r, column=1).value  # A列
            length = ws4.cell(row=r, column=3).value  # C列=Length
            int_part = ws4.cell(row=r, column=4).value  # D列=全体数値
            dec_part = ws4.cell(row=r, column=5).value  # E列=少数桁
            digit_key = _make_table_digits_key(length, int_part, dec_part)
            key = _make_lookup_key(str(name or ""), digit_key or None)
            ws4.cell(row=r, column=ws4_key_col, value=key)
        dedup_key_col_t = get_column_letter(ws4_key_col)
        ws2_key_col_letter = get_column_letter(key_col_t)
        # テーブル定義_重複排除の検索キー列でMATCHし、F列をINDEX
        for r in range(2, ws2.max_row + 1):
            ws2.cell(
                row=r,
                column=vlookup_col_t,
                value=(
                    f'=IFERROR(INDEX(テーブル定義_重複排除!{get_column_letter(ws4_domain_col)}:{get_column_letter(ws4_domain_col)},'
                    f'MATCH({ws2_key_col_letter}{r},テーブル定義_重複排除!{dedup_key_col_t}:{dedup_key_col_t},0)),"")'
                ),
            )

        # VLOOKUP列を分割（一致ドメイン名 / 判定結果 / 備考）
        split_start = key_col_t + 1
        ws2.cell(row=1, column=split_start, value="一致ドメイン名")
        ws2.cell(row=1, column=split_start + 1, value="判定結果")
        ws2.cell(row=1, column=split_start + 2, value="備考")
        for r in range(2, ws2.max_row + 1):
            ws2.cell(
                row=r,
                column=split_start,
                value=(
                    f'=IFERROR(INDEX(テーブル定義_重複排除!{get_column_letter(ws4_domain_col)}:{get_column_letter(ws4_domain_col)},'
                    f'MATCH({ws2_key_col_letter}{r},テーブル定義_重複排除!{dedup_key_col_t}:{dedup_key_col_t},0)),"")'
                ),
            )
            ws2.cell(
                row=r,
                column=split_start + 1,
                value=(
                    f'=IFERROR(INDEX(テーブル定義_重複排除!{get_column_letter(ws4_result_col)}:{get_column_letter(ws4_result_col)},'
                    f'MATCH({ws2_key_col_letter}{r},テーブル定義_重複排除!{dedup_key_col_t}:{dedup_key_col_t},0)),"")'
                ),
            )
            ws2.cell(
                row=r,
                column=split_start + 2,
                value=(
                    f'=IFERROR(INDEX(テーブル定義_重複排除!{get_column_letter(ws4_remark_col)}:{get_column_letter(ws4_remark_col)},'
                    f'MATCH({ws2_key_col_letter}{r},テーブル定義_重複排除!{dedup_key_col_t}:{dedup_key_col_t},0)),"")'
                ),
            )

    # 判定結果列に色付け
    for sheet_name in ("画面項目_重複排除", "テーブル定義_重複排除"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _apply_result_colors(ws, ws.max_row)

    wb.save(output_file)
    print(f"[INFO] 保存完了: {output_file}")
    return output_file


def _select_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """DataFrameから存在する列のみを選択する。"""
    existing = [c for c in cols if c in df.columns]
    return df[existing].copy()


def _first_value(*values: Any) -> str:
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if not s or s in _EMPTY_VALUES:
            continue
        return s
    return ""


def _has_digit_mismatch(details: dict | None) -> bool:
    if not details:
        return False
    keys = {
        "string_length_match",
        "string_max_chars_match",
        "string_max_bytes_match",
        "numeric_integer_digits_match",
        "numeric_decimal_digits_match",
        "numeric_min_max_match",
    }
    for k in keys:
        v = details.get(k)
        if v is False:
            return True
    return False


def _normalize_proposal_name(name: str) -> str:
    if not name:
        return ""
    s = normalize_text(name)
    s = re.sub(r"(コード|id|番号)", "", s)
    s = re.sub(r"[\u3000\s]+", " ", s).strip()
    return s


def _proposal_signature(row: pd.Series) -> tuple[str, str]:
    dt = _first_value(row.get("データ型"))
    digits = _first_value(
        row.get("整数部桁数"), row.get("小数部桁数"),
        row.get("最大文字数"), row.get("最大バイト長"),
    )
    return (dt, digits)


def _group_new_domain_proposals(df: pd.DataFrame, name_col: str, llm_call) -> pd.DataFrame:
    """新規提案の中で汎化できそうな候補をグルーピングする（LLM優先）。"""
    if df.empty or name_col not in df.columns:
        df["汎化候補（新規内）"] = ""
        df["同義候補（新規内）"] = ""
        return df

    names = df[name_col].fillna("").astype(str).tolist()
    items: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        items.append({
            "name": row.get(name_col, ""),
            "data_type": _first_value(row.get("データ型")),
            "digits": _first_value(
                row.get("整数部桁数"), row.get("小数部桁数"),
                row.get("最大文字数"), row.get("最大バイト長"),
            ),
        })

    generalized = [""] * len(df)
    group_names = [""] * len(df)

    if group_new_domain_proposals and llm_call is not None:
        groups = group_new_domain_proposals(items, llm_call)
        for g in groups:
            members = [i - 1 for i in g.get("group", []) if isinstance(i, int)]
            members = [i for i in members if 0 <= i < len(df)]
            if len(members) <= 1:
                continue
            master = names[members[0]]
            member_names = [names[i] for i in members if names[i]]
            for i in members:
                generalized[i] = master if names[i] != master else ""
                group_names[i] = ", ".join(member_names)
    else:
        from difflib import SequenceMatcher
        groups: list[dict[str, Any]] = []
        for idx, name in enumerate(names):
            norm = _normalize_proposal_name(name)
            sig = _proposal_signature(df.iloc[idx])
            if not norm:
                continue
            best_group = None
            best_score = 0.0
            for g in groups:
                if g["signature"] != sig:
                    continue
                score = SequenceMatcher(None, norm, g["norm"]).ratio()
                if score > best_score:
                    best_score = score
                    best_group = g
            if best_group and best_score >= 0.75:
                best_group["members"].append(idx)
                best_group["names"].append(name)
            else:
                groups.append({
                    "signature": sig,
                    "norm": norm,
                    "members": [idx],
                    "names": [name],
                })

        for g in groups:
            if len(g["members"]) <= 1:
                continue
            master = g["names"][0]
            members = [n for n in g["names"] if n]
            for i in g["members"]:
                generalized[i] = master if names[i] != master else ""
                group_names[i] = ", ".join(members)

    df = df.copy()
    df["汎化候補（新規内）"] = generalized
    df["同義候補（新規内）"] = group_names
    return df


def _build_domain_proposal_df(
    screen_dedup_df: pd.DataFrame,
    table_dedup_df: pd.DataFrame,
    cfg: Dict[str, Any],
    include_screen: bool,
    include_table: bool,
) -> pd.DataFrame:
    domain_cols = _domain_column_names(cfg)
    rows: List[Dict[str, Any]] = []

    def add_screen_rows(df: pd.DataFrame) -> None:
        if df.empty:
            return
        target = df[df.get("判定結果") == "提案（新規）"]
        for _, row in target.iterrows():
            name = _first_value(row.get("一致ドメイン名"), _propose_new_domain_name(row.get("項目名称")))
            rows.append({
                "由来": "画面",
                domain_cols["name"]: name,
                domain_cols["data_type"]: _first_value(row.get("D_データ型"), row.get("型")),
                domain_cols["min_char"]: _first_value(row.get("D_最小文字数"), row.get("最小桁")),
                domain_cols["max_char"]: _first_value(row.get("D_最大文字数"), row.get("最大桁")),
                domain_cols["min_byte"]: _first_value(row.get("D_最小バイト長")),
                domain_cols["max_byte"]: _first_value(row.get("D_最大バイト長"), row.get("最大バイト数")),
                domain_cols["integer_digits"]: _first_value(row.get("D_整数部桁数")),
                domain_cols["decimal_digits"]: _first_value(row.get("D_小数部桁数")),
                domain_cols["min_value"]: _first_value(row.get("D_最小値"), row.get("最小値")),
                domain_cols["max_value"]: _first_value(row.get("D_最大値"), row.get("最大値")),
                domain_cols["regex"]: _first_value(row.get("D_書式（正規表現）")),
                domain_cols["code_id"]: _first_value(row.get("D_参照外部コード"), row.get("外部コード")),
                "根拠": _first_value(row.get("備考"), row.get("判定理由"), row.get("理由")),
                "汎化候補": _first_value(row.get("汎化候補")),
            })

    def add_table_rows(df: pd.DataFrame) -> None:
        if df.empty:
            return
        target = df[df.get("判定結果") == "提案（新規）"]
        for _, row in target.iterrows():
            name = _first_value(row.get("一致ドメイン名"), _propose_new_domain_name(row.get("論理項目名")))
            rows.append({
                "由来": "テーブル",
                domain_cols["name"]: name,
                domain_cols["data_type"]: _first_value(row.get("D_データ型"), row.get("データ型")),
                domain_cols["min_char"]: _first_value(row.get("D_最小文字数")),
                domain_cols["max_char"]: _first_value(row.get("D_最大文字数"), row.get("Length")),
                domain_cols["min_byte"]: _first_value(row.get("D_最小バイト長")),
                domain_cols["max_byte"]: _first_value(row.get("D_最大バイト長")),
                domain_cols["integer_digits"]: _first_value(row.get("D_整数部桁数"), row.get("全体数値")),
                domain_cols["decimal_digits"]: _first_value(row.get("D_小数部桁数"), row.get("少数桁")),
                domain_cols["min_value"]: _first_value(row.get("D_最小値")),
                domain_cols["max_value"]: _first_value(row.get("D_最大値")),
                domain_cols["regex"]: _first_value(row.get("D_書式（正規表現）")),
                domain_cols["code_id"]: _first_value(row.get("D_参照外部コード")),
                "根拠": _first_value(row.get("備考"), row.get("判定理由"), row.get("理由")),
                "汎化候補": _first_value(row.get("汎化候補")),
            })

    if include_screen:
        add_screen_rows(screen_dedup_df)
    if include_table:
        add_table_rows(table_dedup_df)

    if not rows:
        return pd.DataFrame()
    proposal_cols = ["由来"] + [domain_cols[k] for k in domain_cols] + ["根拠", "汎化候補", "汎化候補（新規内）", "同義候補（新規内）"]
    proposal_df = pd.DataFrame(rows, columns=proposal_cols)
    proposal_df = proposal_df.drop_duplicates().reset_index(drop=True)
    llm_call = build_llm_call(cfg) if build_llm_call else None
    proposal_df = _group_new_domain_proposals(proposal_df, domain_cols["name"], llm_call)
    return proposal_df


def _apply_result_colors(ws, max_row: int) -> None:
    """判定結果列にセル色を適用する。"""
    from openpyxl.styles import PatternFill

    fills = {
        "完全一致": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "提案": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "選択必須": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "対象外": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    }
    header_row = [cell.value for cell in ws[1]]
    if "判定結果" not in header_row:
        return
    col_idx = header_row.index("判定結果") + 1
    for row_num in range(2, max_row + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        value = str(cell.value or "")
        # 詳細表示（"確定（...）" など）に対応
        fill = None
        for key, color in fills.items():
            if value.startswith(key):
                fill = color
                break
        if fill:
            cell.fill = fill
