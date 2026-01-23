#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ドメイン定義総合チェックツール（LLM補助）

機能1【ドメイン提案】: 対象一覧のドメイン指定をチェックし、適切なドメインを提案
機能2【整合性チェック】: ドメイン定義とテーブル定義の型・桁数の整合性をチェック
"""
from __future__ import annotations

import argparse
import concurrent.futures as cf
import json
import os
import re
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from dotenv import load_dotenv

try:
    from .config import get_domain_config
except ImportError:
    from config import get_domain_config

# ====== GUI（任意） ===========================================================
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

# ====== 定数 ==================================================================
DEFAULT_CONFIG: Dict[str, Any] = get_domain_config()

# ====== 正規化関数 ============================================================
def normalize_text(text: str) -> str:
    """NFKC正規化 + 小文字化 + 同義語統一"""
    if text is None or text == "":
        return ""
    s = unicodedata.normalize("NFKC", str(text)).lower().strip()
    s = re.sub(r"[\u3000\s]+", " ", s)

    # 同義語の統一（類似度向上のため）

    # 日付・時刻関連
    s = re.sub(r"取引日$", "日付", s)  # 取引日 → 日付
    s = re.sub(r"年月日$", "日付", s)  # 年月日 → 日付
    s = re.sub(r"(\w+)日$", r"\1日付", s)  # XX日 → XX日付
    s = re.sub(r"日時$", "日時", s)   # 日時（統一）
    s = re.sub(r"時刻$", "時刻", s)   # 時刻（統一）
    s = re.sub(r"タイムスタンプ$", "日時", s)  # タイムスタンプ → 日時

    # 金額・数量関連
    s = re.sub(r"金額$", "金額", s)   # 金額（統一）
    s = re.sub(r"価格$", "金額", s)   # 価格 → 金額
    s = re.sub(r"料金$", "金額", s)   # 料金 → 金額
    s = re.sub(r"単価$", "単価", s)   # 単価（統一）
    s = re.sub(r"数量$", "数量", s)   # 数量（統一）
    s = re.sub(r"件数$", "数量", s)   # 件数 → 数量
    s = re.sub(r"個数$", "数量", s)   # 個数 → 数量

    # コード・ID関連
    s = re.sub(r"コード$", "コード", s)  # コード（統一）
    s = re.sub(r"cd$", "コード", s)     # cd → コード
    s = re.sub(r"識別子$", "id", s)     # 識別子 → id
    s = re.sub(r"番号$", "番号", s)     # 番号（統一）
    s = re.sub(r"no$", "番号", s)       # no → 番号

    # 名称・名前関連
    s = re.sub(r"名称$", "名称", s)   # 名称（統一）
    s = re.sub(r"名前$", "名称", s)   # 名前 → 名称
    s = re.sub(r"氏名$", "名称", s)   # 氏名 → 名称
    s = re.sub(r"name$", "名称", s)   # name → 名称

    # 区分・種別関連
    s = re.sub(r"区分$", "区分", s)   # 区分（統一）
    s = re.sub(r"種別$", "区分", s)   # 種別 → 区分
    s = re.sub(r"タイプ$", "区分", s)  # タイプ → 区分
    s = re.sub(r"type$", "区分", s)   # type → 区分

    # フラグ・状態関連
    s = re.sub(r"フラグ$", "フラグ", s)  # フラグ（統一）
    s = re.sub(r"flag$", "フラグ", s)   # flag → フラグ
    s = re.sub(r"状態$", "状態", s)     # 状態（統一）
    s = re.sub(r"ステータス$", "状態", s)  # ステータス → 状態
    s = re.sub(r"status$", "状態", s)   # status → 状態

    # 備考・メモ関連
    s = re.sub(r"備考$", "備考", s)   # 備考（統一）
    s = re.sub(r"メモ$", "備考", s)   # メモ → 備考
    s = re.sub(r"コメント$", "備考", s)  # コメント → 備考
    s = re.sub(r"摘要$", "備考", s)   # 摘要 → 備考

    return s

def normalize_data_type(dtype: str) -> str:
    """データ型の正規化"""
    if not dtype:
        return ""
    dt = normalize_text(dtype)
    type_map = {
        "varchar": "varchar", "varchar2": "varchar", "char": "char",
        "int": "integer", "integer": "integer", "bigint": "bigint",
        "decimal": "decimal", "numeric": "decimal", "number": "decimal",
        "date": "date", "datetime": "datetime", "timestamp": "timestamp",
        "boolean": "boolean", "bool": "boolean"
    }
    for key, value in type_map.items():
        if key in dt:
            return value
    return dt

# ====== データクラス ==========================================================
@dataclass
class ScreenItem:
    """画面項目定義"""
    item_name: str          # 項目名称
    field_type: str         # フィールド
    length: Optional[str]   # 長さ
    format: str             # 編集形式
    required: str           # 必須
    row_number: int         # Excel行番号
    source_file: str
    source_sheet: str

@dataclass
class DomainDef:
    """ドメイン定義"""
    name: str                   # ドメイン名
    data_type: str              # データ型
    min_char: Optional[str]     # 最小文字数
    max_char: Optional[str]     # 最大文字数
    min_byte: Optional[str]     # 最小バイト数
    max_byte: Optional[str]     # 最大バイト数
    decimal: Optional[str]      # 小数部バイト数
    min_value: Optional[str]    # 最小値
    max_value: Optional[str]    # 最大値
    regex: str                  # 書式正規表現
    code_id: str                # 参照外部コードID
    row_number: int             # Excel行番号
    source_file: str
    source_sheet: str

# ====== Excel I/O =============================================================
HEADER_DETECT = os.getenv("HEADER_DETECT", "true").lower() != "false"
HEADER_SCAN_ROWS = int(os.getenv("HEADER_SCAN_ROWS", "10"))

def _pick_matching_sheets(xls: pd.ExcelFile, preferred: Optional[str]) -> List[str]:
    if not preferred:
        return [xls.sheet_names[0]]
    norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
    patterns = [p.strip() for p in preferred.split(",") if p.strip()]
    if not patterns:
        return [xls.sheet_names[0]]
    all_matches = []
    for pattern in patterns:
        want = norm(pattern)
        regex_pattern = want.replace('*', '.*')
        for name in xls.sheet_names:
            if re.match(regex_pattern, norm(name)):
                all_matches.append(name)
    result = []
    seen = set()
    for sheet in all_matches:
        if sheet not in seen:
            result.append(sheet)
            seen.add(sheet)
    return result if result else [xls.sheet_names[0]]

def _detect_header_row(path: Path, sheet_name: str, required_cols: List[str], scan_rows: int) -> int:
    head_df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows)
    req = {normalize_text(c) for c in required_cols if c}
    for i in range(len(head_df)):
        row_vals = {normalize_text(x) for x in head_df.iloc[i].values if str(x) not in {"", "nan"}}
        if req.issubset(row_vals):
            return i
    raise KeyError(f"必須列{sorted(required_cols)}を含むヘッダ行が見つかりません: {path.name}/{sheet_name}")

def read_excel_auto(path: Path, sheet_name: Optional[str], required_cols: List[str]) -> pd.DataFrame:
    if not HEADER_DETECT:
        return pd.read_excel(path, sheet_name=sheet_name)
    header_row = _detect_header_row(path, sheet_name, required_cols, HEADER_SCAN_ROWS)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row)


def read_excel_with_header_detection(path: Path, sheet_name: Optional[str], required_cols: List[str],
                                     explicit_header_row_1based: Optional[int] = None,
                                     scan_rows: int = 30) -> Tuple[pd.DataFrame, int]:
    """ヘッダ行が1行目とは限らないExcelに対応。"""
    if explicit_header_row_1based is not None:
        hdr0 = max(0, explicit_header_row_1based - 1)
        return pd.read_excel(path, sheet_name=sheet_name, header=hdr0), hdr0
    if not HEADER_DETECT:
        return pd.read_excel(path, sheet_name=sheet_name), 0
    # ヘッダ行の自動検出
    header_row = _detect_header_row(path, sheet_name, required_cols, scan_rows)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row), header_row

def load_screen_items(dir_path: Path, cfg: Dict[str, Any]) -> List[ScreenItem]:
    """画面項目定義を読み込み（行番号付き）"""
    files = sorted(dir_path.glob(cfg["SCREEN_GLOB"]))
    if not files:
        raise FileNotFoundError(f"画面項目定義ファイルが見つかりません")
    items = []
    for path in files:
        xls = pd.ExcelFile(path)
        sheets = _pick_matching_sheets(xls, cfg["SCREEN_SHEET"])
        for sheet in sheets:
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet,
                    [cfg["SCREEN_ITEM_COL"]],
                    None, 30
                )
                for idx, row in df.iterrows():
                    # 行全体が空かチェック
                    if row.isna().all():
                        continue

                    item_name = str(row.get(cfg["SCREEN_ITEM_COL"], "")).strip()

                    # 項目名が空の行はスキップ
                    if not item_name or item_name in ["nan", "None"]:
                        continue

                    row_number = header_row + idx + 2

                    # 値の取得とnan対策（列名で見つからない場合は列位置でフォールバック）
                    def get_value(col_name, col_idx=-1):
                        """
                        列名で取得を試み、失敗したら列位置インデックスで取得
                        col_idx: 列位置（0始まり）、-1の場合は列位置フォールバックなし
                        """
                        val = None

                        # まず列名で取得を試みる
                        if col_name in df.columns:
                            val = row.get(col_name, "")
                        # 列名が見つからず、列位置が指定されている場合
                        elif col_idx >= 0 and col_idx < len(row):
                            try:
                                val = row.iloc[col_idx]
                            except:
                                val = ""

                        if pd.isna(val):
                            return ""
                        val_str = str(val).strip()
                        # 半角スペースのみ、全角スペースのみ、nanなども空文字として扱う
                        if val_str in ["", "nan", "None", " ", "　"]:
                            return ""
                        # 完全に空白文字のみの場合も空文字として扱う
                        if not val_str or val_str.isspace():
                            return ""
                        return val_str

                    length_val = get_value(cfg["SCREEN_LENGTH_COL"], cfg["SCREEN_LENGTH_COL_IDX"])
                    field_type_val = get_value(cfg["SCREEN_TYPE_COL"], cfg["SCREEN_TYPE_COL_IDX"])
                    format_val = get_value(cfg["SCREEN_FORMAT_COL"], cfg["SCREEN_FORMAT_COL_IDX"])
                    required_val = get_value(cfg["SCREEN_REQUIRED_COL"], cfg["SCREEN_REQUIRED_COL_IDX"])

                    # デバッグログ（最初の5件のみ）
                    if len(items) < 5:
                        print(f"[DEBUG] 行{row_number}: 項目名={item_name}, フィールド={field_type_val}, 長さ={length_val}, 編集形式={format_val}, 必須={required_val}")

                    items.append(ScreenItem(
                        item_name=item_name,
                        field_type=field_type_val,
                        length=length_val if length_val else None,
                        format=format_val,
                        required=required_val,
                        row_number=row_number,
                        source_file=path.name,
                        source_sheet=sheet
                    ))
                print(f"[INFO] 画面項目定義読み込み: {path.name}/{sheet}")
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計画面項目: {len(items)}件")
    return items

def load_domains(dir_path: Path, cfg: Dict[str, Any]) -> Dict[str, DomainDef]:
    """ドメイン定義を読み込み（行番号付き）"""
    files = sorted(dir_path.glob(cfg["DOMAIN_GLOB"]))
    if not files:
        raise FileNotFoundError(f"ドメイン定義ファイルが見つかりません")
    domains = {}
    for path in files:
        xls = pd.ExcelFile(path)
        sheets = _pick_matching_sheets(xls, cfg["DOMAIN_SHEET"])
        for sheet in sheets:
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet,
                    [cfg["DOMAIN_NAME_COL"], cfg["DOMAIN_TYPE_COL"]],
                    None, 30
                )
                for idx, row in df.iterrows():
                    # 行全体が空かチェック
                    if row.isna().all():
                        continue

                    name = str(row.get(cfg["DOMAIN_NAME_COL"], "")).strip()

                    # ドメイン名が空の行はスキップ
                    if not name or name in ["nan", "None"]:
                        continue

                    row_number = header_row + idx + 2

                    # 値の取得とnan対策（列名で見つからない場合は列位置でフォールバック）
                    def get_value(col_name, col_idx=-1):
                        """
                        列名で取得を試み、失敗したら列位置インデックスで取得
                        col_idx: 列位置（0始まり）、-1の場合は列位置フォールバックなし
                        """
                        val = None

                        # まず列名で取得を試みる
                        if col_name in df.columns:
                            val = row.get(col_name, "")
                        # 列名が見つからず、列位置が指定されている場合
                        elif col_idx >= 0 and col_idx < len(row):
                            try:
                                val = row.iloc[col_idx]
                            except:
                                val = ""

                        if pd.isna(val):
                            return ""
                        val_str = str(val).strip()
                        # 半角スペースのみ、全角スペースのみ、nanなども空文字として扱う
                        if val_str in ["", "nan", "None", " ", "　"]:
                            return ""
                        # 完全に空白文字のみの場合も空文字として扱う
                        if not val_str or val_str.isspace():
                            return ""
                        return val_str

                    domains[normalize_text(name)] = DomainDef(
                        name=name,
                        data_type=get_value(cfg["DOMAIN_TYPE_COL"], cfg["DOMAIN_TYPE_COL_IDX"]),
                        min_char=get_value(cfg["DOMAIN_MIN_CHAR_COL"], cfg["DOMAIN_MIN_CHAR_COL_IDX"]) or None,
                        max_char=get_value(cfg["DOMAIN_MAX_CHAR_COL"], cfg["DOMAIN_MAX_CHAR_COL_IDX"]) or None,
                        min_byte=get_value(cfg["DOMAIN_MIN_BYTE_COL"], cfg["DOMAIN_MIN_BYTE_COL_IDX"]) or None,
                        max_byte=get_value(cfg["DOMAIN_MAX_BYTE_COL"], cfg["DOMAIN_MAX_BYTE_COL_IDX"]) or None,
                        decimal=get_value(cfg["DOMAIN_DECIMAL_COL"], cfg["DOMAIN_DECIMAL_COL_IDX"]) or None,
                        min_value=get_value(cfg["DOMAIN_MIN_VALUE_COL"], cfg["DOMAIN_MIN_VALUE_COL_IDX"]) or None,
                        max_value=get_value(cfg["DOMAIN_MAX_VALUE_COL"], cfg["DOMAIN_MAX_VALUE_COL_IDX"]) or None,
                        regex=get_value(cfg["DOMAIN_REGEX_COL"], cfg["DOMAIN_REGEX_COL_IDX"]),
                        code_id=get_value(cfg["DOMAIN_CODE_ID_COL"], cfg["DOMAIN_CODE_ID_COL_IDX"]),
                        row_number=row_number,
                        source_file=path.name,
                        source_sheet=sheet
                    )
                print(f"[INFO] ドメイン定義読み込み: {path.name}/{sheet} ({len(domains)}件)")
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計ドメイン定義: {len(domains)}件")
    return domains

# ====== 類似度計算 =============================================================
try:
    from rapidfuzz import fuzz
    FUZZ_AVAILABLE = True
except ImportError:
    from difflib import SequenceMatcher
    FUZZ_AVAILABLE = False

def calc_similarity(s1: str, s2: str) -> float:
    """2つの文字列の類似度を計算（0.0-1.0）"""
    if FUZZ_AVAILABLE:
        return fuzz.ratio(s1, s2) / 100.0
    else:
        return SequenceMatcher(None, s1, s2).ratio()

# ====== 突合処理 ===============================================================
def match_item_with_domains(item: ScreenItem, domains: Dict[str, DomainDef],
                            cfg: Dict[str, Any]) -> Tuple[str, Optional[DomainDef], float, str]:
    """
    画面項目とドメインを突合

    Returns:
        (match_type, matched_domain, score, reason)
        match_type: "完全一致", "類似一致", "不一致（提案必要）", "不一致（チェック不要）"
    """
    item_name_norm = normalize_text(item.item_name)
    threshold = cfg["FUZZY_THRESHOLD"]

    # 完全一致チェック
    if item_name_norm in domains:
        matched_domain = domains[item_name_norm]
        reason = f"項目名「{item.item_name}」がドメイン「{matched_domain.name}」と完全に一致しました"
        return ("完全一致", matched_domain, 1.0, reason)

    # 類似一致チェック
    best_score = 0.0
    best_domain = None
    for domain_key, domain in domains.items():
        score = calc_similarity(item_name_norm, domain_key)
        if score > best_score:
            best_score = score
            best_domain = domain

    if best_score >= threshold:
        reason = f"項目名「{item.item_name}」がドメイン「{best_domain.name}」と類似しています（類似度: {best_score:.1%}）"
        return ("類似一致", best_domain, best_score, reason)

    # 不一致の場合、項目桁数や編集形式をチェック
    has_length = False
    has_format = False

    if item.length is not None:
        length_str = str(item.length).strip()
        has_length = length_str and length_str not in ["", "nan", "None"]

    if item.format:
        format_str = str(item.format).strip()
        has_format = format_str and format_str not in ["", "nan", "None"]

    if has_length or has_format:
        details = []
        if has_length:
            details.append(f"長さ制約: {item.length}")
        if has_format:
            details.append(f"編集形式: {item.format}")
        reason = f"一致するドメインが見つかりませんでしたが、画面項目定義書に{', '.join(details)}が定義されているため、新規ドメインの提案が必要です"
        return ("不一致（提案必要）", None, 0.0, reason)
    else:
        reason = f"一致するドメインが見つかりませんでした。画面項目定義書に長さや編集形式の定義がないため、ドメイン設定は不要と判断されます"
        return ("不一致（チェック不要）", None, 0.0, reason)

def suggest_domain_from_screen_item(item: ScreenItem) -> Dict[str, str]:
    """画面項目定義からドメイン提案を生成"""
    suggestions = {}

    # フィールドタイプからデータ型を推測
    if item.field_type:
        field_type_lower = str(item.field_type).lower().strip()
        if "text" in field_type_lower or "string" in field_type_lower or "varchar" in field_type_lower:
            suggestions["データ型"] = "VARCHAR"
        elif "number" in field_type_lower or "integer" in field_type_lower or "int" in field_type_lower:
            suggestions["データ型"] = "INTEGER"
        elif "date" in field_type_lower:
            suggestions["データ型"] = "DATE"
        elif "decimal" in field_type_lower or "float" in field_type_lower or "numeric" in field_type_lower:
            suggestions["データ型"] = "DECIMAL"
        elif "boolean" in field_type_lower or "bool" in field_type_lower:
            suggestions["データ型"] = "BOOLEAN"
        else:
            suggestions["データ型"] = item.field_type

    # 長さから最大文字数/バイト数を設定
    if item.length is not None:
        length_value = str(item.length).strip()
        if length_value and length_value not in ["", "nan", "None"]:
            suggestions["最大文字数"] = length_value
            suggestions["最大バイト数"] = length_value

    # 編集形式から書式正規表現を設定
    if item.format:
        format_value = str(item.format).strip()
        if format_value and format_value not in ["", "nan", "None"]:
            suggestions["書式正規表現"] = format_value

    return suggestions

def process_screen_domain_matching(screen_items: List[ScreenItem],
                                   domains: Dict[str, DomainDef],
                                   cfg: Dict[str, Any],
                                   progress_callback=None) -> pd.DataFrame:
    """
    画面項目とドメインの突合処理

    Args:
        screen_items: 画面項目リスト
        domains: ドメイン定義辞書
        cfg: 設定
        progress_callback: 進捗コールバック関数 (processed_count, total_items)
    """
    results = []
    total_items = len(screen_items)
    processed_count = 0

    for item in screen_items:
        match_type, matched_domain, score, reason = match_item_with_domains(item, domains, cfg)

        # 不一致（提案必要）の場合、画面項目定義から借り入れ
        if match_type == "不一致（提案必要）":
            suggestions = suggest_domain_from_screen_item(item)
            result = {
                "元ファイル": item.source_file,
                "元シート": item.source_sheet,
                "行番号": item.row_number,
                "項目名称": item.item_name,
                "判定結果": match_type,
                "理由": reason,
                "一致ドメイン名": "",
                "類似度": "",
                "フィールド": item.field_type,
                "長さ": item.length,
                "編集形式": item.format,
                "必須": item.required,
                "データ型": suggestions.get("データ型", ""),
                "最小文字数": "",
                "最大文字数": suggestions.get("最大文字数", ""),
                "最小バイト数": "",
                "最大バイト数": suggestions.get("最大バイト数", ""),
                "小数部バイト数": "",
                "最小値": "",
                "最大値": "",
                "書式正規表現": suggestions.get("書式正規表現", ""),
                "参照外部コードID": "",
            }
        else:
            # 一致した場合はドメイン情報を使用
            # Noneの場合は空文字列に変換
            def get_domain_value(val):
                if val is None or pd.isna(val):
                    return ""
                return str(val).strip()

            result = {
                "元ファイル": item.source_file,
                "元シート": item.source_sheet,
                "行番号": item.row_number,
                "項目名称": item.item_name,
                "判定結果": match_type,
                "理由": reason,
                "一致ドメイン名": matched_domain.name if matched_domain else "",
                "類似度": f"{score:.2%}" if score > 0 else "",
                "フィールド": item.field_type,
                "長さ": item.length if item.length else "",
                "編集形式": item.format if item.format else "",
                "必須": item.required if item.required else "",
                "データ型": get_domain_value(matched_domain.data_type) if matched_domain else "",
                "最小文字数": get_domain_value(matched_domain.min_char) if matched_domain else "",
                "最大文字数": get_domain_value(matched_domain.max_char) if matched_domain else "",
                "最小バイト数": get_domain_value(matched_domain.min_byte) if matched_domain else "",
                "最大バイト数": get_domain_value(matched_domain.max_byte) if matched_domain else "",
                "小数部バイト数": get_domain_value(matched_domain.decimal) if matched_domain else "",
                "最小値": get_domain_value(matched_domain.min_value) if matched_domain else "",
                "最大値": get_domain_value(matched_domain.max_value) if matched_domain else "",
                "書式正規表現": get_domain_value(matched_domain.regex) if matched_domain else "",
                "参照外部コードID": get_domain_value(matched_domain.code_id) if matched_domain else "",
            }
        results.append(result)

        # 進捗更新
        processed_count += 1
        pct = processed_count * 100 / total_items if total_items else 0
        if processed_count % 10 == 0 or processed_count == total_items:
            print(f"[INFO] {processed_count}/{total_items} 件処理済み ({pct:.1f}%)")
        # コールバック呼び出し（リアルタイム進捗）
        if progress_callback:
            progress_callback(processed_count, total_items)

    df = pd.DataFrame(results)

    # NaNや"nan"文字列を空文字列に置換
    df = df.fillna("")
    df = df.replace("nan", "")
    df = df.replace("None", "")

    print(f"[DEBUG] DataFrame作成完了: {len(df)}行, 列名: {list(df.columns)}")
    return df

# ====== 旧関数（削除予定） =====================================================
def load_tables(dir_path: Path, cfg: Dict[str, Any]) -> Dict[str, Any]:
    """テーブル定義を読み込み（項目名でキー、行番号付き）"""
    files = sorted(dir_path.glob(cfg["TABLE_GLOB"]))
    if not files:
        raise FileNotFoundError(f"テーブル定義ファイルが見つかりません")
    tables = {}
    for path in files:
        xls = pd.ExcelFile(path)
        sheets = _pick_matching_sheets(xls, cfg["TABLE_SHEET"])
        for sheet in sheets:
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet,
                    [cfg["TABLE_NAME_COL"], cfg["TABLE_ITEM_COL"], cfg["TABLE_COLUMN_COL"]],
                    None, 30
                )
                for idx, row in df.iterrows():
                    table_name = str(row.get(cfg["TABLE_NAME_COL"], "")).strip()
                    item_name = str(row.get(cfg["TABLE_ITEM_COL"], "")).strip()
                    column_name = str(row.get(cfg["TABLE_COLUMN_COL"], "")).strip()
                    if table_name and item_name and column_name:
                        row_number = header_row + idx + 2
                        # キーは項目名（論理名）のみ
                        key = normalize_text(item_name)
                        tables[key] = TableDef(
                            table_name=table_name,
                            item_name=item_name,
                            column_name=column_name,
                            data_type=str(row.get(cfg["TABLE_TYPE_COL"], "")).strip() if cfg["TABLE_TYPE_COL"] in df.columns else "",
                            length=str(row.get(cfg["TABLE_LENGTH_COL"], "")).strip() if cfg["TABLE_LENGTH_COL"] in df.columns else None,
                            row_number=row_number,
                            source_file=path.name,
                            source_sheet=sheet
                        )
                print(f"[INFO] テーブル定義読み込み: {path.name}/{sheet}")
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計テーブルカラム: {len(tables)}件")
    return tables

def load_targets(dir_path: Path, cfg: Dict[str, Any]) -> List[TargetItem]:
    """対象一覧を読み込み（項目名のみ、行番号付き）"""
    files = sorted(dir_path.glob(cfg["TARGET_GLOB"]))
    if not files:
        print(f"[警告] 対象一覧ファイルが見つかりません")
        return []
    items = []
    for path in files:
        xls = pd.ExcelFile(path)
        sheets = _pick_matching_sheets(xls, cfg["TARGET_SHEET"])
        for sheet in sheets:
            try:
                df, header_row = read_excel_with_header_detection(
                    path, sheet,
                    [cfg["TARGET_ITEM_COL"]],
                    None, 30
                )
                for idx, row in df.iterrows():
                    item_name = str(row.get(cfg["TARGET_ITEM_COL"], "")).strip()
                    if item_name:
                        row_number = header_row + idx + 2
                        items.append(TargetItem(
                            item_name=item_name,
                            row_number=row_number,
                            source_file=path.name,
                            source_sheet=sheet
                        ))
                print(f"[INFO] 対象一覧読み込み: {path.name}/{sheet}")
            except Exception as e:
                print(f"[警告] {path.name}({sheet}) エラー: {e}")
    print(f"[INFO] 合計対象項目: {len(items)}件")
    return items

# ====== APIクライアント =======================================================
class ApiClient:
    def __init__(self, cfg: Dict[str, Any]):
        self.base_url = cfg["OPENAI_BASE_URL"].rstrip("/")
        self.path = cfg["OPENAI_PATH"]
        self.timeout = cfg["TIMEOUT_SEC"]
        self.verify = cfg["VERIFY_SSL"]
        self.session = requests.Session()
        proxies: Dict[str, str] = {}
        if cfg.get("HTTP_PROXY"):
            proxies["http"] = cfg["HTTP_PROXY"]
        if cfg.get("HTTPS_PROXY"):
            proxies["https"] = cfg["HTTPS_PROXY"]
        if proxies:
            self.session.proxies.update(proxies)
        headers: Dict[str, str] = {"Content-Type": "application/json"}
        if cfg.get("OPENAI_SEND_AUTH") and cfg.get("OPENAI_API_KEY"):
            headers["Authorization"] = f"Bearer {cfg['OPENAI_API_KEY']}"
        extra = cfg.get("OPENAI_HEADERS_JSON")
        if extra:
            try:
                headers.update(json.loads(extra))
            except:
                pass
        self.headers = headers

    def post_json(self, body: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{self.base_url}{self.path}"
        resp = self.session.post(url, headers=self.headers, json=body, timeout=self.timeout, verify=self.verify)
        resp.raise_for_status()
        return resp.json()

# ====== LLMプロンプト（機能1: ドメイン提案） ==================================
LLM_SUGGESTION_SYSTEM = (
    "あなたはデータベース設計の専門家です。\n"
    "項目名の業務的意味を分析し、判断材料を提供します（推奨はしません）。\n"
    "\n"
    "重要な分析視点:\n"
    "1. 業務的意味: 項目名が表す業務上の概念を特定\n"
    "2. 技術的項目の判定: 業務的意味を持たない技術的カラムかどうか\n"
    "3. 類似性分析: 既存ドメインとの意味的な類似性を評価\n"
    "4. 特性分析: この項目に必要な特性（バリデーション、制約等）を抽出\n"
    "\n"
    "技術的項目の例:\n"
    "- id, created_at, updated_at, version, seq_no\n"
    "- delete_flag, is_active, enabled\n"
    "- lock_version, row_version\n"
    "\n"
    "注意: 同じデータ型・桁数でも、意味が異なれば別ドメイン（例: 顧客コード vs 商品コード）\n"
)

LLM_SUGGESTION_USER = """対象項目: {item_name}
テーブル: {table_name}.{column_name}
テーブル定義: データ型={table_type}, 桁数={table_length}

既存ドメイン定義一覧:
{domain_candidates}

タスク:
以下の分析を実施し、人間が判断するための材料を提供してください。

1. 業務的意味の分析
2. 技術的項目かどうかの判定
3. 既存ドメインとの類似性分析（意味的な類似性）
4. この項目に必要な特性の抽出

出力JSON:
{{
  "business_meaning": string,
  "is_technical": boolean,
  "similar_domains": [string],
  "required_characteristics": string,
  "notes": string
}}

フィールド説明:
- business_meaning: 項目の業務的意味を簡潔に説明（例: "顧客を一意に識別するコード"）
  技術的項目の場合も説明を記載（例: "レコードの作成日時を保持する技術的カラム"）
- is_technical: 技術的項目=true、業務的項目=false
- similar_domains: 既存ドメインの中で意味的に類似するものをリスト（最大3件）
  完全一致がなくても、関連性があるものを含める
  例: ["顧客コード", "取引先コード"]
- required_characteristics: この項目に必要と思われる特性（バリデーション、制約等）
  例: "英数字8桁固定、ユニーク制約"
- notes: その他の補足情報や注意点

JSON以外出力禁止。
"""

# ====== LLMプロンプト（機能2: 整合性チェック） ================================
LLM_VALIDATION_SYSTEM = (
    "あなたはデータベース設計の専門家です。\n"
    "ドメイン定義とテーブル定義の差異を分析し、重要度を評価します。\n"
)

LLM_VALIDATION_USER = """テーブル: {table_name}.{column_name}
指定ドメイン: {domain_name}

ドメイン定義: 型={domain_type}, 桁数={domain_length}, 精度={domain_precision}, スケール={domain_scale}
テーブル定義: 型={table_type}, 桁数={table_length}, 精度={table_precision}, スケール={table_scale}

差異: {differences}

出力JSON:
{{
  "severity": "critical" | "warning" | "info" | "acceptable",
  "is_valid": true | false,
  "reason": string,
  "recommendation": string
}}

重要度:
- critical: データ損失の危険
- warning: 業務上問題の可能性
- info: 統一推奨だが問題なし
- acceptable: 実質的に同じ

JSON以外出力禁止。
"""

def call_llm(prompt_system: str, prompt_user: str, cfg: Dict[str, Any],
             client: ApiClient, api_semaphore: threading.Semaphore) -> Dict[str, Any]:
    payload = {
        "model": cfg["OPENAI_MODEL"],
        "max_completion_tokens": cfg["MAX_TOKENS"],  # GPT-5 Cline Proxy用
        "n": 1,
        "reasoning_effort": "high",
        "verbosity": "high",
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": prompt_system},
            {"role": "user", "content": prompt_user}
        ]
    }
    if api_semaphore:
        api_semaphore.acquire()
    try:
        for attempt in range(cfg["RETRY"] + 1):
            try:
                data = client.post_json(payload)
                content = data["choices"][0]["message"]["content"]
                return json.loads(content)
            except Exception as e:
                if attempt < cfg["RETRY"]:
                    time.sleep(1.2 * (attempt + 1))
                    continue
                return {"error": str(e)}
    finally:
        if api_semaphore:
            api_semaphore.release()

# ====== 処理関数 ==============================================================
def compare_spec(domain: DomainDef, table: TableDef) -> Tuple[bool, List[str]]:
    """ドメイン定義とテーブル定義を比較"""
    type_match = normalize_data_type(domain.data_type) == normalize_data_type(table.data_type)
    diffs = []
    if not type_match:
        diffs.append(f"データ型: {domain.data_type} vs {table.data_type}")

    # 桁数比較
    if domain.length and table.length:
        if normalize_text(domain.length) != normalize_text(table.length):
            diffs.append(f"桁数: {domain.length} vs {table.length}")

    # 精度・スケール比較
    if domain.precision and table.precision:
        if normalize_text(domain.precision) != normalize_text(table.precision):
            diffs.append(f"精度: {domain.precision} vs {table.precision}")
    if domain.scale and table.scale:
        if normalize_text(domain.scale) != normalize_text(table.scale):
            diffs.append(f"スケール: {domain.scale} vs {table.scale}")

    return (type_match and len(diffs) == 0), diffs


def analyze_compatibility(table_def: TableDef, domains: Dict[str, DomainDef]) -> Tuple[List[str], str]:
    """テーブル定義と各ドメインの整合性を分析（提案時の参考情報）"""
    domain_list = []
    analysis_notes = []

    table_dtype_norm = normalize_data_type(table_def.data_type)

    # データ型一致のドメインを優先、その後は全ドメイン
    type_matched = []
    type_mismatched = []

    for domain_name, domain in domains.items():
        domain_dtype_norm = normalize_data_type(domain.data_type)
        is_match, diffs = compare_spec(domain, table_def)

        domain_info = f"{domain.name} ({domain.data_type}"
        if domain.length:
            domain_info += f", {domain.length}桁"
        domain_info += ")"

        if domain_dtype_norm == table_dtype_norm:
            # データ型一致
            if is_match:
                # 完全一致
                domain_info += " - 完全一致 ✓"
                analysis_notes.append(f"✓ {domain.name}: データ型・桁数完全一致")
                type_matched.insert(0, domain_info)  # 先頭に追加
            elif len(diffs) == 1 and "桁数" in diffs[0]:
                # 型一致、桁数のみ差異
                domain_info += " - 型一致、桁数差異"
                analysis_notes.append(f"△ {domain.name}: 型一致、{diffs[0]}")
                type_matched.append(domain_info)
            else:
                # 型一致、その他差異
                domain_info += " - 型一致"
                type_matched.append(domain_info)
        else:
            # データ型不一致（参考として含める）
            domain_info += " - 型不一致"
            type_mismatched.append(domain_info)

    # データ型一致を優先、その後全ドメイン
    domain_list = type_matched + type_mismatched

    # 上位候補のみ返却（LLMへの入力を減らす）
    top_domains = domain_list[:30]  # 上位30件（意味重視なので多めに）
    analysis_text = "\n".join(analysis_notes[:15]) if analysis_notes else "データ型が一致するドメインがありません（新規定義を推奨）"

    return (top_domains, analysis_text)


def process_domain_suggestion(targets: List[TargetItem], domains: Dict[str, DomainDef],
                              tables: Dict[str, TableDef],
                              cfg: Dict[str, Any], client: ApiClient,
                              api_semaphore: threading.Semaphore) -> pd.DataFrame:
    """
    ドメイン提案機能（新仕様）

    処理フロー:
    1. 対象一覧から項目名を読み込み
    2. ドメイン定義で完全一致チェック → 一致あれば採用して終了
    3. テーブル定義から該当項目を検索
    4. LLMで判断材料を提供（提案はしない）
    5. 結果を出力（人間が最終判断）
    """
    rows = []
    for item in targets:
        # Step 1: 項目名で正規化キーを作成
        item_key = normalize_text(item.item_name)

        # Step 2: ドメイン定義で完全一致チェック
        if item_key in domains:
            matched_domain = domains[item_key]
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "判定結果": "完全一致",
                "一致ドメイン": matched_domain.name,
                "ドメイン行番号": matched_domain.row_number,
                "ドメインファイル": matched_domain.source_file,
                "データ型": matched_domain.data_type,
                "桁数": matched_domain.length or "-",
                "バリデーション": matched_domain.validation,
                "備考": "ドメイン定義に完全一致"
            })
            continue

        # Step 3: テーブル定義から該当項目を検索
        table_def = tables.get(item_key)
        if not table_def:
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "判定結果": "エラー",
                "備考": "テーブル定義が見つかりません"
            })
            continue

        # Step 4: LLMで判断材料を提供
        # すべてのドメインを候補として渡す（型でフィルタしない）
        domain_list = []
        for domain in domains.values():
            domain_list.append(
                f"- {domain.name}: {domain.data_type}({domain.length or '-'}), バリデーション: {domain.validation}"
            )

        llm_result = call_llm(
            LLM_SUGGESTION_SYSTEM,
            LLM_SUGGESTION_USER.format(
                item_name=item.item_name,
                table_name=table_def.table_name,
                column_name=table_def.column_name,
                table_type=table_def.data_type,
                table_length=table_def.length or "未指定",
                domain_candidates="\n".join(domain_list) if domain_list else "(ドメイン候補なし)"
            ),
            cfg, client, api_semaphore
        )

        if "error" in llm_result:
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "テーブル名": table_def.table_name,
                "カラム名": table_def.column_name,
                "テーブル行番号": table_def.row_number,
                "データ型": table_def.data_type,
                "桁数": table_def.length or "-",
                "判定結果": "LLMエラー",
                "備考": llm_result.get("error", "不明なエラー")
            })
        else:
            # Step 5: LLM分析結果を出力（人間が最終判断）
            rows.append({
                "対象行番号": item.row_number,
                "対象ファイル": item.source_file,
                "項目名": item.item_name,
                "テーブル名": table_def.table_name,
                "カラム名": table_def.column_name,
                "テーブル行番号": table_def.row_number,
                "テーブルファイル": table_def.source_file,
                "データ型": table_def.data_type,
                "桁数": table_def.length or "-",
                "判定結果": "要判断",
                "業務的意味": llm_result.get("business_meaning", ""),
                "類似ドメイン": ", ".join(llm_result.get("similar_domains", [])),
                "必要な特性": llm_result.get("required_characteristics", ""),
                "技術的項目判定": "はい" if llm_result.get("is_technical", False) else "いいえ",
                "備考": llm_result.get("notes", "")
            })

    return pd.DataFrame(rows)

def process_domain_validation(tables: Dict[Tuple[str, str], TableDef],
                              domains: Dict[str, DomainDef],
                              cfg: Dict[str, Any], client: ApiClient,
                              api_semaphore: threading.Semaphore) -> pd.DataFrame:
    """機能2: 整合性チェック"""
    rows = []
    for (table_key), table_def in tables.items():
        if not table_def.domain:
            continue

        domain_key = normalize_text(table_def.domain)
        domain = domains.get(domain_key)

        if not domain:
            rows.append({
                "table_name": table_def.table_name,
                "column_name": table_def.column_name,
                "domain_name": table_def.domain,
                "check_result": "エラー",
                "severity": "critical",
                "reason": "ドメイン定義が存在しません",
                "recommendation": "ドメイン定義を追加"
            })
            continue

        is_match, diffs = compare_spec(domain, table_def)

        if is_match:
            rows.append({
                "table_name": table_def.table_name,
                "column_name": table_def.column_name,
                "domain_name": domain.name,
                "check_result": "OK",
                "severity": "acceptable",
                "domain_type": domain.data_type,
                "table_type": table_def.data_type,
                "reason": "完全一致",
                "recommendation": "問題なし"
            })
        else:
            # LLM判定
            llm_result = call_llm(
                LLM_VALIDATION_SYSTEM,
                LLM_VALIDATION_USER.format(
                    table_name=table_def.table_name,
                    column_name=table_def.column_name,
                    domain_name=domain.name,
                    domain_type=domain.data_type,
                    domain_length=domain.length or "未指定",
                    domain_precision=domain.precision or "未指定",
                    domain_scale=domain.scale or "未指定",
                    table_type=table_def.data_type,
                    table_length=table_def.length or "未指定",
                    table_precision=table_def.precision or "未指定",
                    table_scale=table_def.scale or "未指定",
                    differences="\n".join(diffs)
                ),
                cfg, client, api_semaphore
            )
            if "error" not in llm_result:
                rows.append({
                    "table_name": table_def.table_name,
                    "column_name": table_def.column_name,
                    "domain_name": domain.name,
                    "check_result": "OK" if llm_result.get("is_valid") else "不一致",
                    "severity": llm_result.get("severity", "warning"),
                    "domain_type": domain.data_type,
                    "table_type": table_def.data_type,
                    "reason": llm_result.get("reason", ""),
                    "recommendation": llm_result.get("recommendation", "")
                })
    return pd.DataFrame(rows)

# ====== 出力 ==================================================================
def save_outputs(df_result: pd.DataFrame, cfg: Dict[str, Any]) -> None:
    """結果をExcel形式で保存"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule

    out_dir = Path(cfg["OUT_DIR"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    output_file = out_dir / "screen_domain_check.xlsx"

    # Excelに保存
    df_result.to_excel(output_file, sheet_name="ドメインチェック結果", index=False, engine="openpyxl")

    # 条件付き書式を追加
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb["ドメインチェック結果"]

    # 判定結果列に色付け
    if "判定結果" in df_result.columns:
        result_col_idx = list(df_result.columns).index("判定結果") + 1
        result_col = get_column_letter(result_col_idx)

        fill_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        fill_gray = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        for row in range(2, len(df_result) + 2):
            cell = ws[f"{result_col}{row}"]
            if cell.value == "完全一致":
                cell.fill = fill_green
            elif cell.value == "類似一致":
                cell.fill = fill_yellow
            elif cell.value == "不一致（チェック不要）":
                cell.fill = fill_gray
            elif cell.value == "不一致（提案必要）":
                cell.fill = fill_red

    wb.save(output_file)
    print(f"[INFO] 保存完了: {output_file}")

# ====== CLI ===================================================================
def app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

def ask_directory(title: str) -> Optional[str]:
    if not TK_AVAILABLE:
        return None
    try:
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askdirectory(title=title)
        root.destroy()
        return path or None
    except:
        return None

def main() -> None:
    load_dotenv()

    parser = argparse.ArgumentParser(description="画面項目ドメインチェックツール")
    parser.add_argument("--dir", help="入力ディレクトリ")
    parser.add_argument("--out-dir", help="出力ディレクトリ")
    parser.add_argument("--no-gui", action="store_true")
    args = parser.parse_args()

    cfg = DEFAULT_CONFIG.copy()

    in_dir = args.dir
    if not in_dir and not args.no_gui:
        in_dir = ask_directory("入力ディレクトリを選択")
    if not in_dir:
        print("入力ディレクトリが指定されていません")
        return

    if args.out_dir:
        cfg["OUT_DIR"] = args.out_dir

    root_dir = Path(in_dir)
    if not root_dir.exists():
        print(f"ディレクトリが存在しません: {root_dir}")
        return

    print("[INFO] 画面項目ドメインチェック処理を開始します...")

    # データ読み込み
    print("\n[ステップ1] データ読み込み")
    screen_items = load_screen_items(root_dir, cfg)
    domains = load_domains(root_dir, cfg)

    # 突合処理
    print("\n[ステップ2] 画面項目とドメインの突合")
    df_result = process_screen_domain_matching(screen_items, domains, cfg)
    print(f"[DEBUG] df_result type: {type(df_result)}, shape: {df_result.shape if hasattr(df_result, 'shape') else 'N/A'}")
    if len(df_result) > 0:
        print(f"[DEBUG] df_result columns: {list(df_result.columns)}")

    # 結果保存
    print("\n[ステップ3] 結果保存")
    save_outputs(df_result, cfg)

    # サマリ表示
    print("\n[処理完了]")
    print(f"  総項目数: {len(df_result)}件")
    if '判定結果' in df_result.columns:
        print(f"  完全一致: {len(df_result[df_result['判定結果'] == '完全一致'])}件")
    else:
        print(f"[ERROR] '判定結果' 列が存在しません。列名: {list(df_result.columns)}")
        print(f"  完全一致: 集計不可")
        print(f"  類似一致: {len(df_result[df_result['判定結果'] == '類似一致'])}件")
        print(f"  不一致（提案必要）: {len(df_result[df_result['判定結果'] == '不一致（提案必要）'])}件")
        print(f"  不一致（チェック不要）: {len(df_result[df_result['判定結果'] == '不一致（チェック不要）'])}件")

    if TK_AVAILABLE and not args.no_gui:
        try:
            messagebox.showinfo("完了", f"処理完了\n保存先: {Path(cfg['OUT_DIR']).resolve()}")
        except:
            pass

if __name__ == "__main__":
    main()
