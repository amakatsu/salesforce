from __future__ import annotations

import concurrent.futures as cf
import difflib
import json
import re
import threading
from dataclasses import dataclass
from pathlib import Path
from string import Template
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .config import HARD_EXACT_SCORE
from .llm import ApiClient, LLM_SYSTEM, LLM_USER_TEMPLATE, call_llm


def zenkaku_hankaku_norm(text: str) -> str:
    if text is None:
        return ""
    s = re.sub(r"[　\s]+", " ", str(text).strip().lower())
    s = re.sub(r"[\-/・·•･,()\[\]_]+", " ", s)
    return re.sub(r"\s+", " ", s)


def local_similarity(a: str, b: str) -> float:
    a_n, b_n = zenkaku_hankaku_norm(a), zenkaku_hankaku_norm(b)
    if not a_n or not b_n:
        return 0.0
    if a_n == b_n:
        return 1.0
    if a_n in b_n or b_n in a_n:
        return 0.9
    return difflib.SequenceMatcher(None, a_n, b_n).ratio()


@dataclass
class Candidate:
    term: str
    score: float


@dataclass
class ComponentMatchResult:
    matched_terms: List[str]
    matched_norms: List[str]
    unmatched_segments: List[str]
    coverage_ratio: float
    unmatched_count: int
    has_non_digit: bool
    digit_segments: List[str]


class ComponentMatcher:
    def __init__(self, norm_to_term: Dict[str, str]):
        trimmed: Dict[str, str] = {}
        for norm_value, original in norm_to_term.items():
            key = norm_value.replace(" ", "")
            if not key:
                continue
            trimmed.setdefault(key, original)
        self.norm_to_term = trimmed
        self.prefix_map: Dict[str, List[str]] = {}
        for norm_value in self.norm_to_term.keys():
            first = norm_value[0]
            self.prefix_map.setdefault(first, []).append(norm_value)
        for values in self.prefix_map.values():
            values.sort(key=len, reverse=True)
        self.nondigit_len: Dict[str, int] = {
            key: sum(1 for ch in key if not ch.isdigit()) for key in self.norm_to_term.keys()
        }

    def analyze(self, screen_name: str) -> ComponentMatchResult:
        normalized = zenkaku_hankaku_norm(screen_name)
        flat = normalized.replace(" ", "")
        if not flat:
            return ComponentMatchResult([], [], [], 1.0, 0, False, [])
        total_non_digit = sum(1 for ch in flat if not ch.isdigit())
        has_non_digit = total_non_digit > 0

        from functools import lru_cache

        @lru_cache(maxsize=None)
        def walk(idx: int):
            if idx >= len(flat):
                return (0, 0, 0, [])
            best = None
            ch = flat[idx]
            if ch.isdigit():
                j = idx
                while j < len(flat) and flat[j].isdigit():
                    j += 1
                nxt = walk(j)
                if nxt is not None:
                    candidate = (nxt[0], nxt[1], nxt[2], [("digit", flat[idx:j], None)] + nxt[3])
                    best = self._better(best, candidate)
            if ch in self.prefix_map:
                for norm_term in self.prefix_map[ch]:
                    end = idx + len(norm_term)
                    if flat.startswith(norm_term, idx):
                        nxt = walk(end)
                        if nxt is None:
                            continue
                        term_non_digit = self.nondigit_len.get(norm_term, 0)
                        candidate = (
                            nxt[0],
                            nxt[1] + term_non_digit,
                            nxt[2] + 1,
                            [("term", norm_term, self.norm_to_term.get(norm_term, norm_term))] + nxt[3],
                        )
                        best = self._better(best, candidate)
            if not ch.isdigit():
                nxt = walk(idx + 1)
                if nxt is not None:
                    candidate = (nxt[0] + 1, nxt[1], nxt[2], [("unmatched", ch, None)] + nxt[3])
                    best = self._better(best, candidate)
            return best

        result = walk(0)
        if result is None:
            return ComponentMatchResult([], [], [], 0.0, 0, has_non_digit, [])
        unmatched_len, matched_len, matched_count, parts = result
        matched_terms: List[str] = []
        matched_norms: List[str] = []
        unmatched_segments: List[str] = []
        digit_segments: List[str] = []
        parts.reverse()
        for kind, raw, original in parts:
            if kind == "term" and original:
                matched_terms.append(original)
                matched_norms.append(raw)
            elif kind == "unmatched":
                unmatched_segments.append(raw)
            elif kind == "digit":
                digit_segments.append(raw)
        coverage = matched_len / (matched_len + unmatched_len) if (matched_len + unmatched_len) else 1.0
        return ComponentMatchResult(
            matched_terms,
            matched_norms,
            unmatched_segments,
            round(coverage, 4),
            matched_count,
            has_non_digit,
            digit_segments,
        )

    @staticmethod
    def _better(current, candidate):
        if current is None:
            return candidate
        if candidate is None:
            return current
        if candidate[1] > current[1]:
            return candidate
        if candidate[1] == current[1]:
            if candidate[2] > current[2]:
                return candidate
            if candidate[2] == current[2] and len(candidate[3]) < len(current[3]):
                return candidate
        return current


def top_k_candidates(screen_name: str, vocab_terms: List[str], k: int, threshold: float) -> List[Candidate]:
    scored = [Candidate(term, local_similarity(screen_name, term)) for term in vocab_terms]
    scored.sort(key=lambda c: c.score, reverse=True)
    return [c for c in scored[:k] if c.score >= threshold]


def phrase_candidates(screen_name: str, vocab_terms: List[str], k: int, threshold: float) -> List[Candidate]:
    tokens = [t for t in zenkaku_hankaku_norm(screen_name).split(" ") if t]
    grams = set(tokens)
    for i in range(len(tokens) - 1):
        grams.add(tokens[i] + " " + tokens[i + 1])
    pool: List[Candidate] = []
    for g in grams:
        for vt in vocab_terms:
            s = local_similarity(g, vt)
            if s >= threshold:
                pool.append(Candidate(vt, s))
    best_by_term: Dict[str, float] = {}
    for cand in pool:
        best_by_term[cand.term] = max(best_by_term.get(cand.term, 0.0), cand.score)
    merged = [Candidate(t, sc) for t, sc in best_by_term.items()]
    merged.sort(key=lambda c: c.score, reverse=True)
    return merged[: max(k, 10)]


def simple_proposal(text: str) -> str:
    s = zenkaku_hankaku_norm(text)
    tokens = [t for t in re.split(r"\s+", s) if t]
    if not tokens:
        return "newItem"
    stop_words = {"コード", "番号", "名称名", "名称名称"}
    core = [w for w in tokens[:3] if w not in stop_words] or tokens[:2]
    name = core[0].lower() + "".join(w.capitalize() for w in core[1:])
    return name or "newItem"


def build_llm_payload(
    screen_name: str,
    candidates: List[Candidate],
    cfg: Dict[str, Any],
    term_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    component_result: Optional[ComponentMatchResult] = None,
) -> Dict[str, Any]:
    cand_payload: List[Dict[str, Any]] = []
    component_terms = set(component_result.matched_terms) if component_result else set()
    for c in candidates:
        item: Dict[str, Any] = {"term": c.term, "local_score": round(c.score, 4)}
        if component_terms:
            item["from_component"] = c.term in component_terms
        if term_meta and c.term in term_meta:
            meta = term_meta[c.term]
            phys = meta.get("_phys_abbr") or meta.get("_phys")
            if phys:
                item["physical_name"] = phys
        cand_payload.append(item)
    component_payload: Optional[Dict[str, Any]] = None
    if component_result:
        hint = "no_component_hit"
        if component_result.matched_terms:
            if not component_result.unmatched_segments and (
                component_result.has_non_digit or not component_result.digit_segments
            ):
                hint = "component_full"
            elif not component_result.has_non_digit and component_result.digit_segments:
                hint = "digits_only"
            else:
                hint = "component_partial"
        component_payload = {
            "normalized_screen": zenkaku_hankaku_norm(screen_name),
            "component_tokens": component_result.matched_terms,
            "component_tokens_norm": component_result.matched_norms,
            "unmatched_fragments": component_result.unmatched_segments,
            "digit_segments": component_result.digit_segments,
            "coverage_ratio": round(component_result.coverage_ratio, 4),
            "has_non_digit": component_result.has_non_digit,
            "expected_match_hint": hint,
        }
    return {
        "model": cfg["OPENAI_MODEL"],
        "max_tokens": cfg["MAX_TOKENS"],
        "temperature": cfg["TEMPERATURE"],
        "top_p": cfg["TOP_P"],
        "presence_penalty": cfg["PRESENCE_PENALTY"],
        "frequency_penalty": cfg["FREQUENCY_PENALTY"],
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": LLM_SYSTEM},
            {
                "role": "user",
                "content": Template(LLM_USER_TEMPLATE).safe_substitute(
                    screen_name=screen_name,
                    component_json=(
                        json.dumps(component_payload, ensure_ascii=False, indent=2)
                        if component_payload is not None
                        else "null"
                    ),
                    candidates_json=json.dumps(cand_payload, ensure_ascii=False, indent=2),
                ),
            },
        ],
    }


def fallback_reason(
    screen_name: str,
    candidates: List[Candidate],
    term_meta: Optional[Dict[str, Dict[str, Any]]],
    error_message: Optional[str] = None,
) -> Dict[str, Any]:
    message = f"APIエラー: {error_message}" if error_message else "API応答なし"
    top_term = candidates[0].term if candidates else None
    if top_term:
        message = f"{message} / ローカル候補: {top_term}"
    return {
        "match_type": "一致なし",
        "matched_term": None,
        "reason": message,
        "proposed_name": simple_proposal(screen_name if not top_term else top_term),
    }


def process(
    dir_path: Path,
    screen_col: Optional[str],
    vocab_col: Optional[str],
    cfg: Dict[str, Any],
    progress_callback=None,
) -> pd.DataFrame:
    df_screen, df_vocab = load_screen_and_vocab(dir_path, cfg, screen_col, vocab_col)
    total_items = len(df_screen)
    processed_count = 0
    vocab_terms = df_vocab["_term"].astype(str).tolist()
    term_meta = (
        df_vocab[["_term", "_phys", "_phys_abbr", "_no"]]
        .drop_duplicates("_term")
        .set_index("_term")
        .to_dict(orient="index")
    )
    norm_to_term = (
        df_vocab[["__term_norm", "_term"]]
        .drop_duplicates("__term_norm")
        .set_index("__term_norm")["_term"]
        .to_dict()
    )
    component_matcher = ComponentMatcher(norm_to_term)
    rows: List[Dict[str, Any]] = []
    api_client = ApiClient(cfg)
    max_concurrent_api = cfg.get("MAX_CONCURRENT_API", 3)
    api_semaphore = threading.Semaphore(max_concurrent_api)

    def _format_no(no_value) -> Optional[int]:
        if no_value is None:
            return None
        try:
            return int(float(no_value))
        except (ValueError, TypeError):
            return None

    def meta_of(term: Optional[str]) -> Dict[str, Any]:
        if not term:
            return {"no": None, "phys": None, "phys_abbr": None}
        m = term_meta.get(str(term)) or {}
        return {
            "no": _format_no(m.get("_no")),
            "phys": m.get("_phys"),
            "phys_abbr": m.get("_phys_abbr"),
        }

    def worker(screen_name: str, src_file: str, src_sheet: Optional[str]) -> Dict[str, Any]:
        normalized = zenkaku_hankaku_norm(screen_name)
        exact_term = norm_to_term.get(normalized)
        if exact_term:
            m = term_meta.get(exact_term) or {}
            proposed = m.get("_phys_abbr") or m.get("_phys") or simple_proposal(exact_term)
            return {
                "source_file": src_file,
                "source_sheet": src_sheet,
                "screen_item": screen_name,
                "match_type": "完全一致",
                "matched_term": exact_term,
                "matched_term_no": _format_no(m.get("_no")),
                "matched_term_phys": (m.get("_phys_abbr") or m.get("_phys")),
                "matched_terms": None,
                "matched_terms_nos": None,
                "matched_terms_phys": None,
                "local_top_term": exact_term,
                "local_top_term_no": _format_no(m.get("_no")),
                "local_top_term_phys": (m.get("_phys_abbr") or m.get("_phys")),
                "local_top_score": 1.0,
                "coverage_ratio": 1.0,
                "proposed_name": proposed,
                "reason": "正規化完全一致（LLM未呼び出し）",
                "unmatched_terms": None,
                "unmatched_note": None,
            }
        broad_candidates = phrase_candidates(screen_name, vocab_terms, max(cfg["TOP_K"], 6), cfg["FUZZY_THRESHOLD"])
        direct_candidates = top_k_candidates(screen_name, vocab_terms, cfg["TOP_K"], cfg["FUZZY_THRESHOLD"])
        comp_result = component_matcher.analyze(screen_name)
        component_candidates: List[Candidate] = []
        if comp_result.matched_terms:
            for term in comp_result.matched_terms:
                score = local_similarity(screen_name, term)
                component_candidates.append(Candidate(term, max(score, cfg["FUZZY_THRESHOLD"])))
        component_terms_set = {c.term for c in component_candidates}
        other_candidates_scores: Dict[str, float] = {}
        for c in broad_candidates + direct_candidates:
            if c.term not in component_terms_set:
                other_candidates_scores[c.term] = max(other_candidates_scores.get(c.term, 0.0), c.score)
        other_candidates = [Candidate(t, s) for t, s in other_candidates_scores.items()]
        other_candidates.sort(key=lambda c: c.score, reverse=True)
        max_other = max(cfg["TOP_K"], 10) - len(component_candidates)
        merged = component_candidates + other_candidates[:max_other]
        merged.sort(key=lambda c: c.score, reverse=True)
        if comp_result.matched_terms and not comp_result.unmatched_segments:
            seg_terms_meta = [term_meta.get(t, {}) for t in comp_result.matched_terms]
            phys_list = [
                (m.get("_phys_abbr") or m.get("_phys"))
                for m in seg_terms_meta
                if (m.get("_phys_abbr") or m.get("_phys"))
            ]
            if phys_list:
                proposed = phys_list[0][:1].lower() + phys_list[0][1:] + "".join(phys_list[1:])
            else:
                proposed = simple_proposal(screen_name)
            return {
                "source_file": src_file,
                "source_sheet": src_sheet,
                "screen_item": screen_name,
                "match_type": "完全一致（部品ごと）",
                "matched_term": None,
                "matched_term_no": None,
                "matched_term_phys": None,
                "matched_terms": ", ".join(comp_result.matched_terms),
                "matched_terms_nos": ", ".join(
                    [
                        str(_format_no(term_meta.get(t, {}).get("_no")))
                        for t in comp_result.matched_terms
                        if _format_no(term_meta.get(t, {}).get("_no")) is not None
                    ]
                )
                or None,
                "matched_terms_phys": ", ".join(
                    [
                        str((term_meta.get(t, {}).get("_phys_abbr") or term_meta.get(t, {}).get("_phys")))
                        for t in comp_result.matched_terms
                        if (term_meta.get(t, {}).get("_phys_abbr") or term_meta.get(t, {}).get("_phys"))
                    ]
                )
                or None,
                "local_top_term": comp_result.matched_terms[0] if comp_result.matched_terms else None,
                "local_top_term_no": _format_no(term_meta.get(comp_result.matched_terms[0], {}).get("_no"))
                if comp_result.matched_terms
                else None,
                "local_top_term_phys": (
                    term_meta.get(comp_result.matched_terms[0], {}).get("_phys_abbr")
                    or term_meta.get(comp_result.matched_terms[0], {}).get("_phys")
                )
                if comp_result.matched_terms
                else None,
                "local_top_score": 1.0,
                "coverage_ratio": 1.0,
                "proposed_name": proposed,
                "reason": "辞書最長一致で部品完全一致のため LLM 未使用",
                "unmatched_terms": None,
                "unmatched_note": None,
            }
        if merged and merged[0].score >= HARD_EXACT_SCORE:
            top = merged[0]
            norm_screen = zenkaku_hankaku_norm(screen_name)
            norm_term = zenkaku_hankaku_norm(top.term)
            if norm_screen == norm_term:
                m = term_meta.get(top.term) or {}
                proposed = m.get("_phys_abbr") or m.get("_phys") or simple_proposal(top.term)
                return {
                    "source_file": src_file,
                    "source_sheet": src_sheet,
                    "screen_item": screen_name,
                    "match_type": "完全一致",
                    "matched_term": top.term,
                    "matched_term_no": _format_no(m.get("_no")),
                    "matched_term_phys": (m.get("_phys_abbr") or m.get("_phys")),
                    "matched_terms": None,
                    "matched_terms_nos": None,
                    "matched_terms_phys": None,
                    "local_top_term": top.term,
                    "local_top_term_no": _format_no(m.get("_no")),
                    "local_top_term_phys": (m.get("_phys_abbr") or m.get("_phys")),
                    "local_top_score": top.score,
                    "coverage_ratio": 1.0,
                    "proposed_name": proposed,
                    "reason": f"ローカル完全一致（score={top.score:.2f}、LLM未呼び出し）",
                    "unmatched_terms": None,
                    "unmatched_note": None,
                }
        payload = build_llm_payload(
            screen_name,
            merged,
            cfg,
            term_meta=term_meta,
            component_result=comp_result,
        )
        try:
            llm = call_llm(payload, cfg, screen_name, api_client, api_semaphore)
        except Exception as exc:
            llm = fallback_reason(screen_name, merged, term_meta, str(exc))
        llm_match_type = llm.get("match_type")
        llm_matched_term = llm.get("matched_term")
        llm_matched_terms = llm.get("matched_terms") or []
        if comp_result.matched_terms:
            matched_terms = comp_result.matched_terms
            matched_term_val: Optional[str] = None
            coverage_ratio = comp_result.coverage_ratio
        else:
            matched_terms = llm_matched_terms
            matched_term_val = llm_matched_term
            cov_raw = llm.get("coverage_ratio")
            try:
                coverage_ratio = float(cov_raw) if cov_raw is not None else None
            except Exception:
                coverage_ratio = None
        mt_meta = meta_of(matched_term_val)
        mts_metas = [meta_of(t) for t in matched_terms]
        llm_unmatched_terms: List[str] = []
        llm_unmatched_notes: List[str] = []
        try:
            llm_unmatched_terms = llm.get("unmatched_terms") or []
            llm_unmatched_notes = llm.get("unmatched_notes") or []
        except Exception:
            pass
        unmatched_terms = None
        unmatched_note = None
        if comp_result.matched_terms:
            if comp_result.unmatched_segments:
                unmatched_terms = ", ".join(comp_result.unmatched_segments)
                if llm_unmatched_notes:
                    unmatched_note = " / ".join(llm_unmatched_notes)
        else:
            if llm_unmatched_terms and llm_unmatched_notes:
                unmatched_terms = ", ".join(llm_unmatched_terms)
                unmatched_note = " / ".join(llm_unmatched_notes)
        if unmatched_note is None and llm_unmatched_notes:
            unmatched_note = " / ".join(llm_unmatched_notes)
        final_match_type = llm_match_type or "一致なし"
        if comp_result.matched_terms:
            final_match_type = "一部一致" if comp_result.unmatched_segments else "完全一致（部品ごと）"
        return {
            "source_file": src_file,
            "source_sheet": src_sheet,
            "screen_item": screen_name,
            "match_type": final_match_type,
            "matched_term": matched_term_val,
            "matched_term_no": mt_meta.get("no"),
            "matched_term_phys": (mt_meta.get("phys_abbr") or mt_meta.get("phys")),
            "matched_terms": ", ".join(matched_terms) or None,
            "matched_terms_nos": ", ".join(
                [str(m.get("no")) for m in mts_metas if m.get("no") is not None]
            )
            or None,
            "matched_terms_phys": ", ".join(
                [
                    str((m.get("phys_abbr") or m.get("phys")))
                    for m in mts_metas
                    if (m.get("phys_abbr") or m.get("phys"))
                ]
            )
            or None,
            "local_top_term": (merged[0].term if merged else None),
            "local_top_term_no": _format_no((term_meta.get(merged[0].term) or {}).get("_no")) if merged else None,
            "local_top_term_phys": (
                (term_meta.get(merged[0].term) or {}).get("_phys_abbr")
                or (term_meta.get(merged[0].term) or {}).get("_phys")
            )
            if merged
            else None,
            "local_top_score": (merged[0].score if merged else None),
            "coverage_ratio": coverage_ratio,
            "reason": llm.get("reason"),
            "proposed_name": llm.get("proposed_name"),
            "unmatched_terms": unmatched_terms,
            "unmatched_note": unmatched_note,
        }

    items = df_screen[["_screen", "_src_file", "_src_sheet"]].astype(str).values.tolist()
    max_workers = min(cfg["MAX_WORKERS"], max(1, len(items)))
    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(worker, it[0], it[1], it[2]) for it in items]
        for fut in cf.as_completed(futures):
            try:
                row = fut.result()
                rows.append(row)
                processed_count += 1
                pct = processed_count * 100 / total_items if total_items else 0
                if processed_count % 10 == 0 or processed_count == total_items:
                    print(f"[INFO] {processed_count}/{total_items} 件処理済み ({pct:.1f}%)")
                if progress_callback:
                    progress_callback(processed_count, total_items)
            except Exception as e:
                rows.append(
                    {
                        "source_file": "<error>",
                        "source_sheet": "-",
                        "screen_item": "-",
                        "match_type": "一致なし",
                        "reason": f"worker error: {e}",
                        "proposed_name": None,
                    }
                )
                processed_count += 1
                pct = processed_count * 100 / total_items if total_items else 0
                if processed_count % 10 == 0 or processed_count == total_items:
                    print(f"[INFO] {processed_count}/{total_items} 件処理済み ({pct:.1f}%)")
                if progress_callback:
                    progress_callback(processed_count, total_items)
    return pd.DataFrame(rows).reset_index(drop=True)


def load_screen_and_vocab(
    dir_path: Path,
    cfg: Dict[str, Any],
    screen_col: Optional[str],
    vocab_col: Optional[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    screen_files = list(dir_path.glob(cfg["SCREEN_GLOB"]))
    vocab_files = list(dir_path.glob(cfg["VOCAB_GLOB"]))
    if not screen_files:
        raise FileNotFoundError(f"画面項目定義ファイルが見つかりません: {cfg['SCREEN_GLOB']}")
    if not vocab_files:
        raise FileNotFoundError(f"単語帳ファイルが見つかりません: {cfg['VOCAB_GLOB']}")
    screen_df_list = []
    for f in screen_files:
        sheets = cfg["SCREEN_SHEET"].split(",") if cfg.get("SCREEN_SHEET") else [None]
        screen_df_list.extend(load_excel_sheets(f, sheets, screen_col or cfg["SCREEN_COL"], "_screen"))
    vocab_df_list = []
    vocab_col_name = vocab_col or cfg["VOCAB_TERM_COL"]
    for f in vocab_files:
        sheets = cfg["VOCAB_SHEET"].split(",") if cfg.get("VOCAB_SHEET") else [None]
        vocab_df_list.extend(load_excel_sheets(f, sheets, vocab_col_name, "_term"))
    if not screen_df_list:
        raise ValueError("画面項目定義が読み込めませんでした。")
    if not vocab_df_list:
        raise ValueError("単語帳が読み込めませんでした。")
    screen_df = pd.concat(screen_df_list, ignore_index=True)
    vocab_df = pd.concat(vocab_df_list, ignore_index=True)
    vocab_df["__term_norm"] = vocab_df["_term"].astype(str).map(zenkaku_hankaku_norm)
    return screen_df, vocab_df


def load_excel_sheets(path: Path, sheets: List[Optional[str]], col_name: str, alias: str) -> List[pd.DataFrame]:
    collected: List[pd.DataFrame] = []
    xls = pd.ExcelFile(path)
    target_sheets = xls.sheet_names if sheets == ["*"] else sheets
    for sheet in target_sheets:
        df = xls.parse(sheet_name=sheet) if sheet else xls.parse()
        if col_name not in df.columns:
            continue
        df = df[[col_name]].copy()
        df.columns = [alias]
        df["_src_file"] = path.name
        df["_src_sheet"] = sheet or "<default>"
        collected.append(df)
    return collected


def save_outputs(df: pd.DataFrame, cfg: Dict[str, Any]) -> None:
    out_dir = Path(cfg["OUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "match_result.xlsx"
    ordered_cols = [
        "source_file",
        "source_sheet",
        "screen_item",
        "match_type",
        "matched_term",
        "matched_term_no",
        "matched_term_phys",
        "matched_terms",
        "matched_terms_nos",
        "matched_terms_phys",
        "proposed_name",
        "coverage_ratio",
        "reason",
        "local_top_term",
        "local_top_term_no",
        "local_top_term_phys",
        "local_top_score",
        "unmatched_terms",
        "unmatched_note",
    ]
    for c in ordered_cols:
        if c not in df.columns:
            df[c] = None

    def _warn(row):
        mt = row.get("match_type") or ""
        if mt == "一部一致":
            return "⚠️ 部分一致：内容を確認の上、必要に応じて単語帳を追加・修正してください"
        if mt == "一致なし":
            return "⚠️ 一致なし：単語帳への追加を検討してください"
        return ""

    df["注意事項"] = df.apply(_warn, axis=1)
    df = df[ordered_cols + ["注意事項"]]
    header_map = {
        "source_file": ("【元情報】", "ファイル名"),
        "source_sheet": ("【元情報】", "シート名"),
        "screen_item": ("【対象項目】", "項目名"),
        "match_type": ("【結果】", "一致状況"),
        "matched_term": ("【一致した単語】", "論理名"),
        "matched_term_no": ("【一致した単語】", "列番号"),
        "matched_term_phys": ("【一致した単語】", "物理名"),
        "matched_terms": ("【複数一致した単語群】", "論理名"),
        "matched_terms_nos": ("【複数一致した単語群】", "列番号"),
        "matched_terms_phys": ("【複数一致した単語群】", "物理名"),
        "proposed_name": ("【提案】", "推奨物理名"),
        "coverage_ratio": ("【判定詳細】", "カバー率"),
        "reason": ("【判定詳細】", "理由"),
        "local_top_term": ("【ローカル候補】", "論理名"),
        "local_top_term_no": ("【ローカル候補】", "列番号"),
        "local_top_term_phys": ("【ローカル候補】", "物理名"),
        "local_top_score": ("【ローカル候補】", "スコア"),
        "注意事項": ("【注意事項】", "注意事項"),
        "unmatched_terms": ("【登録候補】", "未登録語"),
        "unmatched_note": ("【登録候補】", "コメント"),
    }
    df.columns = pd.MultiIndex.from_tuples([header_map[c] for c in df.columns])
    df.columns.names = [None, None]
    status_counts = df[("【結果】", "一致状況")].value_counts(dropna=False).to_dict()
    summary_df = pd.DataFrame(
        [
            {"メトリクス": "完全一致", "件数": status_counts.get("完全一致", 0)},
            {"メトリクス": "一部一致", "件数": status_counts.get("一部一致", 0)},
            {"メトリクス": "一致なし", "件数": status_counts.get("一致なし", 0)},
            {"メトリクス": "合計", "件数": len(df)},
        ]
    )
    group_keys = [("【元情報】", "ファイル名"), ("【元情報】", "シート名")]
    target_col = ("【対象項目】", "項目名")
    by_file_df = (
        df.groupby(group_keys)
        .agg(
            項目数=(target_col, "count"),
            項目名サンプル=(target_col, lambda s: ", ".join(map(str, s.dropna().head(50)))),
        )
        .reset_index()
    )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df.drop(df.index).to_excel(w, sheet_name="結果", index=True)
        df.to_excel(w, sheet_name="結果", startrow=1, header=False, index=True)
        summary_df.to_excel(w, sheet_name="サマリ", index=True)
        by_file_df.to_excel(w, sheet_name="ファイル別サマリ", index=True)
        if len(df) > 0:
            ws = w.sheets["結果"]
            start_row = 3
            end_row = start_row + len(df) - 1
            col_index = {col: i + 1 for i, col in enumerate(df.columns)}
            mt_idx = col_index[("【結果】", "一致状況")]
            si_idx = col_index[("【対象項目】", "項目名")]
            mt_col = get_column_letter(mt_idx)
            si_col = get_column_letter(si_idx)
            fill_yellow = PatternFill(start_color="FFF3B3", end_color="FFF3B3", fill_type="solid")
            fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            rng_mt = f"{mt_col}{start_row}:{mt_col}{end_row}"
            ws.conditional_formatting.add(
                rng_mt,
                get_formula_rule(f'{mt_col}{start_row}="一部一致"', fill_yellow),
            )
            ws.conditional_formatting.add(
                rng_mt,
                get_formula_rule(f'{mt_col}{start_row}="一致なし"', fill_red),
            )
            rng_si = f"{si_col}{start_row}:{si_col}{end_row}"
            ws.conditional_formatting.add(
                rng_si,
                get_formula_rule(f'${mt_col}{start_row}="一部一致"', fill_yellow),
            )
            ws.conditional_formatting.add(
                rng_si,
                get_formula_rule(f'${mt_col}{start_row}="一致なし"', fill_red),
            )
    print(f"保存: {xlsx_path}")


def get_formula_rule(formula: str, fill: PatternFill):
    from openpyxl.formatting.rule import FormulaRule

    return FormulaRule(formula=[formula], fill=fill)


__all__ = [
    "Candidate",
    "ComponentMatchResult",
    "ComponentMatcher",
    "build_llm_payload",
    "fallback_reason",
    "local_similarity",
    "phrase_candidates",
    "process",
    "save_outputs",
    "simple_proposal",
    "top_k_candidates",
    "zenkaku_hankaku_norm",
]
