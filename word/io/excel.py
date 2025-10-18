#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel入出力関連の関数"""

import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

from ..utils import zenkaku_hankaku_norm
from ..settings import HEADER_DETECT, HEADER_SCAN_ROWS


def _pick_sheet_or_fallback(xls: pd.ExcelFile, preferred: Optional[str]) -> str:
    """優先シートが無ければ先頭。NFKCで厳密同名も考慮。"""

    if preferred and preferred in xls.sheet_names:
        return preferred
    if preferred:
        norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
        want = norm(preferred)
        for name in xls.sheet_names:
            if norm(name) == want:
                return name
    return xls.sheet_names[0]


def _pick_matching_sheets(xls: pd.ExcelFile, preferred: Optional[str]) -> List[str]:
    """設定シート名にマッチする全てのシートを返す。ワイルドカード対応。"""

    if not preferred:
        return [xls.sheet_names[0]]
    norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
    # 複数パターンをカンマ区切りで分割
    patterns = [p.strip() for p in preferred.split(",") if p.strip()]
    if not patterns:
        return [xls.sheet_names[0]]
    all_matches = []
    for pattern in patterns:
        want = norm(pattern)
        # ワイルドカード対応（*を正規表現に変換）
        import re as regex_module
        regex_pattern = want.replace('*', '.*')
        for name in xls.sheet_names:
            if regex_module.match(regex_pattern, norm(name)):
                all_matches.append(name)
    # 重複を除去して順序を保持
    result = []
    seen = set()
    for sheet in all_matches:
        if sheet not in seen:
            result.append(sheet)
            seen.add(sheet)
    # マッチするものがなければ先頭シート
    return result if result else [xls.sheet_names[0]]


def _find_column_in_candidates(df: pd.DataFrame, col_candidates: str) -> Optional[str]:
    """カンマ区切りの列名候補から最初に見つかった列名を返す。"""

    candidates = [c.strip() for c in col_candidates.split(",") if c.strip()]
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        # 正規化した列名でも検索
        norm_candidate = zenkaku_hankaku_norm(candidate)
        for col in df.columns:
            if zenkaku_hankaku_norm(str(col)) == norm_candidate:
                return col
    return None


def _detect_header_row(path: Path, sheet_name: str, required_cols: List[str], scan_rows: int) -> int:
    """先頭scan_rows行で必須列がそろう行を**ヘッダ行**とみなす。
    required_colsの各要素はカンマ区切りで複数候補を指定可能。"""

    head_df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows)
    # 各必須列について、いずれかの候補が含まれていればOK
    req_sets = []
    for col_spec in required_cols:
        if col_spec:
            candidates = {zenkaku_hankaku_norm(c.strip()) for c in col_spec.split(",") if c.strip()}
            req_sets.append(candidates)
    for i in range(len(head_df)):
        row_vals = {zenkaku_hankaku_norm(x) for x in head_df.iloc[i].values if str(x) not in {"", "nan", "一致なし"}}
        # すべての必須列グループについて、いずれかの候補が含まれているか確認
        if all(any(cand in row_vals for cand in req_set) for req_set in req_sets):
            return i
    raise KeyError(f"必須列{sorted(required_cols)}を含むヘッダ行が見つかりません: {path.name}/{sheet_name}")


def read_excel_with_header_detection(path: Path, sheet_name: Optional[str], required_cols: List[str],
                                     explicit_header_row_1based: Optional[int] = None,
                                     scan_rows: int = 30) -> Tuple[pd.DataFrame, int]:
    """ヘッダ行が1行目とは限らないExcelに対応。"""

    if explicit_header_row_1based is not None:
        hdr0 = max(0, explicit_header_row_1based - 1)
        return pd.read_excel(path, sheet_name=sheet_name, header=hdr0), hdr0
    if not HEADER_DETECT:
        return pd.read_excel(path, sheet_name=sheet_name), 0
    # ヘッダ行の自動検出
    header_row = _detect_header_row(path, sheet_name, required_cols, scan_rows)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row), header_row


def load_screen_and_vocab(dir_path: Path, cfg: Dict[str, Any],
                          screen_col_override: Optional[str] = None,
                          vocab_col_override: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """単一ディレクトリから『画面項目定義』群と『単語帳』群を収集して前処理。"""

    screen_files = sorted(dir_path.glob(cfg["SCREEN_GLOB"]))
    vocab_files = sorted(dir_path.glob(cfg["VOCAB_GLOB"]))
    if not screen_files:
        raise FileNotFoundError(f"画面項目定義ファイルが見つかりません: {dir_path}/{cfg['SCREEN_GLOB']}")
    if not vocab_files:
        raise FileNotFoundError(f"単語帳ファイルが見つかりません: {dir_path}/{cfg['VOCAB_GLOB']}")
    screen_col = screen_col_override or cfg["SCREEN_COL"]
    term_col = vocab_col_override or cfg["VOCAB_TERM_COL"]
    phys_col = cfg["VOCAB_PHYS_COL"]
    phys_abbr_col = cfg["VOCAB_PHYS_ABBR_COL"]
    no_col = cfg["VOCAB_NO_COL"]
    # --- 画面項目定義を縦結合（複数シート対応）
    screen_frames: List[pd.DataFrame] = []
    for path in screen_files:
        xls = pd.ExcelFile(path)
        matching_sheets = _pick_matching_sheets(xls, cfg["SCREEN_SHEET"]) if cfg["SCREEN_SHEET"] else [xls.sheet_names[0]]
        for sheet_used in matching_sheets:
            try:
                df_s, _ = read_excel_with_header_detection(path, sheet_used, [screen_col], cfg.get("SCREEN_HEADER_ROW"), HEADER_SCAN_ROWS)
                if screen_col not in df_s.columns:
                    print(f"[警告] {path.name}（{sheet_used}）: 必須列 '{screen_col}' が存在しません - スキップ")
                    continue
                df_s["__source_file"], df_s["__source_sheet"] = path.name, sheet_used
                screen_frames.append(df_s)
                print(f"[INFO] 画面項目定義読み込み: {path.name} / {sheet_used} ({len(df_s)}件)")
            except Exception as e:
                print(f"[警告] {path.name}（{sheet_used}）の読み込みエラー: {e} - スキップ")
    # --- 単語帳を縦結合（複数シート対応）
    vocab_frames: List[pd.DataFrame] = []
    for path in vocab_files:
        xls = pd.ExcelFile(path)
        matching_sheets = _pick_matching_sheets(xls, cfg["VOCAB_SHEET"]) if cfg["VOCAB_SHEET"] else [xls.sheet_names[0]]
        for sheet_used in matching_sheets:
            try:
                df_v, _ = read_excel_with_header_detection(path, sheet_used, [term_col, phys_col, no_col], cfg.get("VOCAB_HEADER_ROW"), HEADER_SCAN_ROWS)
                # 複数候補から実際の列名を解決
                actual_term_col = _find_column_in_candidates(df_v, term_col)
                actual_phys_col = _find_column_in_candidates(df_v, phys_col)
                actual_no_col = _find_column_in_candidates(df_v, no_col)
                actual_phys_abbr_col = _find_column_in_candidates(df_v, phys_abbr_col)
                missing = []
                if not actual_term_col:
                    missing.append(term_col)
                if not actual_phys_col:
                    missing.append(phys_col)
                if not actual_no_col:
                    missing.append(no_col)
                if missing:
                    print(f"[警告] {path.name}（{sheet_used}）: 必須列が不足: {missing} - スキップ")
                    continue
                # 列名を統一した名前にリネーム
                df_v = df_v.rename(columns={
                    actual_term_col: "_term",
                    actual_phys_col: "_phys",
                    actual_no_col: "_no"
                })
                if actual_phys_abbr_col:
                    df_v = df_v.rename(columns={actual_phys_abbr_col: "_phys_abbr"})
                else:
                    df_v["_phys_abbr"] = ""  # 任意列は空で補完
                vocab_frames.append(df_v)
                print(f"[INFO] 単語帳読み込み: {path.name} / {sheet_used} ({len(df_v)}件)")
            except Exception as e:
                print(f"[警告] {path.name}（{sheet_used}）の読み込みエラー: {e} - スキップ")
    # データフレームの結合（空チェック付き）
    if not screen_frames:
        raise FileNotFoundError(f"読み込み可能な画面項目定義シートが見つかりません")
    if not vocab_frames:
        raise FileNotFoundError(f"読み込み可能な単語帳シートが見つかりません")
    df_screen = pd.concat(screen_frames, ignore_index=True)
    df_vocab = pd.concat(vocab_frames, ignore_index=True)
    print(f"[INFO] 合計読み込み: 画面項目定義 {len(df_screen)}件, 単語帳 {len(df_vocab)}件")
    # --- 欠損・空行の整理
    df_screen[screen_col] = df_screen[screen_col].fillna("")
    # 単語帳は既にリネーム済みなので "_" 付きの列名を使用
    for c in ["_term", "_phys", "_no", "_phys_abbr"]:
        df_vocab[c] = df_vocab[c].fillna("")
    df_screen = df_screen[df_screen[screen_col].astype(str).str.strip() != ""].copy()
    df_vocab = df_vocab[df_vocab["_term"].astype(str).str.strip() != ""].copy()
    # --- 照合キー追加・列名整理
    df_vocab["__term_norm"] = df_vocab["_term"].astype(str).map(zenkaku_hankaku_norm)
    df_screen = df_screen.rename(columns={screen_col: "_screen", "__source_file": "_src_file", "__source_sheet": "_src_sheet"})
    return df_screen, df_vocab


def save_outputs(df: pd.DataFrame, cfg: Dict[str, Any]) -> None:
    """結果をExcel/CSV/JSONLで保存（MultiIndex対応・条件付き色付け付き）"""

    out_dir = Path(cfg["OUT_DIR"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "match_result.xlsx"

    # 列順定義
    ordered_cols = [
        # 元情報
        "source_file","source_sheet",
        # 照合対象
        "screen_item",
        # 結果
        "match_type",
        # 一致した単語（完全一致）
        "matched_term","matched_term_no","matched_term_phys",
        # 複数一致した単語群（部品ごと）
        "matched_terms","matched_terms_nos","matched_terms_phys",
        # 提案を後ろに移動
        "proposed_name",
        # 判定詳細
        "coverage_ratio","reason",
        # 参考（ローカル候補）
        "local_top_term","local_top_term_no","local_top_term_phys","local_top_score",
        # 登録候補
        "unmatched_terms","unmatched_note",
    ]
    for c in ordered_cols:
        if c not in df.columns:
            df[c] = None
    # 注意事項列
    def _warn(row):
        mt = (row.get("match_type") or "")
        if mt == "一部一致":
            return "⚠️ 部分一致：内容を確認の上、必要に応じて単語帳を追加・修正してください"
        if mt == "一致なし":
            return "⚠️ 一致なし：単語帳への追加を検討してください"
        return ""
    df["注意事項"] = df.apply(_warn, axis=1)
    # 一次元で並べ替え
    df = df[ordered_cols + ["注意事項"]]
    # --- 2) 一括で MultiIndex 化（列数とタプル数を一致させる）
    header_map = {
        # 元情報
        "source_file": ("【元情報】", "ファイル名"),
        "source_sheet": ("【元情報】", "シート名"),
        # 照合対象
        "screen_item": ("【対象項目】", "項目名"),
        # 結果
        "match_type": ("【結果】", "一致状況"),
        # 一致した単語（主要）
        "matched_term": ("【一致した単語】", "論理名"),
        "matched_term_no": ("【一致した単語】", "列番号"),
        "matched_term_phys": ("【一致した単語】", "物理名"),
        # 複数一致した単語群
        "matched_terms": ("【複数一致した単語群】", "論理名"),
        "matched_terms_nos": ("【複数一致した単語群】", "列番号"),
        "matched_terms_phys": ("【複数一致した単語群】", "物理名"),
        # 提案
        "proposed_name": ("【提案】", "推奨物理名"),
        # 判定詳細
        "coverage_ratio": ("【判定詳細】", "カバー率"),
        "reason": ("【判定詳細】", "理由"),
        # ローカル候補
        "local_top_term": ("【ローカル候補】", "論理名"),
        "local_top_term_no": ("【ローカル候補】", "列番号"),
        "local_top_term_phys": ("【ローカル候補】", "物理名"),
        "local_top_score": ("【ローカル候補】", "スコア"),
        # 注意事項
        "注意事項": ("【注意事項】", "注意事項"),
        # 登録候補
        "unmatched_terms": ("【登録候補】", "未登録語"),
        "unmatched_note": ("【登録候補】", "コメント"),
    }
    df.columns = pd.MultiIndex.from_tuples([header_map[c] for c in df.columns])
    df.columns.names = [None, None]  # ← "Length of names must match..." 回避
    # --- 3) サマリ/ファイル別サマリ
    status_counts = df[("【結果】", "一致状況")].value_counts(dropna=False).to_dict()
    summary_df = pd.DataFrame([
        {"メトリクス": "完全一致", "件数": status_counts.get("完全一致", 0)},
        {"メトリクス": "一部一致", "件数": status_counts.get("一部一致", 0)},
        {"メトリクス": "一致なし", "件数": status_counts.get("一致なし", 0)},
        {"メトリクス": "合計", "件数": len(df)},
    ])

    group_keys = [("【元情報】", "ファイル名"),
                  ("【元情報】", "シート名")]
    target_col = ("【対象項目】", "項目名")
    by_file_df = (
        df.groupby(group_keys)
          .agg(
              項目数=(target_col, "count"),
              項目名サンプル=(target_col, lambda s: ", ".join(map(str, s.dropna().head(50))))
          )
          .reset_index()
    )
    # --- 4) Excel へ保存
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        # MultiIndex列はExcel保存時にヘッダーとデータの間に空白行が入る問題があるため、
        # ヘッダーとデータを分けて書き込むことで余分な空白行を回避する。
        # 最初にヘッダーのみ（データなし）を書き込み
        df.drop(df.index).to_excel(w, sheet_name="結果", index=True)
        # 次に実際のデータをヘッダーなしで1行下から追記
        df.to_excel(w, sheet_name="結果", startrow=1, header=False, index=True)
        summary_df.to_excel(w, sheet_name="サマリ", index=True)
        by_file_df.to_excel(w, sheet_name="ファイル別サマリ", index=True)
        # --- 5) 条件付き書式（「一部一致」「一致なし」を色付け）
        if len(df) > 0:
            ws = w.sheets["結果"]
            # 2段見出しなのでデータ開始行は 3 行目
            start_row = 3
            end_row = start_row + len(df) - 1
            # 列インデックス（1始まり）を取得
            col_index = {col: i+1 for i, col in enumerate(df.columns)}
            mt_idx = col_index[("【結果】", "一致状況")]
            si_idx = col_index[("【対象項目】", "項目名")]
            mt_col = get_column_letter(mt_idx)
            si_col = get_column_letter(si_idx)
            # 塗り色
            fill_yellow = PatternFill(start_color="FFF3B3", end_color="FFF3B3", fill_type="solid")
            fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            # 一部一致（match_type 列自体）
            rng_mt = f"{mt_col}{start_row}:{mt_col}{end_row}"
            ws.conditional_formatting.add(
                rng_mt,
                FormulaRule(formula=[f'{mt_col}{start_row}="一部一致"'], fill=fill_yellow)
            )
            # 一致なし（match_type 列自体）
            ws.conditional_formatting.add(
                rng_mt,
                FormulaRule(formula=[f'{mt_col}{start_row}="一致なし"'], fill=fill_red)
            )
            # 画面項目名列も同じ条件で色付け（列は固定、行は相対）
            rng_si = f"{si_col}{start_row}:{si_col}{end_row}"
            ws.conditional_formatting.add(
                rng_si,
                FormulaRule(formula=[f'${mt_col}{start_row}="一部一致"'], fill=fill_yellow)
            )
            ws.conditional_formatting.add(
                rng_si,
                FormulaRule(formula=[f'${mt_col}{start_row}="一致なし"'], fill=fill_red)
            )
    print(f"保存: {xlsx_path}")
