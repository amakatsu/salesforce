#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse
import concurrent.futures as cf
import difflib
import json
import logging
import os
import re
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
import pandas as pd
import requests
from dotenv import load_dotenv
from string import Template

try:
    from .config import get_word_config
except ImportError:  # script execution fallback
    from config import get_word_config

# ====== GUI（任意） ===========================================================
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

# ====== 定数（読みやすさのため一箇所に集約） ===============================
# 「事実上の完全一致」と見なすスコア（difflibは1.0に非常に近づく）
HARD_EXACT_SCORE = 1.0  # 完全一致のみに限定
# LLMフォールバック時に"完全一致"扱いにする安全側の下限
FALLBACK_EXACT_FLOOR = 0.95

# ====== 設定（.envで上書き可） =============================================
DEFAULT_CONFIG: Dict[str, Any] = get_word_config()

# ====== ロガー設定 ============================================================
logger = logging.getLogger("word_tool.api")
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter("[WORD-API] %(asctime)s %(levelname)s %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
logger.setLevel(logging.INFO)
logger.propagate = False

SENSITIVE_HEADER_KEYS = {"authorization", "api-key", "apikey", "x-api-key"}


def _mask_headers(headers: Dict[str, Any]) -> Dict[str, Any]:
    masked = {}
    for key, value in headers.items():
        if key.lower() in SENSITIVE_HEADER_KEYS and value:
            masked[key] = "***"
        else:
            masked[key] = value
    return masked


def _dump_payload(payload: Dict[str, Any]) -> str:
    try:
        return json.dumps(payload, ensure_ascii=False)
    except Exception:
        return str(payload)

# ====== 文字列正規化／類似度 =================================================


def zenkaku_hankaku_norm(text: str) -> str:
    """NFKC正規化 + 小文字化 + 記号/空白の正規化で**照合の土台**を整える。"""

    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text)).lower().strip()
    s = re.sub(r"[\u3000\s]+", " ", s)          # 全角/半角スペースを単一化
    s = re.sub(r"[\-/・·•･,()\[\]_]+", " ", s)    # 区切り記号はスペースへ
    return re.sub(r"\s+", " ", s)


def local_similarity(a: str, b: str) -> float:
    """簡易類似度：完全一致=1.0 / 片包含=0.9 / それ以外はdifflibのratio。"""

    a_n, b_n = zenkaku_hankaku_norm(a), zenkaku_hankaku_norm(b)
    if not a_n or not b_n:
        return 0.0
    if a_n == b_n:
        return 1.0
    if a_n in b_n or b_n in a_n:
        return 0.9
    return difflib.SequenceMatcher(None, a_n, b_n).ratio()
@dataclass
class Candidate:
    term: str
    score: float

@dataclass
class ComponentMatchResult:
    matched_terms: List[str]
    matched_norms: List[str]
    unmatched_segments: List[str]
    coverage_ratio: float
    unmatched_count: int
    has_non_digit: bool
    digit_segments: List[str]

# ====== ユーティリティ ========================================================


def normalize_term_no(value: Any) -> Optional[int]:
    """列番号を int に正規化（文字列や浮動小数を許容）。"""

    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        if isinstance(value, str):
            trimmed = value.strip()
            return int(trimmed) if trimmed.isdigit() else None
        return None



class ComponentMatcher:
    """辞書語彙による簡易分割。数字はカバー率計算から除外。"""

    def __init__(self, norm_to_term: Dict[str, str]):
        trimmed: Dict[str, str] = {}
        for norm_value, original in norm_to_term.items():
            key = norm_value.replace(" ", "")
            if not key:
                continue
            trimmed.setdefault(key, original)
        self.norm_to_term = trimmed
        self.prefix_map: Dict[str, List[str]] = {}
        for norm_value in self.norm_to_term.keys():
            first = norm_value[0]
            self.prefix_map.setdefault(first, []).append(norm_value)
        for values in self.prefix_map.values():
            values.sort(key=len, reverse=True)
        self.nondigit_len: Dict[str, int] = {
            key: sum(1 for ch in key if not ch.isdigit()) for key in self.norm_to_term.keys()
        }
    def analyze(self, screen_name: str) -> ComponentMatchResult:
        normalized = zenkaku_hankaku_norm(screen_name)
        flat = normalized.replace(" ", "")
        if not flat:
            return ComponentMatchResult([], [], [], 1.0, 0, False, [])
        total_non_digit = sum(1 for ch in flat if not ch.isdigit())
        has_non_digit = total_non_digit > 0
        @lru_cache(maxsize=None)
        def walk(idx: int):
            if idx >= len(flat):
                return (0, 0, 0, [])
            best = None
            ch = flat[idx]
            if ch.isdigit():
                j = idx
                while j < len(flat) and flat[j].isdigit():
                    j += 1
                nxt = walk(j)
                if nxt is not None:
                    candidate = (nxt[0], nxt[1], nxt[2], [("digit", flat[idx:j], None)] + nxt[3])
                    best = self._better(best, candidate)
            if ch in self.prefix_map:
                for norm_term in self.prefix_map[ch]:
                    end = idx + len(norm_term)
                    if flat.startswith(norm_term, idx):
                        nxt = walk(end)
                        if nxt is None:
                            continue
                        term_non_digit = self.nondigit_len.get(norm_term, 0)
                        candidate = (
                            nxt[0],
                            nxt[1] + term_non_digit,
                            nxt[2] + 1,
                            [("term", norm_term, self.norm_to_term.get(norm_term, norm_term))] + nxt[3],
                        )
                        best = self._better(best, candidate)
            if not ch.isdigit():
                nxt = walk(idx + 1)
                if nxt is not None:
                    candidate = (nxt[0] + 1, nxt[1], nxt[2], [("unmatched", ch, None)] + nxt[3])
                    best = self._better(best, candidate)
            return best
        result = walk(0)
        if result is None:
            return ComponentMatchResult([], [], [], 0.0, total_non_digit, has_non_digit, [])
        _, matched_non_digit, _, steps = result
        matched_terms: List[str] = []
        matched_norms: List[str] = []
        unmatched_segments: List[str] = []
        digit_segments: List[str] = []
        buffer_unmatched = ""
        for step_type, value, extra in steps:
            if step_type == "term":
                if buffer_unmatched:
                    unmatched_segments.append(buffer_unmatched)
                    buffer_unmatched = ""
                matched_norms.append(value)
                matched_terms.append(extra or value)
            elif step_type == "digit":
                if buffer_unmatched:
                    unmatched_segments.append(buffer_unmatched)
                    buffer_unmatched = ""
                digit_segments.append(value)
            elif step_type == "unmatched":
                buffer_unmatched += value
        if buffer_unmatched:
            unmatched_segments.append(buffer_unmatched)
        coverage = 1.0
        if has_non_digit:
            coverage = matched_non_digit / total_non_digit if total_non_digit else 0.0
            if coverage > 1.0:
                coverage = 1.0
        unmatched_count = sum(len(seg) for seg in unmatched_segments)
        return ComponentMatchResult(
            matched_terms=matched_terms,
            matched_norms=matched_norms,
            unmatched_segments=unmatched_segments,
            coverage_ratio=coverage,
            unmatched_count=unmatched_count,
            has_non_digit=has_non_digit,
            digit_segments=digit_segments,
        )
    @staticmethod
    def _better(current, candidate):
        """重み付きで候補を比較する。未一致が少なく、被覆が高く、語数が少ないものを優先。"""

        if current is None:
            return candidate
        # 未一致文字数が少ない方を優先
        if candidate[0] != current[0]:
            return candidate if candidate[0] < current[0] else current
        # 被覆率（非数字の一致文字数）が多い方を優先
        if candidate[1] != current[1]:
            return candidate if candidate[1] > current[1] else current
        # 同条件なら語数が少ない（=長い語を選択）方を優先
        if candidate[2] != current[2]:
            return candidate if candidate[2] < current[2] else current
        return candidate

# ====== Excel I/O =============================================================


def _int_env(name: str) -> Optional[int]:
    v = os.getenv(name, "")
    try:
        return int(v) if v else None
    except Exception:
        return None
HEADER_DETECT = os.getenv("HEADER_DETECT", "true").lower() != "false"
HEADER_SCAN_ROWS = int(os.getenv("HEADER_SCAN_ROWS", "30"))
SCREEN_HEADER_ROW = _int_env("SCREEN_HEADER_ROW")
VOCAB_HEADER_ROW = _int_env("VOCAB_HEADER_ROW") 


def _pick_sheet_or_fallback(xls: pd.ExcelFile, preferred: Optional[str]) -> str:
    """優先シートが無ければ先頭。NFKCで厳密同名も考慮。"""

    if preferred and preferred in xls.sheet_names:
        return preferred
    if preferred:
        norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
        want = norm(preferred)
        for name in xls.sheet_names:
            if norm(name) == want:
                return name
    return xls.sheet_names[0]


def _pick_matching_sheets(xls: pd.ExcelFile, preferred: Optional[str]) -> List[str]:
    """設定シート名にマッチする全てのシートを返す。ワイルドカード対応。"""

    if not preferred:
        return [xls.sheet_names[0]]
    norm = lambda s: unicodedata.normalize("NFKC", s).strip().lower()
    # 複数パターンをカンマ区切りで分割
    patterns = [p.strip() for p in preferred.split(",") if p.strip()]
    if not patterns:
        return [xls.sheet_names[0]]
    all_matches = []
    for pattern in patterns:
        want = norm(pattern)
        # ワイルドカード対応（*を正規表現に変換）
        import re as regex_module
        regex_pattern = want.replace('*', '.*')
        for name in xls.sheet_names:
            if regex_module.match(regex_pattern, norm(name)):
                all_matches.append(name)
    # 重複を除去して順序を保持
    result = []
    seen = set()
    for sheet in all_matches:
        if sheet not in seen:
            result.append(sheet)
            seen.add(sheet)
    # マッチするものがなければ先頭シート
    return result if result else [xls.sheet_names[0]]


def _find_column_in_candidates(df: pd.DataFrame, col_candidates: str) -> Optional[str]:
    """カンマ区切りの列名候補から最初に見つかった列名を返す。"""

    candidates = [c.strip() for c in col_candidates.split(",") if c.strip()]
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
        # 正規化した列名でも検索
        norm_candidate = zenkaku_hankaku_norm(candidate)
        for col in df.columns:
            if zenkaku_hankaku_norm(str(col)) == norm_candidate:
                return col
    return None


def _detect_header_row(path: Path, sheet_name: str, required_cols: List[str], scan_rows: int) -> int:
    """先頭scan_rows行で必須列がそろう行を**ヘッダ行**とみなす。
    required_colsの各要素はカンマ区切りで複数候補を指定可能。"""

    head_df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows)
    # 各必須列について、いずれかの候補が含まれていればOK
    req_sets = []
    for col_spec in required_cols:
        if col_spec:
            candidates = {zenkaku_hankaku_norm(c.strip()) for c in col_spec.split(",") if c.strip()}
            req_sets.append(candidates)
    for i in range(len(head_df)):
        row_vals = {zenkaku_hankaku_norm(x) for x in head_df.iloc[i].values if str(x) not in {"", "nan", "一致なし"}}
        # すべての必須列グループについて、いずれかの候補が含まれているか確認
        if all(any(cand in row_vals for cand in req_set) for req_set in req_sets):
            return i
    raise KeyError(f"必須列{sorted(required_cols)}を含むヘッダ行が見つかりません: {path.name}/{sheet_name}")


def read_excel_with_header_detection(path: Path, sheet_name: Optional[str], required_cols: List[str],
                                     explicit_header_row_1based: Optional[int] = None,
                                     scan_rows: int = 30) -> Tuple[pd.DataFrame, int]:
    """ヘッダ行が1行目とは限らないExcelに対応。"""

    if explicit_header_row_1based is not None:
        hdr0 = max(0, explicit_header_row_1based - 1)
        return pd.read_excel(path, sheet_name=sheet_name, header=hdr0), hdr0
    if not HEADER_DETECT:
        return pd.read_excel(path, sheet_name=sheet_name), 0
    # ヘッダ行の自動検出
    header_row = _detect_header_row(path, sheet_name, required_cols, scan_rows)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row), header_row


def load_screen_and_vocab(dir_path: Path, cfg: Dict[str, Any],
                          screen_col_override: Optional[str] = None,
                          vocab_col_override: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """単一ディレクトリから『画面項目定義』群と『単語帳』群を収集して前処理。"""

    screen_files = sorted(dir_path.glob(cfg["SCREEN_GLOB"]))
    vocab_files = sorted(dir_path.glob(cfg["VOCAB_GLOB"]))
    if not screen_files:
        raise FileNotFoundError(f"画面項目定義ファイルが見つかりません: {dir_path}/{cfg['SCREEN_GLOB']}")
    if not vocab_files:
        raise FileNotFoundError(f"単語帳ファイルが見つかりません: {dir_path}/{cfg['VOCAB_GLOB']}")
    screen_col = screen_col_override or cfg["SCREEN_COL"]
    term_col = vocab_col_override or cfg["VOCAB_TERM_COL"]
    phys_col = cfg["VOCAB_PHYS_COL"]
    phys_abbr_col = cfg["VOCAB_PHYS_ABBR_COL"]
    no_col = cfg["VOCAB_NO_COL"]
    # --- 画面項目定義を縦結合（複数シート対応）
    screen_frames: List[pd.DataFrame] = []
    for path in screen_files:
        xls = pd.ExcelFile(path)
        matching_sheets = _pick_matching_sheets(xls, cfg["SCREEN_SHEET"]) if cfg["SCREEN_SHEET"] else [xls.sheet_names[0]]
        for sheet_used in matching_sheets:
            try:
                df_s, _ = read_excel_with_header_detection(path, sheet_used, [screen_col], SCREEN_HEADER_ROW, HEADER_SCAN_ROWS)
                if screen_col not in df_s.columns:
                    print(f"[警告] {path.name}（{sheet_used}）: 必須列 '{screen_col}' が存在しません - スキップ")
                    continue
                df_s["__source_file"], df_s["__source_sheet"] = path.name, sheet_used
                screen_frames.append(df_s)
                print(f"[INFO] 画面項目定義読み込み: {path.name} / {sheet_used} ({len(df_s)}件)")
            except Exception as e:
                print(f"[警告] {path.name}（{sheet_used}）の読み込みエラー: {e} - スキップ")
    # --- 単語帳を縦結合（複数シート対応）
    vocab_frames: List[pd.DataFrame] = []
    for path in vocab_files:
        xls = pd.ExcelFile(path)
        matching_sheets = _pick_matching_sheets(xls, cfg["VOCAB_SHEET"]) if cfg["VOCAB_SHEET"] else [xls.sheet_names[0]]
        for sheet_used in matching_sheets:
            try:
                df_v, _ = read_excel_with_header_detection(path, sheet_used, [term_col, phys_col, no_col], VOCAB_HEADER_ROW, HEADER_SCAN_ROWS)
                # 複数候補から実際の列名を解決
                actual_term_col = _find_column_in_candidates(df_v, term_col)
                actual_phys_col = _find_column_in_candidates(df_v, phys_col)
                actual_no_col = _find_column_in_candidates(df_v, no_col)
                actual_phys_abbr_col = _find_column_in_candidates(df_v, phys_abbr_col)
                missing = []
                if not actual_term_col:
                    missing.append(term_col)
                if not actual_phys_col:
                    missing.append(phys_col)
                if not actual_no_col:
                    missing.append(no_col)
                if missing:
                    print(f"[警告] {path.name}（{sheet_used}）: 必須列が不足: {missing} - スキップ")
                    continue
                # 列名を統一した名前にリネーム
                df_v = df_v.rename(columns={
                    actual_term_col: "_term",
                    actual_phys_col: "_phys",
                    actual_no_col: "_no"
                })
                if actual_phys_abbr_col:
                    df_v = df_v.rename(columns={actual_phys_abbr_col: "_phys_abbr"})
                else:
                    df_v["_phys_abbr"] = ""  # 任意列は空で補完
                vocab_frames.append(df_v)
                print(f"[INFO] 単語帳読み込み: {path.name} / {sheet_used} ({len(df_v)}件)")
            except Exception as e:
                print(f"[警告] {path.name}（{sheet_used}）の読み込みエラー: {e} - スキップ")
    # データフレームの結合（空チェック付き）
    if not screen_frames:
        raise FileNotFoundError(f"読み込み可能な画面項目定義シートが見つかりません")
    if not vocab_frames:
        raise FileNotFoundError(f"読み込み可能な単語帳シートが見つかりません")
    df_screen = pd.concat(screen_frames, ignore_index=True)
    df_vocab = pd.concat(vocab_frames, ignore_index=True)
    print(f"[INFO] 合計読み込み: 画面項目定義 {len(df_screen)}件, 単語帳 {len(df_vocab)}件")
    # --- 欠損・空行の整理
    df_screen[screen_col] = df_screen[screen_col].fillna("")
    # 単語帳は既にリネーム済みなので "_" 付きの列名を使用
    for c in ["_term", "_phys", "_no", "_phys_abbr"]:
        df_vocab[c] = df_vocab[c].fillna("")
    df_screen = df_screen[df_screen[screen_col].astype(str).str.strip() != ""].copy()
    df_vocab = df_vocab[df_vocab["_term"].astype(str).str.strip() != ""].copy()
    # --- 照合キー追加・列名整理
    df_vocab["__term_norm"] = df_vocab["_term"].astype(str).map(zenkaku_hankaku_norm)
    df_screen = df_screen.rename(columns={screen_col: "_screen", "__source_file": "_src_file", "__source_sheet": "_src_sheet"})
    return df_screen, df_vocab

# ====== 候補生成（ローカル） ==================================================


def top_k_candidates(screen_name: str, vocab_terms: List[str], k: int, threshold: float) -> List[Candidate]:
    """画面項目 vs 単語帳の**直接照合**で上位k件を返却。"""

    scored = [Candidate(term, local_similarity(screen_name, term)) for term in vocab_terms]
    scored.sort(key=lambda c: c.score, reverse=True)
    if k and k > 0:
        scored = scored[:k]
    return [c for c in scored if c.score >= threshold]


def phrase_candidates(screen_name: str, vocab_terms: List[str], k: int, threshold: float) -> List[Candidate]:
    """複合語対策：unigram/bigram に分解してパーツ単位で候補を拾う。"""

    tokens = [t for t in zenkaku_hankaku_norm(screen_name).split(" ") if t]
    grams = set(tokens)
    for i in range(len(tokens) - 1):
        grams.add(tokens[i] + " " + tokens[i + 1])
    pool: List[Candidate] = []
    for g in grams:
        for vt in vocab_terms:
            s = local_similarity(g, vt)
            if s >= threshold:
                pool.append(Candidate(vt, s))
    # 同一単語は最大スコアを採用
    best_by_term: Dict[str, float] = {}
    for cand in pool:
        best_by_term[cand.term] = max(best_by_term.get(cand.term, 0.0), cand.score)
    merged = [Candidate(t, sc) for t, sc in best_by_term.items()]
    merged.sort(key=lambda c: c.score, reverse=True)
    if k and k > 0:
        return merged[:k]
    return merged



def build_candidate_payload(
    candidates: List[Candidate],
    component_result: Optional[ComponentMatchResult],
    term_meta: Optional[Dict[str, Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    """候補リストを LLM へ渡す辞書形式に整形する。"""

    payload: List[Dict[str, Any]] = []
    component_terms = set(component_result.matched_terms) if component_result else set()
    component_order: Dict[str, int] = {}
    if component_result:
        for order, term in enumerate(component_result.matched_terms, start=1):
            component_order.setdefault(term, order)

    for rank, candidate in enumerate(candidates, start=1):
        item: Dict[str, Any] = {
            "term": candidate.term,
            "normalized_term": zenkaku_hankaku_norm(candidate.term),
            "local_score": round(candidate.score, 4),
            "local_rank": rank,
            "from_component": candidate.term in component_terms,
        }
        if item["from_component"]:
            component_rank = component_order.get(candidate.term)
            if component_rank is not None:
                item["component_rank"] = component_rank
        if term_meta:
            meta = term_meta.get(candidate.term)
            if meta:
                term_no = normalize_term_no(meta.get("_no"))
                if term_no is not None:
                    item["term_no"] = term_no
                phys_abbr = (meta.get("_phys_abbr") or "").strip()
                phys_full = (meta.get("_phys") or "").strip()
                if phys_abbr:
                    item["physical_name"] = phys_abbr
                    if phys_full and phys_full != phys_abbr:
                        item["physical_name_full"] = phys_full
                elif phys_full:
                    item["physical_name"] = phys_full
        payload.append(item)
    return payload


def summarize_matched_terms(
    matched_terms: List[str],
    meta_lookup: Callable[[str], Dict[str, Any]],
) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """一致語とメタ情報を表示用に整形する。"""

    if not matched_terms:
        return None, None, None

    metas = [meta_lookup(term) for term in matched_terms]
    numbers = [
        str(no)
        for no in (normalize_term_no(meta.get("no")) for meta in metas)
        if no is not None
    ]
    phys_names = [
        str(meta.get("phys_abbr") or meta.get("phys") or "")
        for meta in metas
        if (meta.get("phys_abbr") or meta.get("phys"))
    ]
    display = ", ".join(str(term) for term in matched_terms)
    joined_numbers = ", ".join(numbers) if numbers else None
    joined_phys = ", ".join(phys_names) if phys_names else None
    return display or None, joined_numbers, joined_phys


# ====== APIクライアント =======================================================


class ApiClient:
    """OpenAI互換APIへの最小ラッパ（ヘッダ/プロキシ/SSL検証対応）。"""

    def __init__(self, cfg: Dict[str, Any]):
        self.base_url = cfg["OPENAI_BASE_URL"].rstrip("/")
        self.path = cfg["OPENAI_PATH"]
        self.timeout = cfg["TIMEOUT_SEC"]
        self.verify = cfg["VERIFY_SSL"]
        self.session = requests.Session()  # ThreadPool内ではスレッド毎の生成を推奨
        # プロキシ
        proxies: Dict[str, str] = {}
        if cfg.get("HTTP_PROXY"):
            proxies["http"] = cfg["HTTP_PROXY"]
        if cfg.get("HTTPS_PROXY"):
            proxies["https"] = cfg["HTTPS_PROXY"]
        if proxies:
            self.session.proxies.update(proxies)
        # ヘッダ
        headers: Dict[str, str] = {"Content-Type": "application/json"}
        if cfg.get("OPENAI_SEND_AUTH") and cfg.get("OPENAI_API_KEY"):
            headers["Authorization"] = f"Bearer {cfg['OPENAI_API_KEY']}"
        if cfg.get("OPENAI_ORG_ID"):
            headers["OpenAI-Organization"] = cfg["OPENAI_ORG_ID"]
        extra = cfg.get("OPENAI_HEADERS_JSON")
        if extra:
            try:
                headers.update(json.loads(extra))
            except Exception:
                pass
        self.headers = headers
    def post_json(self, body: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{self.base_url}{self.path}"
        logger.info(
            "HTTP POST開始 url=%s model=%s tokens=%s messages=%s headers=%s",
            url,
            body.get("model"),
            body.get("max_completion_tokens"),
            len(body.get("messages", [])),
            _mask_headers(self.headers),
        )
        logger.info("HTTP リクエストボディ:\n%s", _dump_payload(body))
        start = time.time()
        resp = None
        try:
            resp = self.session.post(
                url,
                headers=self.headers,
                json=body,
                timeout=self.timeout,
                verify=self.verify,
            )
            resp.raise_for_status()
            elapsed = time.time() - start
            resp_text = resp.text
            logger.info(
                "HTTP POST完了 status=%s elapsed=%.2fs request_id=%s",
                resp.status_code,
                elapsed,
                resp.headers.get("x-request-id") or resp.headers.get("X-Request-ID") or "-",
            )
            logger.info("HTTP レスポンスボディ:\n%s", resp_text)
            return resp.json()
        except requests.RequestException as exc:
            elapsed = time.time() - start
            status = resp.status_code if resp is not None else getattr(exc.response, "status_code", "n/a")
            logger.error(
                "HTTP POST失敗: status=%s elapsed=%.2fs error=%s",
                status,
                elapsed,
                exc,
            )
            if resp is not None:
                logger.error("HTTP レスポンスボディ(エラー):\n%s", resp.text)
            raise

# ====== LLM呼び出し（プロンプト詳細は割愛） ================================
LLM_SYSTEM = """あなたは業務システム開発における命名規則の専門家です。
画面項目名と単語帳を照合し、lowerCamelCase 形式の物理名を提案することが使命です。
以降の規則は優先度の高い順に記載します。衝突時は上位規則を必ず優先してください。
### 入力コンテキスト
- component_analysis: ローカル分割結果の JSON。
  - component_tokens: 一致済み語
  - unmatched_fragments: 未登録語（数字は含まない）
  - digit_segments: 数字のみの断片
  - expected_match_hint: "component_full" / "component_partial" / "digits_only" / "no_component_hit"
  - coverage_ratio: 分割語ベースの被覆率（0.0〜1.0）
- 候補リストの要素は {term, physical_name, from_component, score?} を含む。
- from_component が true の要素は component_tokens と同一。候補に存在しない語を創作しない。
### 判定ポリシー（component_analysis を最優先）
1) expected_match_hint = "component_full"
   - match_type = 「完全一致（部品ごと）」
   - matched_terms = component_tokens（順序を保持）
   - unmatched_terms は null または []
2) expected_match_hint = "component_partial"
   - match_type = 「一部一致」
   - matched_terms = component_tokens
   - unmatched_terms = unmatched_fragments（数字は含めない）
   - coverage_ratio = component_analysis.coverage_ratio をそのまま採用
   - ★重要: 一部一致判定時の注意点
     * 候補の意味的コンテキストを必ず考慮すること
     * **より長い文字列で一致する候補がある場合、必ずそちらを優先すること（最長一致の原則）**
     * 例: 「重要度」→「重要」(2文字一致) vs 「要」(1文字一致) → 「重要」を優先
     * 例: 「有無フラグ」→「有無」(2文字一致) vs 「有」(1文字一致) or 「無」(1文字一致) → 「有無」を優先
     * 例: 「担当者名」→「担当者」(3文字一致) vs 「担当」(2文字一致) → 「担当者」を優先
     * 例: 「顧客番号」→「顧客番号」(4文字一致) vs 「顧客」(2文字一致) + 「番号」(2文字一致) → 「顧客番号」を優先
     * 部分文字列一致だけでなく、業務的な意味の整合性を確認すること
     * 最長一致の原則: 常に最も長く一致する候補を選択すること
3) expected_match_hint = "digits_only"
   - 数字は照合対象外（unmatched_termsに含めない）
   - 数字以外の部分が全て一致していれば「完全一致（部品ごと）」
   - 例: 「顧客コード1」→「顧客コード」が一致 → 「完全一致（部品ごと）」、proposed_name=「customerCode1」
   - 数字は物理名に必ず含めるが、一致判定には影響しない
4) expected_match_hint = "no_component_hit"
   - ComponentMatcherで一致なし。候補リストから文字列として完全に含まれるものを探す
   - 判定基準: 画面項目名を分解したとき、候補の論理名が **文字列として完全一致** する部分があるか
   - 例: 「顧客担当者名」→「顧客」「担当者」「名」のように分割でき、各部分が候補に存在すれば「一部一致」
   - 例: 「重要度」→「重要」は文字列として含まれるが、「要」は部分一致なので不適切
   - 例: 「有無フラグ」→「有無」は文字列として完全一致するが、「有」「無」への分割は不適切（意味が異なる）
   - スコアは参考程度。文字列の完全一致性を最優先すること
   - 曖昧な場合や、文字列として完全に分割できない場合は「一致なし」を選ぶ
5) component_tokens も unmatched_fragments も空
   - 画面項目が空 or 数字のみ。創作をせず安全側で判定
6) unmatched_terms / unmatched_notes は同じ長さの配列。数字のみ断片は unmatched_terms に含めない
7) coverage_ratio は component_analysis.coverage_ratio を採用。未提供時のみ自分で計算
8) ★重要: 略称の優先
   - 候補に略称（_phys_abbr）が存在する場合は、必ず略称を優先して使用
   - 略称がない場合のみ、正式名称（_phys）を使用
### 物理名命名ルール（厳密）
1) lowerCamelCase を必ず守る（例: shinseiDate）
2) 文字数制限（match_typeによって異なる）:
   - 「完全一致（部品ごと）」: 文字数制限なし
     理由: 候補の物理名を全て連結するため、自然に15文字を超える場合がある
     例: "collection" + "RingiNo" = "collectionRingiNo" (17文字)
   - 「一部一致」: 文字数制限なし
     理由: 候補の物理名を連結+未登録語を補完するため、15文字を超えても可
     例: "customer" + "PersonInCharge" + "Name" = "customerPersonInChargeName" (27文字)
   - 「一致なし」: 推奨 ~8 文字、最大 15 文字
     理由: 完全に創作するため、読みやすさを優先
3) 英語を基調。必要に応じヘボン式ローマ字（例: 稟議 → ringi）
4) 1つの物理名内で英語とローマ字の混在は禁止（組織内の慣用のみ例外）
5) 略語は 9 文字以上で明快さが増す場合のみ
6) ★重要: 候補の physical_name がある語の優先順位
   - 略称（_phys_abbr）が存在する場合は、必ず略称を優先して使用
   - 略称がない場合は、正式名称（_phys）を使用
   - 候補にある語の physical_name は**必ず優先**して採用
7) ★重要: 数字の扱い
   - 論理名に数字が含まれている場合、物理名にも必ず数字を含める
   - 数字は物理名の適切な位置に配置する（例: 「顧客コード1」→「customerCode1」）
   - 数字間のハイフン（-）は「and」に置き換える（例: 「1-1」→「1and1」、「項目2-3」→「item2and3」）
   - ただし、数字は一致判定の対象外（候補の不足判定unmatched_termsには含めない）
   - 数字以外の部分が全て一致していれば「完全一致（部品ごと）」として扱う
8) 余計な語を追加しない。画面項目の意味を最少構成で表現
### 表記揺れの扱い（参考）
- コード = CD = code / 番号 = No = number / 名称 = 名 = name / フラグ = flag
### 物理名生成の優先順位（最重要）
1) 候補 physical_name を最優先
2) 不足語のみ命名ルールで補完（英語優先、次点でローマ字）
3) 最小限・短く・明瞭に（8〜15 文字目安）
### 思考プロセス（内部で行う。出力には含めない）
1) 画面項目名を語に分解
2) 各語が候補にあるか確認
3) 一致判定（完全一致/一部一致/一致なし）
4) 候補にある語の physical_name を採用。ない語のみ命名ルールで補完
5) lowerCamelCase で結合
### 出力（JSONのみ）
- **説明文や前置きは禁止。JSON だけを返すこと。**
- proposed_name は match_type に関わらず lowerCamelCase の非空文字列で必ず返す。
- proposed_name は上記の命名ルール 1)〜7) をすべて満たすこと。違反が判明した場合は修正してから出力する。
### reason フィールドの記述ルール（最重要）
reason フィールドには、以下の内容を丁寧に、分かりやすく記述すること：

1) **一致した語の説明**（完全一致・一部一致の場合）
   - どの語が単語帳に登録されていたか
   - それぞれの語に対応する物理名は何か

2) **未登録語の物理名提案とその理由**（一部一致・一致なしの場合）
   - 未登録語に対してどのような物理名を提案したか
   - なぜその物理名を選んだのか（英語/ローマ字の選択理由、業務的な意味など）
   - 例: 「「担当者」は未登録のため「personInCharge」を提案しました。業務上、人を表す語には英語の person を使用し、担当の意味を明確にするため InCharge を付加しました」

3) **数字の扱い**（数字が含まれる場合）
   - 数字部分はどう扱ったか

4) **注意事項**（必要な場合）
   - 単語帳への追加を推奨する語があればその理由

**重要**: reasonフィールドにタグ（[component_full]など）は不要です。自然な日本語で、分かりやすく記述してください。

### unmatched_notes フィールドの記述ルール
unmatched_notes フィールドには、unmatched_terms の各要素について以下を記述：
- その語に対してどのような物理名を提案したか
- なぜその物理名を選んだのか（業務的な意味、英語/ローマ字の選択理由など）
- 例: 「業務担当者を表す英語表現として personInCharge を採用。person で人を表し、inCharge で担当の意味を明確化」

### 出力スキーマ（検証用。必ず満たすこと）
{
  "type": "object",
  "required": ["match_type","matched_terms","unmatched_terms","unmatched_notes","coverage_ratio","reason","proposed_name"],
  "properties": {
    "match_type": { "type": "string", "enum": ["完全一致（部品ごと）","一部一致","一致なし"] },
    "matched_terms": { "type": "array", "items": { "type": "string", "minLength": 1 } },
    "unmatched_terms": { "type": ["array","null"], "items": { "type": "string", "minLength": 1 } },
    "unmatched_notes": { "type": ["array","null"], "items": { "type": "string", "minLength": 10 } },
    "coverage_ratio": { "type": "number", "minimum": 0, "maximum": 1 },
    "reason": { "type": "string", "minLength": 20 },
    "proposed_name": {
      "type": "string",
      "minLength": 1,
      "pattern": "^[a-z]+(?:[A-Z][a-z0-9]*)*$"
    }
  },
  "additionalProperties": false
}
注意: proposed_name の maxLength 制約は match_type によって異なる
- 「完全一致（部品ごと）」「一部一致」: 制限なし（候補物理名の連結を優先）
- 「一致なし」: 15文字推奨（ただしスキーマ上は制限なし）
### 事前セルフチェック（出力直前に内部で確認）
- proposed_name が lowerCamelCase である
- proposed_name を null や空文字にしない
- 「一致なし」の場合は proposed_name を 15 文字以内に収めることを推奨（必須ではない）
- 命名ルール 1)〜7) に全て従っているか最終確認する（違反があれば再検討）
- reason は分かりやすく丁寧に記述されているか（タグなし、20文字以上）
- unmatched_notes は各要素が丁寧に記述されているか（10文字以上）
- coverage_ratio が [0,1] にある
- unmatched_terms と unmatched_notes の長さが一致（双方 null 可）
- expected_match_hint の規則に反していない
- 候補との意味的整合性を確認したか（部分文字列一致だけで判断していないか）
- 曖昧なら **no_match** を選ぶ（創作禁止）

### Few-shot（形式の見本）
例1（完全一致・部品ごと）
Input:
"""
component_analysis: {"component_tokens":["回収","稟議番号"],"unmatched_fragments":[],"digit_segments":[],"expected_match_hint":"component_full","coverage_ratio":1.0}
candidates: [{"term":"回収","physical_name":"collection","from_component":true},
{"term":"稟議番号","physical_name":"ringiNo","from_component":true}]
"""
Output:
{"match_type":"完全一致（部品ごと）","matched_terms":["回収","稟議番号"],"unmatched_terms":null,"unmatched_notes":null,"coverage_ratio":1.0,"reason":"「回収」は単語帳に登録されており物理名「collection」、「稟議番号」は単語帳に登録されており物理名「ringiNo」を使用しました。全ての語が単語帳で充足されています。","proposed_name":"collectionRingiNo"}

例2（一部一致）
Input:
"""
component_analysis: {"component_tokens":["顧客","名"],"unmatched_fragments":["担当者"],"digit_segments":[],"expected_match_hint":"component_partial","coverage_ratio":0.67}
candidates: [{"term":"顧客","physical_name":"customer","from_component":true},
{"term":"名","physical_name":"Name","from_component":true}]
"""
Output:
{"match_type":"一部一致","matched_terms":["顧客","名"],"unmatched_terms":["担当者"],"unmatched_notes":["業務担当者を表す英語表現として personInCharge を採用しました。person で人を表し、inCharge で担当の意味を明確化することで、業務上の役割を正確に表現しています"],"coverage_ratio":0.67,"reason":"「顧客」は単語帳に登録されており物理名「customer」、「名」は単語帳に登録されており物理名「Name」を使用しました。「担当者」は未登録のため、業務担当者を表す英語表現 personInCharge を提案しました。この語は単語帳への追加を推奨します。","proposed_name":"customerPersonInChargeName"}

例3（数字のみ）
Input:
"""
component_analysis: {"component_tokens":[],"unmatched_fragments":[],"digit_segments":["1","1"],"expected_match_hint":"digits_only","coverage_ratio":1.0}
candidates: []
"""
Output:
{"match_type":"完全一致（部品ごと）","matched_terms":[],"unmatched_terms":null,"unmatched_notes":null,"coverage_ratio":1.0,"reason":"画面項目名は数字のみ（1-1）で構成されており、数字は照合対象外ですが、物理名には数字を保持し、ハイフンは「and」に置き換えて「1and1」を提案しました。","proposed_name":"1and1"}

例4（一致なし）
Input:
"""
component_analysis: {"component_tokens":[],"unmatched_fragments":["緊急度"],"digit_segments":[],"expected_match_hint":"no_component_hit","coverage_ratio":0.0}
candidates: []
"""
Output:
{"match_type":"一致なし","matched_terms":null,"unmatched_terms":["緊急度"],"unmatched_notes":["緊急の度合いを表す英語 urgency を採用しました。業務上、優先度や重要度を表す一般的な英語表現として適切です"],"coverage_ratio":0.0,"reason":"「緊急度」は単語帳に登録されていないため、緊急の度合いを表す英語 urgency を提案しました。この語は単語帳への追加を推奨します。","proposed_name":"urgency"}
"""
LLM_USER_TEMPLATE = r"""画面項目名: $screen_name
component_analysis:
$component_json
単語帳候補（local_score降順）:
$candidates_json
---
出力JSON仕様:
{
  "match_type": "完全一致（部品ごと）" | "一部一致" | "一致なし",
  "matched_term": null,
  "matched_terms": string[] | null,
  "reason": string,
  "proposed_name": string,
  "coverage_ratio": number | null,
  "unmatched_terms": string[] | null,
  "unmatched_notes": string[] | null
}
注意事項:
- JSONのみ出力し、前後にテキストを付与しない。
- matched_terms が存在する場合は画面項目内の順序を維持する。
- unmatched_terms と unmatched_notes は同じ長さにする（どちらかが null なら両方 null）。
- reason は自然な日本語で丁寧に記述する（タグ不要、20文字以上）。
  * 一致した語とその物理名を説明
  * 未登録語に対する物理名提案とその理由を詳細に説明
- unmatched_notes は各要素について10文字以上で丁寧に記述する。
  * その語に対する物理名提案
  * なぜその物理名を選んだのかの理由（業務的意味、英語/ローマ字の選択理由など）
- coverage_ratio は component_analysis.coverage_ratio を使い、未提供時のみ自分で算出する。
- 数字だけの差分は unmatched_terms に含めず、理由に明記する。
- 候補との意味的整合性を必ず確認する（最長一致の原則に従い、より長い文字列で一致する候補を優先）。
- 論理名に数字が含まれる場合、物理名にも必ず数字を含める。
  * 例: 「顧客コード1」→「customerCode1」（数字の1を含める）
  * 例: 「申請日2」→「applicationDate2」（数字の2を含める）
- 略称（physical_name_abbr）が存在する場合は、必ず略称を優先して使用する。
---
タスク: 上記の画面項目名に対して、lowerCamelCaseの物理名を提案してください。
処理ステップ:
1. 画面項目名を意味のある語に分解
2. 各語が候補リストに存在するかチェック（意味的整合性も確認）
3. proposed_name生成:
   - 候補にある語 → その physical_name を使用（必須）
   - 候補にない語 → ネーミングルールに従い自分で考案（理由を reason と unmatched_notes に記載）
   - 全てを lowerCamelCase で結合
出力JSON:
{
  "match_type": "完全一致（部品ごと）" | "一部一致" | "一致なし",
  "matched_term": null,
  "matched_terms": string[] | null,
  "reason": string,
  "proposed_name": string,
  "coverage_ratio": number | null,
  "unmatched_terms": string[] | null,
  "unmatched_notes": string[] | null
}
フィールド定義:
- match_type: 「完全一致（部品ごと）」「一部一致」「一致なし」のいずれか
  - 完全一致（部品ごと）: 分解した全ての語が候補で充足（unmatched_terms が空）
  - 一部一致: 一部の語が候補にあり、一部は未登録（unmatched_terms に値あり）
  - 一致なし: どの語も候補にない
- matched_term: 常に null（単一語での完全一致用フィールドは使用しない）
- matched_terms: 完全一致または一部一致時の語リスト（例: ["回収", "稟議番号"]）、一致なしの場合は null
- proposed_name: 必須、lowerCamelCase
  - 「完全一致（部品ごと）」「一部一致」: 文字数制限なし（候補物理名を連結するため）
  - 「一致なし」: 8-15文字推奨（完全に創作するため）
- coverage_ratio: 0.0-1.0、完全一致（部品ごと）=1.0、一致なし=null または 0.0
- unmatched_terms: 候補にない語のリスト（完全一致（部品ごと）の場合は空配列またはnull）
- unmatched_notes: unmatched_terms各要素の説明（10文字以上で丁寧に）
  * その語に対する物理名提案とその理由を記載
  * 英語/ローマ字の選択理由、業務的な意味などを含める
- reason: 判定理由を丁寧に記述（20文字以上）
  * タグ不要、自然な日本語で
  * 一致した語とその物理名、未登録語への提案とその理由を含める
JSON以外の出力は禁止です。必ず上記スキーマそのままのキー構成で 1 行の JSON を返してください。
"""


def build_llm_payload(
    screen_name: str,
    candidates: List[Candidate],
    cfg: Dict[str, Any],
    term_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    component_result: Optional[ComponentMatchResult] = None,
    extra_component_terms: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """LLM呼び出しペイロードを構築（厳密JSON指定）。"""

    cand_payload = build_candidate_payload(candidates, component_result, term_meta)
    component_payload: Optional[Dict[str, Any]] = None
    if component_result:
        hint = "no_component_hit"
        if component_result.matched_terms:
            if not component_result.unmatched_segments and (
                component_result.has_non_digit or not component_result.digit_segments
            ):
                hint = "component_full"
            elif not component_result.has_non_digit and component_result.digit_segments:
                hint = "digits_only"
            else:
                hint = "component_partial"
        tokens_for_llm = extra_component_terms or component_result.matched_terms
        component_payload = {
            "normalized_screen": zenkaku_hankaku_norm(screen_name),
            "component_tokens": tokens_for_llm,
            "component_tokens_norm": component_result.matched_norms,
            "unmatched_fragments": component_result.unmatched_segments,
            "digit_segments": component_result.digit_segments,
            "coverage_ratio": round(component_result.coverage_ratio, 4),
            "has_non_digit": component_result.has_non_digit,
            "expected_match_hint": hint,
        }
        if extra_component_terms and extra_component_terms != component_result.matched_terms:
            component_payload["component_tokens_original"] = component_result.matched_terms
    return {
        "model": cfg["OPENAI_MODEL"],
        "max_completion_tokens": cfg["MAX_TOKENS"],  # GPT-5 Cline Proxy用
        "n": 1,
        "reasoning_effort": "high",
        "verbosity": "high",
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": LLM_SYSTEM},
            {
                "role": "user",
                "content": Template(LLM_USER_TEMPLATE).safe_substitute(
                    screen_name=screen_name,
                    component_json=(
                        json.dumps(component_payload, ensure_ascii=False, indent=2)
                        if component_payload is not None
                        else "null"
                    ),
                    candidates_json=json.dumps(cand_payload, ensure_ascii=False, indent=2),
                ),
            },
        ],
    }


def call_llm(
    screen_name: str,
    candidates: List[Candidate],
    cfg: Dict[str, Any],
    client: Optional[ApiClient] = None,
    api_semaphore: Optional[threading.Semaphore] = None,
    term_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    component_result: Optional[ComponentMatchResult] = None,
    extra_component_terms: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """LLM呼び出し。失敗時はフォールバック。"""

    client = client or ApiClient(cfg)
    payload = build_llm_payload(
        screen_name,
        candidates,
        cfg,
        term_meta=term_meta,
        component_result=component_result,
        extra_component_terms=extra_component_terms,
    )
    # サーバー負荷対策：セマフォで同時実行API数を制限
    if api_semaphore:
        api_semaphore.acquire()
    try:
        for attempt in range(cfg["RETRY"] + 1):
            try:
                data = client.post_json(payload)
                content = data["choices"][0]["message"]["content"]
                result = json.loads(content)
                return result
            except Exception as e:
                # API認証エラーの場合はリトライせずに即座に例外を投げる
                error_msg = str(e)
                print(f"[ERROR] LLM API failed for {screen_name}: {error_msg}", file=sys.stderr)
                if hasattr(e, 'response') and e.response is not None:
                    status_code = e.response.status_code
                    if status_code in [401, 403]:
                        raise Exception(f"API認証エラー (HTTP {status_code}): APIキーまたはユーザIDが無効です") from e
                # その他のエラーの場合はリトライ
                if attempt < cfg["RETRY"]:
                    time.sleep(1.2 * (attempt + 1))  # バックオフ
                    continue
                return fallback_reason(screen_name, candidates, term_meta, error_msg)
    finally:
        if api_semaphore:
            api_semaphore.release() 


def fallback_reason(
    screen_name: str,
    candidates: List[Candidate],
    term_meta: Optional[Dict[str, Dict[str, Any]]] = None,
    error_message: Optional[str] = None,
) -> Dict[str, Any]:
    """ネットワーク障害や破損時の**最小限の結論**。"""

    def _error_prefix() -> str:
        return f"APIエラー: {error_message}" if error_message else "API応答なし"

    if not candidates:
        reason = _error_prefix()
        if not error_message:
            reason = "API不達/候補なし。後日、単語帳の拡充を検討してください。"
        return {
            "match_type": "一致なし",
            "matched_term": None,
            "reason": reason,
            "proposed_name": simple_proposal(screen_name),
        }

    top = candidates[0]
    local_hint = f"ローカル候補: {top.term} (score={top.score:.2f})"
    reason = f"{_error_prefix()} / {local_hint}"
    if not error_message:
        reason = f"API応答なし / {local_hint}"
    return {
        "match_type": "一致なし",
        "matched_term": None,
        "reason": reason,
        "proposed_name": simple_proposal(top.term),
    }


# ====== 簡易物理名生成 =======================================================


def simple_proposal(text: str) -> str:
    """ローワーキャメルの簡易物理名を生成（8〜10文字程度）。"""

    s = zenkaku_hankaku_norm(text)
    tokens = [t for t in re.split(r"\s+", s) if t]
    if not tokens:
        return "newItem"
    # 冗長語の削ぎ落とし
    stop_words = {"コード", "番号", "名称名", "名称名称"}
    core = [w for w in tokens[:3] if w not in stop_words] or tokens[:2]
    # lowerCamelCase化
    name = core[0].lower() + "".join(w.capitalize() for w in core[1:])
    return name or "newItem"


# ====== メイン処理 ============================================================


def process(dir_path: Path, screen_col: Optional[str], vocab_col: Optional[str], cfg: Dict[str, Any], progress_callback=None) -> pd.DataFrame:
    """全体フロー：入力→候補生成→（完全一致なら即決）→LLM判定→集計DataFrame。"""

    df_screen, df_vocab = load_screen_and_vocab(dir_path, cfg, screen_col, vocab_col)
    # 全体件数（進捗ログ用）
    total_items = len(df_screen)
    processed_count = 0
    vocab_terms = df_vocab["_term"].astype(str).tolist()
    term_meta = (
        df_vocab[["_term", "_phys", "_phys_abbr", "_no"]]
        .drop_duplicates("_term").set_index("_term").to_dict(orient="index")
    )
    # 正規化キー→原語の辞書（完全一致ショートサーキット用）
    norm_to_term = (
        df_vocab[["__term_norm", "_term"]]
        .drop_duplicates("__term_norm").set_index("__term_norm")["_term"].to_dict()
    )
    component_matcher = ComponentMatcher(norm_to_term)
    rows: List[Dict[str, Any]] = []
    api_client = ApiClient(cfg)
    # API同時実行数を制限するセマフォ
    max_concurrent_api = cfg.get("MAX_CONCURRENT_API", 3)
    api_semaphore = threading.Semaphore(max_concurrent_api)
    def meta_of(term: Optional[str]) -> Dict[str, Any]:
        if not term:
            return {"no": None, "phys": None, "phys_abbr": None}
        meta = term_meta.get(str(term)) or {}
        return {
            "no": normalize_term_no(meta.get("_no")),
            "phys": meta.get("_phys"),
            "phys_abbr": meta.get("_phys_abbr"),
        }
    def worker(screen_name: str, src_file: str, src_sheet: Optional[str]) -> Dict[str, Any]:
        """1件の画面項目に対する判定ワーカー（スレッドで実行）。"""

        # --- 0) 正規化ベースの完全一致 → LLMスキップ
        normalized = zenkaku_hankaku_norm(screen_name)
        exact_term = norm_to_term.get(normalized)
        if exact_term:
            meta = term_meta.get(exact_term) or {}
            matched_no = normalize_term_no(meta.get("_no"))
            matched_phys = meta.get("_phys_abbr") or meta.get("_phys")
            return {
                "source_file": src_file,
                "source_sheet": src_sheet,
                "screen_item": screen_name,
                "match_type": "完全一致",
                "matched_term": exact_term,
                "matched_term_no": matched_no,
                "matched_term_phys": matched_phys,
                "matched_terms": f"{exact_term}(1.00)",
                "matched_terms_nos": None,
                "matched_terms_phys": None,
                "local_top_term": exact_term,
                "local_top_term_no": matched_no,
                "local_top_term_phys": matched_phys,
                "local_top_score": 1.0,
                "coverage_ratio": 1.0,
                "proposed_name": (matched_phys or simple_proposal(exact_term)),
                "reason": "正規化完全一致（LLM未呼び出し）",
                "unmatched_terms": None,
                "unmatched_note": None,
            }
        # 1) ローカル候補生成
        # 候補はすべて LLM へ渡すため、上限は設けない
        broad_candidates = phrase_candidates(screen_name, vocab_terms, 0, cfg["FUZZY_THRESHOLD"])
        direct_candidates = top_k_candidates(screen_name, vocab_terms, 0, cfg["FUZZY_THRESHOLD"])
        comp_result = component_matcher.analyze(screen_name)

        # ComponentMatcher候補を最優先で保持（LLM判定に必須）
        component_candidates: List[Candidate] = []
        if comp_result.matched_terms:
            for term in comp_result.matched_terms:
                score = local_similarity(screen_name, term)
                component_candidates.append(Candidate(term, max(score, cfg["FUZZY_THRESHOLD"])))

        # Component の部分語（例: 「有無」→「有」「無」）も候補に追加
        # 候補をマージ（重複除外）
        component_terms_set = {c.term for c in component_candidates}
        other_candidates_scores: Dict[str, float] = {}
        for c in broad_candidates + direct_candidates:
            if c.term not in component_terms_set:
                other_candidates_scores[c.term] = max(other_candidates_scores.get(c.term, 0.0), c.score)

        # 画面項目内に含まれる語彙をすべて候補に追加（例: 有・無・有無 を同時に保持）
        normalized_screen_flat = zenkaku_hankaku_norm(screen_name).replace(" ", "")
        for vocab_term in vocab_terms:
            if vocab_term in component_terms_set or vocab_term in other_candidates_scores:
                continue
            term_norm = zenkaku_hankaku_norm(vocab_term).replace(" ", "")
            if term_norm and term_norm in normalized_screen_flat:
                score = local_similarity(screen_name, vocab_term)
                if score >= cfg["FUZZY_THRESHOLD"]:
                    other_candidates_scores[vocab_term] = score

        other_candidates = [Candidate(t, s) for t, s in other_candidates_scores.items()]
        other_candidates.sort(key=lambda c: c.score, reverse=True)

        # 最終候補リスト（全候補をLLMに渡す）
        merged = component_candidates + other_candidates
        merged.sort(key=lambda c: c.score, reverse=True)
        display_terms: List[str] = []
        for cand in merged:
            if cand.term not in display_terms:
                display_terms.append(cand.term)

        # 2) LLM判定（数字処理・略称優先・最長一致の原則を適用）
        llm = call_llm(
            screen_name,
            merged,
            cfg,
            api_client,
            api_semaphore,
            term_meta,
            comp_result,
            extra_component_terms=display_terms,
        )
        llm_match_type = llm.get("match_type")
        llm_matched_term = llm.get("matched_term")
        llm_matched_terms = llm.get("matched_terms") or []
        if comp_result.matched_terms:
            matched_terms = comp_result.matched_terms
            matched_term_val: Optional[str] = None
            coverage_ratio = comp_result.coverage_ratio
        else:
            matched_terms = llm_matched_terms
            matched_term_val = llm_matched_term
            cov_raw = llm.get("coverage_ratio")
            try:
                coverage_ratio = float(cov_raw) if cov_raw is not None else None
            except Exception:
                coverage_ratio = None
        mt_meta = meta_of(matched_term_val)
        matched_terms_display, joined_matched_terms_nos, joined_matched_terms_phys = summarize_matched_terms(display_terms, meta_of)
        llm_unmatched_terms: List[str] = []
        llm_unmatched_notes: List[str] = []
        try:
            llm_unmatched_terms = llm.get("unmatched_terms") or []
            llm_unmatched_notes = llm.get("unmatched_notes") or []
        except Exception:
            pass
        unmatched_terms = None
        unmatched_note = None
        if comp_result.matched_terms:
            if comp_result.unmatched_segments:
                unmatched_terms = ", ".join(comp_result.unmatched_segments)
                if llm_unmatched_notes:
                    unmatched_note = " / ".join(llm_unmatched_notes)
        else:
            if llm_unmatched_terms and llm_unmatched_notes:
                unmatched_terms = ", ".join(llm_unmatched_terms)
                unmatched_note = " / ".join(llm_unmatched_notes)
        if unmatched_note is None and llm_unmatched_notes:
            unmatched_note = " / ".join(llm_unmatched_notes)
        final_match_type = llm_match_type
        if comp_result.matched_terms:
            final_match_type = "一部一致" if comp_result.unmatched_segments else "完全一致（部品ごと）"
        elif final_match_type == "完全一致（部品ごと）":
            final_match_type = "一部一致"

        return {
            "source_file": src_file,
            "source_sheet": src_sheet,
            "screen_item": screen_name,
            "match_type": final_match_type,
            "matched_term": matched_term_val,
            "matched_term_no": mt_meta["no"],
            "matched_term_phys": (mt_meta.get("phys_abbr") or mt_meta.get("phys")),
            "matched_terms": matched_terms_display,
            "matched_terms_nos": joined_matched_terms_nos,
            "matched_terms_phys": joined_matched_terms_phys,
            "local_top_term": (merged[0].term if merged else None),
            "local_top_term_no": normalize_term_no((term_meta.get(merged[0].term) or {}).get("_no")) if merged else None,
            "local_top_term_phys": ((term_meta.get(merged[0].term) or {}).get("_phys_abbr") or (term_meta.get(merged[0].term) or {}).get("_phys")) if merged else None,
            "local_top_score": (merged[0].score if merged else None),
            "coverage_ratio": coverage_ratio,
            "reason": llm.get("reason"),
            "proposed_name": llm.get("proposed_name"),
            "unmatched_terms": unmatched_terms,
            "unmatched_note": unmatched_note,
        }
    # 並列実行（小規模時は自動で縮退）
    items = df_screen[["_screen", "_src_file", "_src_sheet"]].astype(str).values.tolist()
    max_workers = min(cfg["MAX_WORKERS"], max(1, len(items)))
    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(worker, it[0], it[1], it[2]) for it in items]
        for fut in cf.as_completed(futures):
            try:
                row = fut.result()
                rows.append(row)
                processed_count += 1
                # 進捗ログ: 10件ごとに出力
                pct = processed_count * 100 / total_items if total_items else 0
                if processed_count % 10 == 0 or processed_count == total_items:
                    print(f"[INFO] {processed_count}/{total_items} 件処理済み ({pct:.1f}%)")
                # コールバック呼び出し（リアルタイム進捗）
                if progress_callback:
                    progress_callback(processed_count, total_items)
            except Exception as e:
                # ワーカー失敗時も処理を止めない
                error_row = {
                    "source_file": "<error>", "source_sheet": "-", "screen_item": "-",
                    "match_type": "一致なし", "reason": f"worker error: {e}", "proposed_name": None,
                }
                rows.append(error_row)
                processed_count += 1
                # エラー行でも進捗ログ
                pct = processed_count * 100 / total_items if total_items else 0
                if processed_count % 10 == 0 or processed_count == total_items:
                    print(f"[INFO] {processed_count}/{total_items} 件処理済み ({pct:.1f}%)")
                # コールバック呼び出し（リアルタイム進捗）
                if progress_callback:
                    progress_callback(processed_count, total_items)
    return pd.DataFrame(rows).reset_index(drop=True)
    print(f"保存: {xlsx_path}")

# ====== 出力 ================================================================


def save_outputs(df: pd.DataFrame, cfg: Dict[str, Any]) -> None:
    """結果をExcel/CSV/JSONLで保存（MultiIndex対応・条件付き色付け付き）"""

    from pathlib import Path
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule
    out_dir = Path(cfg["OUT_DIR"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "match_result.xlsx"

    # 列順定義
    ordered_cols = [
        # 元情報
        "source_file","source_sheet",
        # 照合対象
        "screen_item",
        # 結果
        "match_type",
        # 一致した単語（完全一致）
        "matched_term","matched_term_no","matched_term_phys",
        # 複数一致した単語群（部品ごと）
        "matched_terms","matched_terms_nos","matched_terms_phys",
        # 提案を後ろに移動
        "proposed_name",
        # 判定詳細
        "coverage_ratio","reason",
        # 参考（ローカル候補）
        "local_top_term","local_top_term_no","local_top_term_phys","local_top_score",
        # 登録候補
        "unmatched_terms","unmatched_note",
    ]
    for c in ordered_cols:
        if c not in df.columns:
            df[c] = None
    # 注意事項列
    def _warn(row):
        mt = (row.get("match_type") or "")
        if mt == "一部一致":
            return "⚠️ 部分一致：内容を確認の上、必要に応じて単語帳を追加・修正してください"
        if mt == "一致なし":
            return "⚠️ 一致なし：単語帳への追加を検討してください"
        return ""
    df["注意事項"] = df.apply(_warn, axis=1)
    # 一次元で並べ替え
    df = df[ordered_cols + ["注意事項"]]
    # --- 2) 一括で MultiIndex 化（列数とタプル数を一致させる）
    header_map = {
        # 元情報
        "source_file": ("【元情報】", "ファイル名"),
        "source_sheet": ("【元情報】", "シート名"),
        # 照合対象
        "screen_item": ("【対象項目】", "項目名"),
        # 結果
        "match_type": ("【結果】", "一致状況"),
        # 一致した単語（主要）
        "matched_term": ("【一致した単語】", "論理名"),
        "matched_term_no": ("【一致した単語】", "列番号"),
        "matched_term_phys": ("【一致した単語】", "物理名"),
        # 複数一致した単語群
        "matched_terms": ("【複数一致した単語群】", "論理名"),
        "matched_terms_nos": ("【複数一致した単語群】", "列番号"),
        "matched_terms_phys": ("【複数一致した単語群】", "物理名"),
        # 提案
        "proposed_name": ("【提案】", "推奨物理名"),
        # 判定詳細
        "coverage_ratio": ("【判定詳細】", "カバー率"),
        "reason": ("【判定詳細】", "理由"),
        # ローカル候補
        "local_top_term": ("【ローカル候補】", "論理名"),
        "local_top_term_no": ("【ローカル候補】", "列番号"),
        "local_top_term_phys": ("【ローカル候補】", "物理名"),
        "local_top_score": ("【ローカル候補】", "スコア"),
        # 注意事項
        "注意事項": ("【注意事項】", "注意事項"),
        # 登録候補
        "unmatched_terms": ("【登録候補】", "未登録語"),
        "unmatched_note": ("【登録候補】", "コメント"),
    }
    df.columns = pd.MultiIndex.from_tuples([header_map[c] for c in df.columns])
    df.columns.names = [None, None]  # ← "Length of names must match..." 回避
    # --- 3) サマリ/ファイル別サマリ
    status_counts = df[("【結果】", "一致状況")].value_counts(dropna=False).to_dict()
    summary_df = pd.DataFrame([
        {"メトリクス": "完全一致", "件数": status_counts.get("完全一致", 0)},
        {"メトリクス": "一部一致", "件数": status_counts.get("一部一致", 0)},
        {"メトリクス": "一致なし", "件数": status_counts.get("一致なし", 0)},
        {"メトリクス": "合計", "件数": len(df)},
    ])

    group_keys = [("【元情報】", "ファイル名"),
                  ("【元情報】", "シート名")]
    target_col = ("【対象項目】", "項目名")
    by_file_df = (
        df.groupby(group_keys)
          .agg(
              項目数=(target_col, "count"),
              項目名サンプル=(target_col, lambda s: ", ".join(map(str, s.dropna().head(50))))
          )
          .reset_index()
    )
    # --- 4) Excel へ保存
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        # MultiIndex列はExcel保存時にヘッダーとデータの間に空白行が入る問題があるため、
        # ヘッダーとデータを分けて書き込むことで余分な空白行を回避する。
        # 最初にヘッダーのみ（データなし）を書き込み
        df.drop(df.index).to_excel(w, sheet_name="結果", index=True)
        # 次に実際のデータをヘッダーなしで1行下から追記
        df.to_excel(w, sheet_name="結果", startrow=1, header=False, index=True)
        summary_df.to_excel(w, sheet_name="サマリ", index=True)
        by_file_df.to_excel(w, sheet_name="ファイル別サマリ", index=True)
        # --- 5) 条件付き書式（「一部一致」「一致なし」を色付け）
        if len(df) > 0:
            ws = w.sheets["結果"]
            # 2段見出しなのでデータ開始行は 3 行目
            start_row = 3
            end_row = start_row + len(df) - 1
            # 列インデックス（1始まり）を取得
            col_index = {col: i+1 for i, col in enumerate(df.columns)}
            mt_idx = col_index[("【結果】", "一致状況")]
            si_idx = col_index[("【対象項目】", "項目名")]
            mt_col = get_column_letter(mt_idx)
            si_col = get_column_letter(si_idx)
            # 塗り色
            fill_yellow = PatternFill(start_color="FFF3B3", end_color="FFF3B3", fill_type="solid")
            fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            # 一部一致（match_type 列自体）
            rng_mt = f"{mt_col}{start_row}:{mt_col}{end_row}"
            ws.conditional_formatting.add(
                rng_mt,
                FormulaRule(formula=[f'{mt_col}{start_row}="一部一致"'], fill=fill_yellow)
            )
            # 一致なし（match_type 列自体）
            ws.conditional_formatting.add(
                rng_mt,
                FormulaRule(formula=[f'{mt_col}{start_row}="一致なし"'], fill=fill_red)
            )
            # 画面項目名列も同じ条件で色付け（列は固定、行は相対）
            rng_si = f"{si_col}{start_row}:{si_col}{end_row}"
            ws.conditional_formatting.add(
                rng_si,
                FormulaRule(formula=[f'${mt_col}{start_row}="一部一致"'], fill=fill_yellow)
            )
            ws.conditional_formatting.add(
                rng_si,
                FormulaRule(formula=[f'${mt_col}{start_row}="一致なし"'], fill=fill_red)
            )
    print(f"保存: {xlsx_path}")

# ====== CLI ================================================================


def app_root() -> Path:
    if getattr(sys, "frozen", False):  # PyInstaller
        return Path(sys.executable).parent
    return Path(__file__).parent


def ask_directory(title: str, initial: Optional[str] = None) -> Optional[str]:
    if not TK_AVAILABLE:
        return None
    try:
        root = tk.Tk(); root.withdraw()
        path = filedialog.askdirectory(title=title, initialdir=initial or str(app_root()))
        root.destroy()
        return path or None
    except Exception:
        return None


def main() -> None:
    load_dotenv()
    # ダブルクリック実行時のカレントずれ防止
    try:
        os.chdir(app_root())
    except Exception:
        pass
    parser = argparse.ArgumentParser(description="Excel 単語照合（LLM 補助）")
    parser.add_argument("--dir", help="入力ディレクトリ（画面項目定義・単語帳）")
    parser.add_argument("--in-dir", help="上と同じ（--dir と同義。後方互換）")
    parser.add_argument("--out-dir", help="出力ディレクトリ（既定: out）")
    parser.add_argument("--screen-col", help="画面項目定義の列名（デフォルト: 項目名称）")
    parser.add_argument("--vocab-col", help="単語帳の列名（デフォルト: 論理名）")
    parser.add_argument("--no-gui", action="store_true", help="ダイアログを使わない（サーバ/CI向け）")
    args = parser.parse_args()
    cfg = DEFAULT_CONFIG.copy()
    if cfg.get("OPENAI_SEND_AUTH") and not cfg.get("OPENAI_API_KEY"):
        print("[警告] OPENAI_SEND_AUTH=true ですが OPENAI_API_KEY が未設定です。Authorization は送信されません。")
    # 入力ディレクトリの解決
    in_dir = args.in_dir or args.dir
    if not in_dir and not args.no_gui:
        in_dir = ask_directory("入力ディレクトリ（画面項目定義・単語帳）を選択")
        if not in_dir:
            try:
                in_dir = input("入力ディレクトリのパスを入力してください: ").strip()
            except Exception:
                in_dir = None
    if not in_dir:
        raise FileNotFoundError("入力ディレクトリが指定されていません。--dir かダイアログで指定してください。")
    # 出力先の解決
    if args.out_dir:
        cfg["OUT_DIR"] = args.out_dir
    else:
        if not args.no_gui:
            chosen = ask_directory("出力ディレクトリ（保存先）を選択（キャンセルで既定の out）")
            if chosen:
                cfg["OUT_DIR"] = chosen
    # 実行
    root_dir = Path(in_dir)
    if not root_dir.exists():
        raise FileNotFoundError(f"ディレクトリが存在しません: {root_dir}")
    df = process(root_dir, args.screen_col, args.vocab_col, cfg)
    save_outputs(df, cfg)
    if TK_AVAILABLE and not args.no_gui:
        try:
            messagebox.showinfo("完了", f"出力が完了しました。保存先: {Path(cfg['OUT_DIR']).resolve()}\\match_result.xlsx")
        except Exception:
            pass

if __name__ == "__main__":
    main()
